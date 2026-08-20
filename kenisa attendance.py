import streamlit as st
import streamlit.components.v1 as components
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import plotly.express as px
from datetime import datetime, timedelta, timezone
import uuid
import json
import random
import string
import jwt
import time
import requests
import hashlib
import hmac
import os
import io
import zipfile
from functools import wraps
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64
from io import BytesIO
import plotly.graph_objects as go
import plotly.io as pio
import re
import html
import qrcode
from PIL import Image, ImageDraw, ImageFont

# Auto-load background image as base64 from image1.jpg
_BG_IMG_PATH = os.path.join(os.path.dirname(__file__), "image1.jpg")
if os.path.exists(_BG_IMG_PATH):
    with open(_BG_IMG_PATH, "rb") as _f:
        BG_IMAGE_BASE64 = base64.b64encode(_f.read()).decode("utf-8")
else:
    BG_IMAGE_BASE64 = ""

# =============================================================================
# الإعدادات العامة والثوابت
# =============================================================================
DEFAULT_JWT_SECRET = "StDemianaChurch2025!Secure#Key"
QUIZ_JWT_SECRET = "StDemianaChurch2025!QuizSecure#Key"
CACHE_TTL_SECONDS = 600
# تحديث مدة الجلسة إلى 24 ساعة
SESSION_TIMEOUT_HOURS = 24
CAIRO_TZ = timezone(timedelta(hours=3), name='Africa/Cairo')

STUDENT_ASSESSMENTS_PAGE = "🏆 المسابقات والاختبارات"
LEGACY_STUDENT_ASSESSMENT_PAGES = {
    "🏆 المسابقات": STUDENT_ASSESSMENTS_PAGE,
    "📊 درجاتي": STUDENT_ASSESSMENTS_PAGE,
    "📋 سجل الامتحانات": STUDENT_ASSESSMENTS_PAGE,
}
ADMIN_ASSESSMENTS_PAGE = "📝 المسابقات والاختبارات"
LEGACY_ADMIN_ASSESSMENTS_PAGE = "📝 إدارة الامتحانات"

# أعمدة سجل التدقيق (AuditLog)
AUDIT_LOG_COLUMNS = [
    "log_id", "timestamp", "username", "user_id", "action", "details",
    "ip_address", "country", "city", "browser", "os", "device_type", "screen_size"
]

# Password hashing
def hash_password(password: str) -> str:
    """Hash a password using SHA-256 with a random salt."""
    salt = os.urandom(32)
    key = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000)
    return salt.hex() + ':' + key.hex()

def verify_password(password: str, stored_hash: str) -> bool:
    """Verify a password against a stored hash (or plaintext fallback)."""
    try:
        salt_hex, key_hex = stored_hash.split(':')
        salt = bytes.fromhex(salt_hex)
        stored_key = bytes.fromhex(key_hex)
        new_key = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000)
        return hmac.compare_digest(new_key, stored_key)
    except (ValueError, AttributeError):
        return stored_hash == password


def get_cairo_now():
    return datetime.now(CAIRO_TZ)


def generate_student_code(db=None, existing_codes=None):
    """
    توليد كود فريد للطالبة.
    الكود: STU + 6 أرقام متسلسلة مع أصفار(left padding)
    مثال: STU000001, STU000002, STU000003
    """
    used_codes = set(existing_codes or [])
    if db is not None:
        try:
            students_df = db.get_students()
            if not students_df.empty and "student_code" in students_df.columns:
                used_codes = set(students_df["student_code"].dropna().astype(str).str.strip().tolist())
        except Exception:
            pass
    
    # استخراج الأرقام من الأكواد الموجودة
    max_num = 0
    for code in used_codes:
        code_str = str(code).strip()
        if code_str.startswith("STU") and len(code_str) == 9:
            try:
                num = int(code_str[3:])
                max_num = max(max_num, num)
            except ValueError:
                pass
    
    # توليد كود جديد بالرقم التالي
    next_num = max_num + 1
    return f"STU{next_num:06d}"


def format_cairo_time(dt):
    if dt is None:
        return "غير متاح"
    return dt.astimezone(CAIRO_TZ).strftime("%Y-%m-%d %I:%M:%S %p")


AR_WEEKDAYS = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
AR_MONTHS = [
    "", "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
    "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر",
]


def parse_cairo_datetime(value):
    """Parse stored datetime string/object into Cairo timezone."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        s = str(value).strip()
        if not s or s.lower() in ("nat", "none", "nan"):
            return None
        try:
            dt = pd.to_datetime(s)
        except Exception:
            return None
    if getattr(dt, "tzinfo", None) is None:
        dt = dt.replace(tzinfo=CAIRO_TZ)
    else:
        dt = dt.astimezone(CAIRO_TZ)
    return dt


def format_arabic_datetime(value, include_time=True):
    """
    Format datetime for Arabic UI display.
    Returns (date_line, time_line) e.g. ('الثلاثاء، 3 يونيو 2026', '03:00 مساءً')
    """
    dt = parse_cairo_datetime(value)
    if dt is None:
        return "—", ""
    weekday = AR_WEEKDAYS[dt.weekday()]
    month = AR_MONTHS[dt.month]
    date_line = f"{weekday}، {dt.day} {month} {dt.year}"
    if not include_time:
        return date_line, ""
    hour = dt.hour
    minute = dt.minute
    period = "صباحًا" if hour < 12 else "مساءً"
    h12 = hour % 12 or 12
    time_line = f"{h12:02d}:{minute:02d} {period}"
    return date_line, time_line


st.set_page_config(
    page_title="نظام- كنيسة الشهيدة دميانة",
    page_icon="⛪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# Telegram & Support
# =============================================================================
def get_telegram_config():
    try:
        return st.secrets["telegram"]["bot_token"], st.secrets["telegram"]["chat_id"]
    except Exception:
        return None, None


def get_support_config():
    try:
        return (
            st.secrets.get("support", {}).get("contact_name", "مسؤول النظام"),
            st.secrets.get("support", {}).get("whatsapp", "")
        )
    except Exception:
        return "مسؤول النظام", ""


# =============================================================================
# Credentials & IDs
# =============================================================================
def get_credentials():
    try:
        return Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
    except Exception as e:
        st.error(f"❌ خطأ في بيانات اعتماد Google: {e}")
        st.stop()


def get_spreadsheet_id():
    try:
        sid = st.secrets["sheets"]["spreadsheet_id"]
        if not sid or not isinstance(sid, str) or sid.strip() == "":
            st.error("❌ معرف جدول البيانات غير صالح.")
            st.stop()
        return sid.strip()
    except Exception as e:
        st.error(f"❌ لم يتم العثور على spreadsheet_id: {e}")
        st.stop()


def get_jwt_secret():
    try:
        return st.secrets["sheets"]["jwt_secret"]
    except Exception:
        return DEFAULT_JWT_SECRET


# =============================================================================
# get sections from users sheet (column "section_id")
# =============================================================================
@st.cache_data(ttl=600)
def get_sections_from_users():
    """
    قراءة قائمة الفصول الفريدة من ورقة Users (عمود section_id).
    ترجع قائمة مرتبة من الفصول غير الفارغة.
    """
    try:
        if 'db_instance' in st.session_state:
            db_local = st.session_state.db_instance
            users = db_local.get_users()
            if not users.empty and "section_id" in users.columns:
                sections = users["section_id"].dropna().unique().tolist()
                sections = [s.strip() for s in sections if s and str(s).strip() and str(s).strip().upper() != "N/A"]
                sections = sorted(set(sections))
                return sections
    except Exception:
        pass
    return []


# =============================================================================
# helpers: age and colors
# =============================================================================
def get_student_age(birthdate):
    if not birthdate:
        return None
    try:
        bd = pd.to_datetime(birthdate)
        now = datetime.now()
        return now.year - bd.year - ((now.month, now.day) < (bd.month, bd.day))
    except Exception:
        return None


STAGE_COLORS = {
    "إعدادي": "#28a745",
    "ثانوي": "#007bff",
    "جامعي": "#6f42c1",
    "KG1": "#fd7e14",
    "KG2": "#e83e8c",
    "الصف الأول": "#20c997",
    "default": "#667eea"
}


def get_stage_color(stage_name):
    for key, color in STAGE_COLORS.items():
        if key in str(stage_name):
            return color
    return STAGE_COLORS["default"]


# =============================================================================
# Premium Design System - Unified CSS
# =============================================================================
def get_design_css():
    """Return the unified premium design system CSS."""
    bg_data_url = f"data:image/jpeg;base64,{BG_IMAGE_BASE64}"
    return f"""
    <style id="premium-design-system">
        @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;500;600;700;800&display=swap');
        @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200');

        /* ===== RTL Setup ===== */
        html, body {{
            direction: rtl !important;
            text-align: right !important;
        }}

        /* ===== Design Tokens ===== */
        :root {{
            --primary: #2563eb;
            --primary-dark: #1d4ed8;
            --primary-light: #dbeafe;
            --secondary: #7c3aed;
            --bg: #f8fafc;
            --card: #ffffff;
            --text: #0f172a;
            --text-muted: #64748b;
            --success: #059669;
            --success-light: #d1fae5;
            --warning: #d97706;
            --warning-light: #fef3c7;
            --danger: #dc2626;
            --danger-light: #fee2e2;
            --border: #e2e8f0;
            --radius: 16px;
            --radius-sm: 10px;
            --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
            --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -2px rgba(0, 0, 0, 0.05);
            --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -4px rgba(0, 0, 0, 0.05);
            --font: 'Cairo', sans-serif;
        }}

        /* ===== Base ===== */
        html, body {{
            font-family: var(--font) !important;
            color: var(--text) !important;
            background: var(--bg) !important;
        }}
        * {{ box-sizing: border-box !important; }}
        .stApp {{
            background: transparent !important;
        }}
        [data-testid="stAppViewContainer"] {{
            background-color: #f8fafc !important;
            background-image: none !important;
            min-height: 100vh !important;
        }}
        /* Hero Banner with Church Image */
        .hero-banner {{
            position: relative !important;
            width: 100% !important;
            height: 240px !important;
            border-radius: 18px !important;
            overflow: hidden !important;
            margin-bottom: 2rem !important;
            background: transparent !important;
        }}
        .hero-banner::before {{
            content: '' !important;
            position: absolute !important;
            inset: 0 !important;
            background-image: url('{bg_data_url}') !important;
            background-position: center center !important;
            background-repeat: no-repeat !important;
            background-size: cover !important;
            z-index: 0 !important;
        }}
        .hero-banner::after {{
            content: '' !important;
            position: absolute !important;
            inset: 0 !important;
            background: rgba(0, 0, 0, 0.65) !important;
            z-index: 1 !important;
        }}
        .hero-content {{
            position: relative !important;
            z-index: 3 !important;
            padding: 2.5rem 3.5rem !important;
            height: 100% !important;
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
            text-align: right !important;
            align-items: flex-start !important;
        }}
        .hero-title {{
            font-size: 2.1rem !important;
            font-weight: 800 !important;
            color: #ffffff !important;
            margin: 0 !important;
            text-shadow: 0 2px 8px rgba(0,0,0,0.3) !important;
            line-height: 1.3 !important;
            font-family: 'Cairo', sans-serif !important;
        }}
        .hero-subtitle {{
            font-size: 1.05rem !important;
            font-weight: 500 !important;
            color: rgba(255, 255, 255, 0.9) !important;
            margin-top: 0.5rem !important;
            text-shadow: 0 1px 4px rgba(0,0,0,0.3) !important;
            font-family: 'Cairo', sans-serif !important;
        }}
        .block-container, section.main {{
            position: relative !important;
            z-index: 1 !important;
        }}
        .content-area {{
            padding: 1.5rem !important;
            max-width: 1400px !important;
            margin: 0 auto !important;
        }}

        /* ===== Typography ===== */
        h1, h2, h3, h4, h5, h6 {{
            font-family: var(--font) !important;
            color: var(--text) !important;
            letter-spacing: -0.01em !important;
        }}
        .main-header {{
            font-size: 1.75rem !important;
            font-weight: 800 !important;
            color: var(--primary) !important;
            text-align: right !important;
            margin-bottom: 1.5rem !important;
            padding: 1rem 1.5rem !important;
            background: var(--card) !important;
            border: 1px solid var(--border) !important;
            border-radius: var(--radius) !important;
            box-shadow: var(--shadow-md) !important;
            border-right: 4px solid var(--primary) !important;
        }}

        /* ===== Cards ===== */
        .glass-card, .user-card, .student-card, .event-card, .profile-stat-card {{
            background: var(--card) !important;
            border-radius: var(--radius) !important;
            border: 1px solid var(--border) !important;
            box-shadow: var(--shadow-sm) !important;
            padding: 1.25rem !important;
            transition: all 0.2s ease !important;
        }}
        .glass-card:hover, .user-card:hover, .student-card:hover, .event-card:hover, .profile-stat-card:hover {{
            box-shadow: var(--shadow-md) !important;
            transform: translateY(-2px) !important;
        }}

        /* ===== Sidebar (RTL - Right Side) ===== */
        section[data-testid="stSidebar"] {{
            background: #ffffff !important;
            border-left: 1px solid var(--border) !important;
            border-right: none !important;
            box-shadow: -2px 0 8px rgba(0,0,0,0.05) !important;
        }}
        
        /* Full screen sidebar overlay for mobile */
        @media (max-width: 768px) {{
            section[data-testid="stSidebar"] {{
                position: fixed !important;
                top: 0 !important;
                right: 0 !important;
                width: 100vw !important;
                height: 100vh !important;
                max-width: 100vw !important;
                max-height: 100vh !important;
                z-index: 999999 !important;
                border-radius: 0 !important;
                overflow-y: auto !important;
                overflow-x: hidden !important;
            }}
            .sidebar-backdrop {{
                position: fixed !important;
                top: 0 !important;
                right: 0 !important;
                width: 100vw !important;
                height: 100vh !important;
                background: rgba(0, 0, 0, 0.7) !important;
                z-index: 999998 !important;
            }}
            .nav-btn-container .stButton > button {{
                font-size: 1rem !important;
                padding: 1rem 1.2rem !important;
                min-height: 56px !important;
            }}
        }}
        .sidebar-brand {{
            display: flex !important;
            align-items: center !important;
            gap: 0.75rem !important;
            padding: 1rem !important;
            margin-bottom: 0.5rem !important;
        }}
        .brand-logo {{
            width: 44px !important;
            height: 44px !important;
            border-radius: 12px !important;
            background: linear-gradient(135deg, var(--primary), var(--secondary)) !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            font-size: 1.4rem !important;
            color: white !important;
            box-shadow: 0 4px 12px rgba(37, 99, 235, 0.25) !important;
        }}
        .brand-text h3 {{
            font-size: 0.95rem !important;
            font-weight: 700 !important;
            color: var(--text) !important;
            margin: 0 !important;
        }}
        .brand-text small {{
            font-size: 0.7rem !important;
            color: var(--text-muted) !important;
        }}
        .sidebar-user {{
            display: flex !important;
            align-items: center !important;
            gap: 0.75rem !important;
            padding: 0.75rem !important;
            background: #f8fafc !important;
            border-radius: var(--radius-sm) !important;
            margin-bottom: 1rem !important;
            border: 1px solid var(--border) !important;
        }}
        .user-avatar-lg {{
            width: 40px !important;
            height: 40px !important;
            border-radius: 50% !important;
            background: linear-gradient(135deg, var(--primary), var(--secondary)) !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            color: white !important;
            font-weight: 700 !important;
            font-size: 0.9rem !important;
            box-shadow: 0 2px 8px rgba(37, 99, 235, 0.2) !important;
        }}
        .user-info {{
            display: flex !important;
            flex-direction: column !important;
        }}
        .user-info strong {{
            font-size: 0.9rem !important;
            color: var(--text) !important;
            font-weight: 700 !important;
        }}
        .user-info span {{
            font-size: 0.75rem !important;
            color: var(--text-muted) !important;
        }}
        .sidebar-footer {{
            padding: 0.5rem !important;
            margin-top: 1rem !important;
            border-top: 1px solid var(--border) !important;
        }}

        /* ===== Navigation Buttons ===== */
        .nav-btn-container .stButton > button {{
            width: 100% !important;
            text-align: right !important;
            justify-content: flex-start !important;
            padding: 0.85rem 1rem !important;
            font-size: 0.9rem !important;
            font-weight: 600 !important;
            border-radius: var(--radius-sm) !important;
            background: transparent !important;
            color: var(--text) !important;
            border: 1px solid transparent !important;
            transition: all 0.2s ease !important;
            font-family: var(--font) !important;
        }}
        .nav-btn-container .stButton > button:hover {{
            background: var(--primary-light) !important;
            border-color: var(--primary) !important;
            transform: translateX(-4px) !important;
        }}
        .nav-btn-container .stButton > button[kind="primary"] {{
            background: linear-gradient(135deg, var(--primary), var(--secondary)) !important;
            color: white !important;
            box-shadow: 0 4px 12px rgba(37, 99, 235, 0.3) !important;
            border: none !important;
        }}
        .nav-btn-container .stButton > button[kind="primary"]:hover {{
            background: linear-gradient(135deg, var(--primary-dark), var(--secondary)) !important;
            box-shadow: 0 6px 16px rgba(37, 99, 235, 0.4) !important;
        }}

        /* ===== Buttons ===== */
        .stButton > button, .stDownloadButton > button {{
            border-radius: var(--radius-sm) !important;
            font-family: var(--font) !important;
            font-weight: 600 !important;
            transition: all 0.2s ease !important;
            border: 1px solid var(--border) !important;
            box-shadow: var(--shadow-sm) !important;
        }}
        .stButton > button:hover, .stDownloadButton > button:hover {{
            transform: translateY(-1px) !important;
            box-shadow: var(--shadow-md) !important;
        }}
        button[kind="primary"], button[kind="primaryFormSubmit"] {{
            background: linear-gradient(135deg, var(--primary), var(--secondary)) !important;
            color: white !important;
            border: none !important;
            box-shadow: var(--shadow-md) !important;
        }}

        /* ===== Inputs ===== */
        input, textarea, select, .stSelectbox > div, .stTextInput > div {{
            border-radius: var(--radius-sm) !important;
            border: 1px solid var(--border) !important;
            font-family: var(--font) !important;
            background: var(--card) !important;
            color: var(--text) !important;
            box-shadow: var(--shadow-sm) !important;
        }}
        input:focus, textarea:focus, select:focus {{
            border-color: var(--primary) !important;
            box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.1) !important;
            outline: none !important;
        }}
        [data-testid="stTextInput"] label, [data-testid="stSelectbox"] label {{
            color: var(--text) !important;
            font-weight: 600 !important;
            font-size: 0.85rem !important;
        }}

        /* ===== Metrics ===== */
        div[data-testid="stMetric"] {{
            background: var(--card) !important;
            border-radius: var(--radius) !important;
            padding: 1.25rem !important;
            border: 1px solid var(--border) !important;
            box-shadow: var(--shadow-md) !important;
            text-align: right !important;
        }}
        div[data-testid="stMetric"] label {{
            color: var(--text-muted) !important;
            font-weight: 600 !important;
            font-size: 0.85rem !important;
            text-align: right !important;
        }}
        div[data-testid="stMetricValue"] {{
            color: var(--primary) !important;
            font-weight: 800 !important;
            font-size: 1.75rem !important;
            text-align: right !important;
        }}

        /* ===== Tabs ===== */
        button[data-testid="stTab"] {{
            border-radius: var(--radius-sm) !important;
            font-weight: 600 !important;
            font-size: 0.9rem !important;
            padding: 0.6rem 1.2rem !important;
        }}
        button[data-testid="stTab"][aria-selected="true"] {{
            color: var(--primary) !important;
            border-bottom: 3px solid var(--primary) !important;
        }}

        /* ===== Expanders ===== */
        .streamlit-expanderHeader {{
            border-radius: var(--radius-sm) !important;
            font-weight: 700 !important;
            color: var(--text) !important;
            background: var(--card) !important;
            border: 1px solid var(--border) !important;
        }}

        /* ===== DataFrames ===== */
        .stDataFrame {{
            background: var(--card) !important;
            border-radius: var(--radius) !important;
            border: 1px solid var(--border) !important;
            box-shadow: var(--shadow-md) !important;
            direction: rtl !important;
        }}

        /* ===== Alerts ===== */
        .stSuccess {{
            background: var(--success-light) !important;
            border: 1px solid #a7f3d0 !important;
            color: #065f46 !important;
            border-radius: var(--radius-sm) !important;
        }}
        .stError {{
            background: var(--danger-light) !important;
            border: 1px solid #fecaca !important;
            color: #991b1b !important;
            border-radius: var(--radius-sm) !important;
        }}
        .stWarning {{
            background: var(--warning-light) !important;
            border: 1px solid #fde68a !important;
            color: #92400e !important;
            border-radius: var(--radius-sm) !important;
        }}
        .stInfo {{
            background: var(--primary-light) !important;
            border: 1px solid #bfdbfe !important;
            color: #1e40af !important;
            border-radius: var(--radius-sm) !important;
        }}

        /* ===== Badges ===== */
        .status-badge, .student-badge, .card-badge {{
            display: inline-block !important;
            padding: 0.3rem 0.8rem !important;
            border-radius: 9999px !important;
            font-size: 0.75rem !important;
            font-weight: 600 !important;
        }}
        .status-badge.active, .student-badge.active, .card-badge.active {{
            background: var(--success-light) !important;
            color: #065f46 !important;
        }}
        .status-badge.inactive, .student-badge.inactive, .card-badge.inactive {{
            background: #f1f5f9 !important;
            color: #475569 !important;
        }}
        .role-badge {{
            display: inline-block !important;
            padding: 0.25rem 0.7rem !important;
            border-radius: 9999px !important;
            font-size: 0.75rem !important;
            font-weight: 600 !important;
        }}
        .role-badge.admin {{
            background: #ede9fe !important;
            color: #5b21b6 !important;
        }}
        .role-badge.priest {{
            background: #fef3c7 !important;
            color: #92400e !important;
        }}
        .role-badge.leader {{
            background: #dbeafe !important;
            color: #1e40af !important;
        }}
        .role-badge.teacher {{
            background: #fce7f3 !important;
            color: #9d174d !important;
        }}
        .role-badge.student {{
            background: var(--success-light) !important;
            color: #065f46 !important;
        }}

        /* ===== Avatars ===== */
        .user-avatar, .student-avatar-large {{
            width: 56px !important;
            height: 56px !important;
            border-radius: 50% !important;
            background: linear-gradient(135deg, var(--primary), var(--secondary)) !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            color: white !important;
            font-size: 1.2rem !important;
            font-weight: 700 !important;
            box-shadow: 0 4px 12px rgba(37, 99, 235, 0.25) !important;
        }}
        .student-avatar-large {{
            width: 48px !important;
            height: 48px !important;
            font-size: 1rem !important;
        }}

        /* ===== Profile Header ===== */
        .profile-header {{
            background: linear-gradient(135deg, var(--primary), var(--secondary)) !important;
            border-radius: var(--radius) !important;
            padding: 2rem !important;
            color: white !important;
            box-shadow: var(--shadow-lg) !important;
            margin-bottom: 1.5rem !important;
        }}
        .profile-header h1 {{
            color: white !important;
        }}

        /* ===== Hide Streamlit Elements ===== */
        header[data-testid="stHeader"] {{ display: none !important; }}
        #MainMenu {{ visibility: hidden; }}
        footer {{ visibility: hidden; }}
        [data-testid="InputInstructions"] {{ display: none !important; }}
        [data-testid="stToolbar"],
        [data-testid="stDeployButton"],
        [data-testid="stStatusWidget"],
        [data-testid="baseButton-header"] {{
            display: none !important;
            visibility: hidden !important;
        }}

        /* ===== Scrollbar ===== */
        ::-webkit-scrollbar {{ width: 8px !important; height: 8px !important; }}
        ::-webkit-scrollbar-track {{ background: transparent !important; }}
        ::-webkit-scrollbar-thumb {{ background: #cbd5e1 !important; border-radius: 4px !important; }}
        ::-webkit-scrollbar-thumb:hover {{ background: var(--primary) !important; }}

        /* ===== Responsive ===== */
        @media (max-width: 768px) {{
            .content-area {{ padding: 1rem !important; }}
            .main-header {{ font-size: 1.5rem !important; padding: 1rem !important; }}
            input, textarea, select, .stButton > button {{
                font-size: 16px !important;
                min-height: 44px !important;
            }}
            .nav-btn-container .stButton > button {{
                font-size: 0.85rem !important;
                padding: 0.6rem 0.8rem !important;
            }}
        }}

        /* ===== RTL Specific ===== */
        .stButton > button {{
            text-align: right !important;
        }}
        [data-testid="stSidebar"] .stButton > button {{
            text-align: right !important;
        }}
    </style>
    """


def inject_css():
    """Inject the unified premium design system."""
    st.markdown(get_design_css(), unsafe_allow_html=True)


def inject_top_bar_css():
    """Styles for the global top bar (Help + Menu buttons)."""
    st.markdown("""
    <style>
    .app-top-bar {
        direction: rtl;
        margin-bottom: 0.5rem;
        padding-bottom: 0.25rem;
        border-bottom: 1px solid #e5e7eb;
    }
    .app-top-bar [data-testid="stButton"] > button {
        background-color: #2563eb !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 12px !important;
        min-height: 48px !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
        box-shadow: 0 2px 8px rgba(37, 99, 235, 0.22) !important;
    }
    .app-top-bar [data-testid="stButton"] > button:hover {
        background-color: #1d4ed8 !important;
        color: #ffffff !important;
    }
    .app-top-title-center {
        text-align: center;
        font-size: 1.05rem;
        font-weight: 800;
        color: #0f172a;
        margin: 0.65rem 0 0.85rem 0;
        line-height: 1.45;
    }
    @media (max-width: 480px) {
        .app-top-bar [data-testid="stButton"] > button {
            min-height: 52px !important;
            font-size: 1rem !important;
        }
        .app-top-title-center { font-size: 0.92rem; }
    }
    </style>
    """, unsafe_allow_html=True)


def render_help_center_button():
    """Blue Help Center button — call at most once per Streamlit run."""
    if st.button("❓ مركز المساعدة", key="app_help_center_btn", use_container_width=True):
        st.session_state.open_help_dialog = True
        st.rerun()


def render_student_menu_button():
    """Blue menu button for student portal — call at most once per Streamlit run."""
    if st.button("القائمة", key="app_student_menu_btn", use_container_width=True):
        st.session_state.sidebar_open = not st.session_state.get("sidebar_open", False)
        st.rerun()


def render_student_top_bar(current_page):
    """Single student top bar: Help (left), Menu (right), optional competitions title below."""
    inject_top_bar_css()
    st.markdown('<div class="app-top-bar">', unsafe_allow_html=True)
    is_competitions = current_page == STUDENT_ASSESSMENTS_PAGE
    if is_competitions:
        c_left, c_right = st.columns(2)
        with c_left:
            render_help_center_button()
        with c_right:
            render_student_menu_button()
        st.markdown(
            '<p class="app-top-title-center">المسابقات والاختبارات 🏆</p>',
            unsafe_allow_html=True,
        )
    else:
        c_left, _c_mid, c_right = st.columns([1, 0.15, 1])
        with c_left:
            render_help_center_button()
        with c_right:
            render_student_menu_button()
    st.markdown('</div>', unsafe_allow_html=True)


def render_login_top_bar():
    """Help Center on login page."""
    inject_top_bar_css()
    st.markdown('<div class="app-top-bar">', unsafe_allow_html=True)
    c1, _c2, _c3 = st.columns([1, 1, 1])
    with c1:
        render_help_center_button()
    st.markdown('</div>', unsafe_allow_html=True)


def render_admin_top_bar(show_menu_button=False):
    """Help Center (+ optional menu) for admin/staff pages."""
    inject_top_bar_css()
    st.markdown('<div class="app-top-bar">', unsafe_allow_html=True)
    if show_menu_button:
        c_left, c_right = st.columns(2)
        with c_left:
            render_help_center_button()
        with c_right:
            if st.button("القائمة", key="app_admin_menu_btn", use_container_width=True):
                st.session_state.show_sidebar = True
                st.rerun()
    else:
        c1, _c2, _c3 = st.columns([1, 1, 1])
        with c1:
            render_help_center_button()
    st.markdown('</div>', unsafe_allow_html=True)


def inject_user_cards_css():
    """Design system already applies globally — no-op for backwards compat."""
    pass


def inject_students_cards_css():
    """Design system already applies globally — no-op for backwards compat."""
    pass


def page_header(title, subtitle=""):
    """Render a premium page header."""
    return f"""
    <div class="profile-header">
        <h1 style="margin:0; font-size:1.5rem; font-weight:800;">{title}</h1>
        {f'<p style="margin:0.5rem 0 0; opacity:0.9; font-size:0.9rem;">{subtitle}</p>' if subtitle else ''}
    </div>
    """


def hero_header(title, subtitle=""):
    """Render the reusable hero section with background image."""
    return f"""
    <div class="hero-banner">
        <div class="hero-content">
            <h1 class="hero-title">{title}</h1>
            {f'<p class="hero-subtitle">{subtitle}</p>' if subtitle else ''}
        </div>
    </div>
    """


def empty_state(message, icon="📭"):
    """Render an empty state message."""
    return f"""
    <div style="text-align:center; padding:3rem 1rem; color:#6b7280;">
        <div style="font-size:3rem; margin-bottom:0.5rem;">{icon}</div>
        <p style="font-size:1rem; font-weight:600; margin:0;">{message}</p>
    </div>
    """


def stat_card(label, value, icon="📊"):
    """Render a stat card."""
    return f"""
    <div class="profile-stat-card" style="text-align:center; padding:1.25rem;">
        <div style="font-size:1.5rem; font-weight:800; color:#1a56db;">{icon} {value}</div>
        <div style="font-size:0.8rem; color:#6b7280; font-weight:600; margin-top:0.25rem;">{label}</div>
    </div>
    """



def under_development_page(title, subtitle, message, button_label="العودة إلى لوحة التحكم", button_key=None, features=None):
    """
    Render a premium Arabic 'Under Development' page with RTL support, Cairo font, soft animations, and Bootstrap 5 styling.
    
    Args:
        title: Main title for the page
        subtitle: Subtitle below the title
        message: Main message/description
        button_label: Label for the back button
        button_key: Unique key for the button
        features: List of feature strings to display as bullets (optional)
    """
    inject_css()
    st.markdown(hero_header(title, subtitle), unsafe_allow_html=True)

    # Build features HTML if provided
    features_html = ""
    if features and isinstance(features, list):
        features_html = "<ul style='text-align: right; line-height: 2;'>"
        for feat in features:
            features_html += f"<li>✅ {feat}</li>"
        features_html += "</ul>"
    
    # Replace message with features if provided
    if features_html:
        message = features_html

    # Custom styles for the under-development page
    st.markdown("""
    <style>
        .under-dev-container {
            max-width: 700px;
            margin: 2rem auto;
            text-align: center;
            padding: 2.5rem 2rem;
            background: #ffffff;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.08);
            border: 1px solid #e2e8f0;
            animation: fadeInUp 0.8s ease-out;
            direction: rtl;
            font-family: 'Cairo', sans-serif;
        }
        .under-dev-icon {
            font-size: 5rem;
            margin-bottom: 1.2rem;
            animation: pulse 2s infinite;
        }
        .under-dev-title {
            font-size: 2rem;
            font-weight: 800;
            color: #0f172a;
            margin-bottom: 0.5rem;
            line-height: 1.4;
        }
        .under-dev-subtitle {
            font-size: 1.2rem;
            color: #475569;
            margin-bottom: 1.5rem;
            font-weight: 500;
            line-height: 1.7;
        }
        .under-dev-message {
            font-size: 1rem;
            color: #64748b;
            margin-bottom: 2rem;
            line-height: 1.8;
            background: #f8fafc;
            padding: 1rem 1.5rem;
            border-radius: 12px;
            border-right: 4px solid #2563eb;
            text-align: right;
        }
        .under-dev-message ul {
            margin: 0;
            padding-right: 1.5rem;
        }
        .under-dev-message li {
            margin-bottom: 0.5rem;
        }
        .under-dev-btn {
            background: linear-gradient(135deg, #2563eb, #7c3aed);
            color: white;
            border: none;
            padding: 0.9rem 2.2rem;
            font-size: 1rem;
            font-weight: 700;
            border-radius: 12px;
            cursor: pointer;
            box-shadow: 0 8px 20px rgba(37,99,235,0.3);
            transition: all 0.3s ease;
            font-family: 'Cairo', sans-serif;
            text-decoration: none;
            display: inline-block;
        }
        .under-dev-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 12px 24px rgba(37,99,235,0.4);
        }
        @keyframes fadeInUp {
            from { opacity:0; transform:translateY(30px); }
            to { opacity:1; transform:translateY(0); }
        }
        @keyframes pulse {
            0% { transform: scale(1); }
            50% { transform: scale(1.1); }
            100% { transform: scale(1); }
        }
        @media (max-width: 768px) {
            .under-dev-container { margin: 1rem; padding: 1.5rem 1rem; }
            .under-dev-icon { font-size: 3.5rem; }
            .under-dev-title { font-size: 1.5rem; }
            .under-dev-subtitle { font-size: 1rem; }
        }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="under-dev-container">
        <div class="under-dev-icon">🚧</div>
        <div class="under-dev-title">{title}</div>
        <div class="under-dev-subtitle">{subtitle}</div>
        <div class="under-dev-message">{message}</div>
    </div>
    """, unsafe_allow_html=True)

    if st.button(button_label, use_container_width=True, key=button_key or "under_dev_back"):
        st.session_state.show_exam_portal = False
        st.rerun()



def info_row(label, value):
    """Render an info row."""
    return f"""
    <div style="display:flex; justify-content:space-between; padding:0.5rem 0; border-bottom:1px solid #f3f4f6;">
        <span style="color:#6b7280; font-size:0.85rem;">{label}</span>
        <span style="color:#1f2937; font-weight:600; font-size:0.85rem;">{value}</span>
    </div>
    """


# =============================================================================
# Cache & Retry
# =============================================================================
def init_data_cache():
    if 'data_cache' not in st.session_state:
        st.session_state.data_cache = {}
    if 'data_dirty' not in st.session_state:
        st.session_state.data_dirty = {}
    if 'cache_stats' not in st.session_state:
        st.session_state.cache_stats = {'hits': 0, 'misses': 0, 'last_cleanup': time.time()}
    # Auto-invalidation: periodically clean expired cache entries
    now = time.time()
    if now - st.session_state.cache_stats.get('last_cleanup', 0) > 300:  # Every 5 minutes
        cache = st.session_state.get('data_cache', {})
        expired_keys = [k for k, v in cache.items() if now - v.get('timestamp', 0) > CACHE_TTL_SECONDS]
        for k in expired_keys:
            del cache[k]
        st.session_state.cache_stats['last_cleanup'] = now


def retry_operation(max_retries=5, base_delay=2):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except gspread.exceptions.APIError as e:
                    last_exception = e
                    if attempt < max_retries - 1:
                        delay = base_delay * (2 ** attempt)
                        st.warning(f"⏳ النظام مشغول، جاري المحاولة تاني... (محاولة {attempt+1})")
                        time.sleep(delay)
                    else:
                        st.error("❌ النظام مشغول حالياً، من فضلك انتظر دقيقة وحمّل الصفحة تاني")
                        raise last_exception
                except Exception as e:
                    last_exception = e
                    if attempt < max_retries - 1:
                        time.sleep(base_delay * (2 ** attempt))
                    else:
                        raise last_exception
            return None
        return wrapper
    return decorator


# =============================================================================
# Audit Log - Client Info (IP, location, browser, OS, device)
# =============================================================================
def _get_client_ip_and_location():
    """
    جلب عنوان IP والموقع الجغرافي باستخدام ipapi.co API المجاني.
    """
    try:
        resp = requests.get("https://ipapi.co/json/", timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            return {
                "ip_address": data.get("ip", ""),
                "country": data.get("country_name", ""),
                "city": data.get("city", "")
            }
    except Exception:
        pass
    return {"ip_address": "", "country": "", "city": ""}


def _parse_user_agent(ua_string):
    """
    تحليل User-Agent لاستخراج: browser, os, device_type.
    """
    ua = ua_string or ""
    result = {"browser": "", "os": "", "device_type": "", "screen_size": ""}

    # Browser detection
    browser_patterns = [
        (r"Edge|Edg/", "Edge"),
        (r"Chrome/", "Chrome"),
        (r"Firefox/", "Firefox"),
        (r"Safari/", "Safari"),
        (r"Opera|OPR/", "Opera"),
        (r"MSIE|Trident/", "Internet Explorer"),
    ]
    for pattern, name in browser_patterns:
        if re.search(pattern, ua):
            result["browser"] = name
            break

    # OS detection
    os_patterns = [
        (r"Windows NT 10\.0", "Windows 10/11"),
        (r"Windows NT 6\.\d", "Windows"),
        (r"Android", "Android"),
        (r"iPhone|iPad|iOS", "iOS"),
        (r"Mac OS X", "macOS"),
        (r"Linux", "Linux"),
    ]
    for pattern, name in os_patterns:
        if re.search(pattern, ua):
            result["os"] = name
            break

    # Device type detection
    if re.search(r"Mobile|Android|iPhone|iPad|iPod", ua):
        if re.search(r"iPad", ua):
            result["device_type"] = "Tablet"
        else:
            result["device_type"] = "Mobile"
    else:
        result["device_type"] = "Desktop"

    return result


def _get_screen_size():
    """
    محاولة جلب حجم الشاشة باستخدام JavaScript عبر streamlit_js_eval.
    إذا فشل، نعيد قيمة افتراضية.
    """
    try:
        # استخدم get_page_location أو get_browser_language كطريقة للحصول على معلومات
        from streamlit_js_eval import get_page_location
        # لا نستطيع الحصول على screen size مباشرة، لذا نعيد قيمة افتراضية
        pass
    except Exception:
        pass
    return ""


def get_client_info():
    """
    تجميع معلومات العميل: IP، الموقع، المتصفح، نظام التشغيل، نوع الجهاز.
    """
    info = _get_client_ip_and_location()
    
    # محاولة الحصول على User-Agent من streamlit_js_eval
    ua_string = ""
    try:
        from streamlit_js_eval import get_user_agent
        ua_result = get_user_agent()
        if ua_result and isinstance(ua_result, str):
            ua_string = ua_result
        elif ua_result and isinstance(ua_result, dict):
            ua_string = ua_result.get("userAgent", "")
    except Exception:
        pass
    
    # إذا لم نتمكن من الحصول على User-Agent من JS، نستخدم طريقة بديلة
    if not ua_string:
        try:
            # محاولة من request headers (قد لا تكون متاحة في Streamlit)
            ua_string = st.context.headers.get("User-Agent", "") if hasattr(st, 'context') else ""
        except Exception:
            pass

    parsed = _parse_user_agent(ua_string)
    info.update(parsed)
    info["screen_size"] = _get_screen_size()
    return info


# =============================================================================
# Database Class
# =============================================================================
class Database:
    _request_times = []
    _lock = threading.Lock()

    @staticmethod
    def _rate_limit():
        now = time.time()
        with Database._lock:
            Database._request_times = [t for t in Database._request_times if now - t < 60]
            if len(Database._request_times) >= 40:
                sleep_time = 60 - (now - Database._request_times[0]) + 1
                if sleep_time > 0:
                    time.sleep(sleep_time)
                Database._request_times = []
            Database._request_times.append(time.time())

    def __init__(self, creds, spreadsheet_id):
        self.client = gspread.authorize(creds)
        self.spreadsheet = self.client.open_by_key(spreadsheet_id)

    def _get_or_create_worksheet(self, name, columns):
        Database._rate_limit()
        try:
            ws = self.spreadsheet.worksheet(name)
        except gspread.WorksheetNotFound:
            ws = self.spreadsheet.add_worksheet(title=name, rows=1000, cols=max(len(columns), 1))
            if columns:
                try:
                    ws.append_row(columns)
                except Exception:
                    pass
        time.sleep(0.2)
        return ws

    def _get_cached_df(self, sheet_name, fetch_func):
        init_data_cache()
        cache = st.session_state.data_cache
        dirty = st.session_state.data_dirty
        now = time.time()
        if sheet_name in cache and not dirty.get(sheet_name, False):
            entry = cache[sheet_name]
            if now - entry['timestamp'] < CACHE_TTL_SECONDS:
                return entry['data'].copy()
        df = fetch_func()
        st.session_state.data_cache[sheet_name] = {'data': df.copy(), 'timestamp': now}
        st.session_state.data_dirty[sheet_name] = False
        return df.copy()

    def _invalidate_cache(self, sheet_name):
        init_data_cache()
        st.session_state.data_dirty[sheet_name] = True

    def _read_sheet_raw(self, sheet_name):
        Database._rate_limit()
        ws = self._get_or_create_worksheet(sheet_name, [])
        values = ws.get_all_values()
        time.sleep(0.2)
        if not values or len(values) < 1:
            return pd.DataFrame()
        raw_headers = [h.strip() for h in values[0]]
        seen = {}
        unique_headers = []
        for h in raw_headers:
            if h in seen:
                seen[h] += 1
                unique_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                unique_headers.append(h)
        data_rows = values[1:]
        df = pd.DataFrame(data_rows, columns=unique_headers)
        df.dropna(how='all', axis=1, inplace=True)
        df.dropna(how='all', inplace=True)
        return df.astype(object)

    def _sheet_to_df(self, sheet_name):
        return self._get_cached_df(sheet_name, lambda: self._read_sheet_raw(sheet_name))

    def _df_to_sheet(self, sheet_name, df, columns):
        if not isinstance(df, pd.DataFrame):
            raise ValueError("df must be a DataFrame")
        if not isinstance(columns, list) or not columns:
            raise ValueError("columns must be a non-empty list")
        Database._rate_limit()
        ws = self._get_or_create_worksheet(sheet_name, columns)
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        work_df = df[columns].copy()
        work_df.fillna("", inplace=True)
        work_df = work_df.astype(str)
        values = [columns] + work_df.values.tolist()
        try:
            ws.resize(rows=len(values), cols=len(columns))
            ws.update(values)
            time.sleep(0.2)
            self._invalidate_cache(sheet_name)
        except Exception as e:
            raise e

    @staticmethod
    def _safe_str(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        if isinstance(value, (dict, list)):
            return str(value)
        return str(value)

    # --- Users ---
    def get_users(self):
        return self._sheet_to_df("Users")

    def ensure_all_sheets_exist(self):
        """Ensure all required sheets exist with proper columns."""
        sheets_config = {
            "Users": ["user_id", "username", "password", "role", "full_name", "section_id", "phone", "email"],
            "Students": ["student_id", "student_code", "student_password", "full_name", "section_id", "teacher_id", "phone", "parent_phone", "birthdate", "address", "notes", "school", "status", "profile_edit_used"],
            "Exams": ["exam_id", "title", "description", "created_by", "stage_id", "section_id", "chapter_lesson", "exam_date", "start_date", "end_date", "duration_minutes", "total_marks", "passing_score", "is_active", "is_published", "created_at"],
            "Stages": self.STAGE_COLUMNS,
            "StageSupervisors": self.STAGE_SUPERVISOR_COLUMNS,
            "SectionTeachers": self.SECTION_TEACHER_COLUMNS,
            "Sections": self.SECTION_COLUMNS,
            "Attendance": self.ATTENDANCE_COLUMNS,
            "FollowUp": ["record_id", "student_id", "teacher_id", "followup_date", "followup_type", "notes", "regularity_status"],
            "Quizzes": self.QUIZ_COLUMNS,
            "QuizQuestions": self.QUIZ_QUESTION_COLUMNS,
            "QuizResults": self.QUIZ_RESULT_COLUMNS,
            "AuditLog": AUDIT_LOG_COLUMNS,
            "Events": self.EVENT_COLUMNS,
            "EventRSVP": self.EVENT_RSVP_COLUMNS,
            "EventAttendance": self.EVENT_ATTENDANCE_COLUMNS,
            "ExamQuestions": ["question_id", "exam_id", "question_text", "question_type", "option1", "option2", "option3", "option4", "correct_answer", "marks"],
            "ExamResults": ["result_id", "exam_id", "student_id", "student_name", "score", "total_marks", "start_time", "submission_time", "answers", "status"],
            "Homeworks": ["homework_id", "title", "description", "created_by", "section_id", "subject", "due_date", "total_marks", "is_active", "created_at"],
            "HomeworkSubmissions": ["submission_id", "homework_id", "student_id", "student_name", "section_id", "image_data", "image_name", "submission_note", "status", "grade", "feedback", "submitted_at", "reviewed_by", "reviewed_at"],
            "Notifications": ["notification_id", "user_id", "title", "message", "notification_type", "is_read", "created_at"]
        }
        for sheet_name, columns in sheets_config.items():
            try:
                self._get_or_create_worksheet(sheet_name, columns)
            except Exception:
                pass

    def add_user(self, user_data):
        df = self.get_users()
        if df.empty:
            df = pd.DataFrame(columns=["user_id", "username", "password", "role",
                                       "full_name", "section_id", "phone", "email"])
        df = pd.concat([df, pd.DataFrame([user_data])], ignore_index=True)
        self._df_to_sheet("Users", df, ["user_id", "username", "password", "role",
                                        "full_name", "section_id", "phone", "email"])

    def update_user(self, user_id, updates):
        df = self.get_users()
        idx = df[df.user_id == user_id].index
        if len(idx) > 0:
            for k, v in updates.items():
                df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("Users", df, df.columns.tolist())

    def delete_user(self, user_id):
        df = self.get_users()
        df = df[df.user_id != user_id]
        self._df_to_sheet("Users", df, df.columns.tolist())

    # --- Stages ---
    STAGE_COLUMNS = ["stage_id", "stage_name", "description", "display_order",
                     "status", "created_date", "created_by", "manager_user_id", "notes"]

    def get_stages(self):
        return self._sheet_to_df("Stages")

    def add_stage(self, stage_data):
        df = self.get_stages()
        if df.empty:
            df = pd.DataFrame(columns=self.STAGE_COLUMNS)
        new_row = {
            "stage_id": stage_data["stage_id"],
            "stage_name": stage_data.get("stage_name", ""),
            "description": stage_data.get("description", ""),
            "display_order": stage_data.get("display_order", ""),
            "status": stage_data.get("status", "active"),
            "created_date": stage_data.get("created_date", get_cairo_now().strftime("%Y-%m-%d")),
            "created_by": stage_data.get("created_by", ""),
            "manager_user_id": stage_data.get("manager_user_id", ""),
            "notes": stage_data.get("notes", "")
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self._df_to_sheet("Stages", df, self.STAGE_COLUMNS)

    def update_stage(self, stage_id, updates):
        df = self.get_stages()
        idx = df[df.stage_id == stage_id].index
        if len(idx) > 0:
            for k, v in updates.items():
                df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("Stages", df, self.STAGE_COLUMNS)

    def delete_stage(self, stage_id):
        df = self.get_stages()
        df = df[df["stage_id"] != stage_id]
        self._df_to_sheet("Stages", df, self.STAGE_COLUMNS)

    # --- StageSupervisors (Many-to-Many) ---
    STAGE_SUPERVISOR_COLUMNS = ["assignment_id", "stage_id", "supervisor_id", "assigned_date"]

    def get_stage_supervisors(self):
        return self._sheet_to_df("StageSupervisors")

    def get_supervisors_for_stage(self, stage_id):
        df = self.get_stage_supervisors()
        if df.empty:
            return []
        assignments = df[df["stage_id"] == stage_id]
        if assignments.empty:
            return []
        return assignments["supervisor_id"].tolist()

    def get_supervisor_names_for_stage(self, stage_id, users_df=None):
        sup_ids = self.get_supervisors_for_stage(stage_id)
        if not sup_ids:
            return []
        if users_df is None or users_df.empty:
            return sup_ids
        names = []
        for sid in sup_ids:
            match = users_df[users_df["user_id"] == sid]
            if not match.empty:
                names.append(match.iloc[0].get("full_name", sid))
            else:
                names.append(sid)
        return names

    def add_stage_supervisor(self, stage_id, supervisor_id):
        df = self.get_stage_supervisors()
        if df.empty:
            df = pd.DataFrame(columns=self.STAGE_SUPERVISOR_COLUMNS)
        duplicate = df[(df["stage_id"] == stage_id) & (df["supervisor_id"] == supervisor_id)]
        if not duplicate.empty:
            return False
        new_row = {
            "assignment_id": str(uuid.uuid4()),
            "stage_id": stage_id,
            "supervisor_id": supervisor_id,
            "assigned_date": get_cairo_now().strftime("%Y-%m-%d")
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self._df_to_sheet("StageSupervisors", df, self.STAGE_SUPERVISOR_COLUMNS)
        return True

    def remove_stage_supervisor(self, stage_id, supervisor_id):
        df = self.get_stage_supervisors()
        if df.empty:
            return
        df = df[~((df["stage_id"] == stage_id) & (df["supervisor_id"] == supervisor_id))]
        self._df_to_sheet("StageSupervisors", df, self.STAGE_SUPERVISOR_COLUMNS)

    def clear_stage_supervisors(self, stage_id):
        df = self.get_stage_supervisors()
        if df.empty:
            return
        df = df[df["stage_id"] != stage_id]
        self._df_to_sheet("StageSupervisors", df, self.STAGE_SUPERVISOR_COLUMNS)

    def migrate_single_supervisors(self):
        stages = self.get_stages()
        if stages.empty or "manager_user_id" not in stages.columns:
            return 0
        existing_assignments = self.get_stage_supervisors()
        migrated = 0
        for _, row in stages.iterrows():
            stage_id = row.get("stage_id", "")
            mgr_id = row.get("manager_user_id", "")
            if not stage_id or not mgr_id:
                continue
            if existing_assignments.empty or (
                existing_assignments[(existing_assignments["stage_id"] == stage_id) &
                                     (existing_assignments["supervisor_id"] == mgr_id)].empty
            ):
                self.add_stage_supervisor(stage_id, mgr_id)
                migrated += 1
        return migrated

    def migrate_student_codes_and_passwords(self):
        """
        Migration: ensure all students have student_code and student_password.
        Returns: (migrated_count, message)
        """
        students = self.get_students()
        if students.empty:
            return 0, "لا توجد طالبات"
        
        updated = False
        count = 0
        for idx, row in students.iterrows():
            updates = {}
            if not row.get("student_code"):
                updates["student_code"] = generate_student_code(self)
            if not row.get("student_password"):
                updates["student_password"] = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            if updates:
                for k, v in updates.items():
                    students.at[idx, k] = v
                count += 1
                updated = True
        
        # التأكد من وجود عمود profile_edit_used
        if "profile_edit_used" not in students.columns:
            students["profile_edit_used"] = ""

        if updated:
            self._df_to_sheet("Students", students, ["student_id", "student_code", "student_password", "full_name", "section_id", "teacher_id",
                                                     "phone", "parent_phone", "birthdate", "address", "notes", "school", "status", "profile_edit_used"])
        return count, f"تم تحديث {count} طالبة"

    def get_stages_for_supervisor(self, supervisor_id):
        df = self.get_stage_supervisors()
        if df.empty:
            return []
        assignments = df[df["supervisor_id"] == supervisor_id]
        if assignments.empty:
            return []
        return assignments["stage_id"].tolist()

    # --- SectionTeachers (Many-to-Many) ---
    SECTION_TEACHER_COLUMNS = ["assignment_id", "section_id", "teacher_id", "assigned_date"]

    def get_section_teachers(self):
        return self._sheet_to_df("SectionTeachers")

    def get_teachers_for_section(self, section_id):
        df = self.get_section_teachers()
        if df.empty:
            return []
        assignments = df[df["section_id"] == section_id]
        if assignments.empty:
            return []
        return assignments["teacher_id"].tolist()

    def get_teacher_names_for_section(self, section_id, users_df=None):
        t_ids = self.get_teachers_for_section(section_id)
        if not t_ids:
            return []
        if users_df is None or users_df.empty:
            return t_ids
        names = []
        for tid in t_ids:
            match = users_df[users_df["user_id"] == tid]
            if not match.empty:
                names.append(match.iloc[0].get("full_name", tid))
            else:
                names.append(tid)
        return names

    def add_section_teacher(self, section_id, teacher_id):
        df = self.get_section_teachers()
        if df.empty:
            df = pd.DataFrame(columns=self.SECTION_TEACHER_COLUMNS)
        duplicate = df[(df["section_id"] == section_id) & (df["teacher_id"] == teacher_id)]
        if not duplicate.empty:
            return False
        new_row = {
            "assignment_id": str(uuid.uuid4()),
            "section_id": section_id,
            "teacher_id": teacher_id,
            "assigned_date": get_cairo_now().strftime("%Y-%m-%d")
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self._df_to_sheet("SectionTeachers", df, self.SECTION_TEACHER_COLUMNS)
        return True

    def remove_section_teacher(self, section_id, teacher_id):
        df = self.get_section_teachers()
        if df.empty:
            return
        df = df[~((df["section_id"] == section_id) & (df["teacher_id"] == teacher_id))]
        self._df_to_sheet("SectionTeachers", df, self.SECTION_TEACHER_COLUMNS)

    def clear_section_teachers(self, section_id):
        df = self.get_section_teachers()
        if df.empty:
            return
        df = df[df["section_id"] != section_id]
        self._df_to_sheet("SectionTeachers", df, self.SECTION_TEACHER_COLUMNS)

    # --- Sections ---
    SECTION_COLUMNS = ["section_id", "section_name", "stage_id", "teacher_id", "leader_id",
                       "max_students", "room", "meeting_day", "meeting_time",
                       "status", "notes", "manager_user_id"]

    def get_sections(self):
        return self._sheet_to_df("Sections")

    def add_section(self, sec_data):
        self._get_or_create_worksheet("Sections", self.SECTION_COLUMNS)
        df = self.get_sections()
        if df.empty:
            df = pd.DataFrame(columns=self.SECTION_COLUMNS)
        new_row = {
            "section_id": sec_data["section_id"],
            "section_name": sec_data.get("section_name", ""),
            "stage_id": sec_data.get("stage_id", ""),
            "teacher_id": sec_data.get("teacher_id", ""),
            "leader_id": sec_data.get("leader_id", ""),
            "max_students": sec_data.get("max_students", ""),
            "room": sec_data.get("room", ""),
            "meeting_day": sec_data.get("meeting_day", ""),
            "meeting_time": sec_data.get("meeting_time", ""),
            "status": sec_data.get("status", "active"),
            "notes": sec_data.get("notes", ""),
            "manager_user_id": sec_data.get("manager_user_id", "")
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self._df_to_sheet("Sections", df, self.SECTION_COLUMNS)

    def update_section(self, section_id, updates):
        df = self.get_sections()
        idx = df[df.section_id == section_id].index
        if len(idx) > 0:
            for k, v in updates.items():
                df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("Sections", df, self.SECTION_COLUMNS)

    def delete_section(self, section_id):
        df = self.get_sections()
        df = df[df.section_id != section_id]
        self._df_to_sheet("Sections", df, self.SECTION_COLUMNS)

    def get_sections_by_stage(self, stage_id):
        df = self.get_sections()
        if df.empty:
            return pd.DataFrame()
        return df[df.stage_id == stage_id]

    def get_sections_by_teacher(self, teacher_id):
        df = self.get_sections()
        if df.empty:
            return pd.DataFrame()
        return df[df.teacher_id == teacher_id]

    def get_sections_by_leader(self, leader_id):
        df = self.get_sections()
        if df.empty:
            return pd.DataFrame()
        return df[df.leader_id == leader_id]

    def get_section_student_count(self, section_id):
        students = self.get_students()
        if students.empty:
            return 0
        return len(students[students.section_id == section_id])

    def move_students_to_section(self, student_ids, new_section_id):
        students = self.get_students()
        if students.empty:
            return
        for sid in student_ids:
            idx = students[students.student_id == sid].index
            if len(idx) > 0:
                students.at[idx[0], "section_id"] = new_section_id
        self._df_to_sheet("Students", students, ["student_id", "student_code", "student_password", "full_name", "section_id", "teacher_id",
                                                 "phone", "parent_phone", "birthdate", "address", "notes", "school", "status", "profile_edit_used"])

    # --- Students ---
    def get_students(self):
        return self._sheet_to_df("Students")

    def ensure_student_profile_edit_column(self):
        """
        التأكد من وجود عمود profile_edit_used في ورقة Students.
        إذا لم يكن موجوداً، يتم إضافته مع تعبئة القيم الفارغة بـ "".
        """
        try:
            df = self.get_students()
            if df.empty:
                return
            if "profile_edit_used" not in df.columns:
                df["profile_edit_used"] = ""
                columns = df.columns.tolist()
                self._df_to_sheet("Students", df, columns)
        except Exception:
            pass

    def add_student(self, student_data):
        df = self.get_students()
        if df.empty:
            df = pd.DataFrame(columns=["student_id", "student_code", "student_password", "full_name", "section_id", "teacher_id",
                                       "phone", "parent_phone", "birthdate", "address", "notes", "school", "status", "profile_edit_used"])
        student_data["teacher_id"] = ""
        student_data.setdefault("profile_edit_used", "")
        # توليد كود فريد إذا لم يكن موجوداً
        if not student_data.get("student_code"):
            student_data["student_code"] = generate_student_code(self)
        # توليد كلمة مرور تلقائياً إذا لم تكن موجودة
        if not student_data.get("student_password"):
            if student_data.get("password"):
                student_data["student_password"] = student_data["password"]
            else:
                student_data["student_password"] = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
        df = pd.concat([df, pd.DataFrame([student_data])], ignore_index=True)
        self._df_to_sheet("Students", df, ["student_id", "student_code", "student_password", "full_name", "section_id", "teacher_id",
                                           "phone", "parent_phone", "birthdate", "address", "notes", "school", "status", "profile_edit_used"])

    def update_student(self, student_id, updates):
        df = self.get_students()
        idx = df[df.student_id == student_id].index
        if len(idx) > 0:
            profile_fields = {"full_name", "phone", "parent_phone", "birthdate", "address", "school", "notes"}
            if any(k in updates for k in profile_fields):
                used_val = str(df.at[idx[0], "profile_edit_used"]).strip().lower()
                if used_val in ("true", "1", "yes", "نعم"):
                    return False
            for k, v in updates.items():
                df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("Students", df, df.columns.tolist())
            return True
        return False

    def delete_student(self, student_id):
        df = self.get_students()
        df = df[df.student_id != student_id]
        self._df_to_sheet("Students", df, df.columns.tolist())

    # --- Attendance ---
    ATTENDANCE_COLUMNS = ["record_id", "date", "time", "user_id", "name", "role", "section_id", "stage_id", "status", "notes", "recorded_by", "attendance_method"]

    def get_attendance(self):
        return self._sheet_to_df("Attendance")

    def batch_add_attendance(self, records_list):
        if not records_list:
            return
        df = self.get_attendance()
        if df.empty:
            df = pd.DataFrame(columns=self.ATTENDANCE_COLUMNS)
        existing_ids = set(df["record_id"].tolist()) if not df.empty else set()
        new_records = []
        for rec in records_list:
            if rec.get("record_id") in existing_ids:
                idx = df[df.record_id == rec["record_id"]].index[0]
                for k, v in rec.items():
                    if k in df.columns:
                        df.at[idx, k] = self._safe_str(v)
            else:
                new_records.append(rec)
        if new_records:
            new_df = pd.DataFrame(new_records)
            df = pd.concat([df, new_df], ignore_index=True)
        self._df_to_sheet("Attendance", df, self.ATTENDANCE_COLUMNS)

    def get_attendance_by_date_user(self, date_str, user_id):
        df = self.get_attendance()
        if df.empty:
            return pd.DataFrame()
        return df[(df.date == date_str) & (df.user_id == user_id)]

    def get_attendance_by_date_section(self, date_str, section_id):
        df = self.get_attendance()
        if df.empty:
            return pd.DataFrame()
        return df[(df.date == date_str) & (df.section_id == section_id)]

    def delete_attendance_record(self, record_id):
        df = self.get_attendance()
        df = df[df.record_id != record_id]
        self._df_to_sheet("Attendance", df, self.ATTENDANCE_COLUMNS)

    def record_qr_attendance(self, student_id, student_name, section_id, stage_id, recorded_by_user_id):
        """
        Record attendance via QR scan.
        Returns dict with success status and message.
        """
        try:
            # Check for duplicate attendance today via QR_SCAN
            today = get_cairo_now().strftime("%Y-%m-%d")
            existing = self.get_attendance_by_date_user(today, student_id)
            if not existing.empty:
                qr_records = existing[existing.get("attendance_method", "") == "QR_SCAN"]
                if not qr_records.empty:
                    return {"success": False, "message": "⚠️ تم تسجيل حضور هذه الطالبة اليوم بالفعل"}
            
            record_id = str(uuid.uuid4())
            now = get_cairo_now()
            current_time = now.strftime("%H:%M:%S")
            
            new_record = {
                "record_id": record_id,
                "date": today,
                "time": current_time,
                "user_id": student_id,
                "name": student_name,
                "role": "Student",
                "section_id": section_id,
                "stage_id": stage_id,
                "status": "حاضر",
                "notes": "تسجيل عبر QR",
                "recorded_by": recorded_by_user_id,
                "attendance_method": "QR_SCAN"
            }
            
            df = self.get_attendance()
            if df.empty:
                df = pd.DataFrame(columns=self.ATTENDANCE_COLUMNS)
            df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)
            self._df_to_sheet("Attendance", df, self.ATTENDANCE_COLUMNS)
            return {"success": True, "message": "تم تسجيل الحضور بنجاح"}
        except Exception as e:
            return {"success": False, "message": f"❌ فشل الاتصال بـ Google Sheets: {str(e)}"}

    # --- FollowUp ---
    def get_followup(self):
        return self._sheet_to_df("FollowUp")

    def add_followup_record(self, record):
        df = self.get_followup()
        if not df.empty:
            duplicate = df[(df.student_id == record["student_id"]) &
                           (df.followup_date == record["followup_date"]) &
                           (df.followup_type == record["followup_type"])]
            if not duplicate.empty:
                raise ValueError("⛔ تم تسجيل نفس الافتقاد مسبقاً لنفس الطالبة في نفس التاريخ ونفس النوع.")
        if df.empty:
            df = pd.DataFrame(columns=["record_id", "student_id", "teacher_id", "followup_date",
                                       "followup_type", "notes", "regularity_status"])
        df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
        self._df_to_sheet("FollowUp", df, ["record_id", "student_id", "teacher_id", "followup_date",
                                           "followup_type", "notes", "regularity_status"])

    def delete_followup_record(self, record_id):
        df = self.get_followup()
        df = df[df.record_id != record_id]
        self._df_to_sheet("FollowUp", df, ["record_id", "student_id", "teacher_id", "followup_date",
                                           "followup_type", "notes", "regularity_status"])

    # --- Quizzes / Unified Assessments ---
    QUIZ_COLUMNS = [
        "quiz_id", "title", "description", "created_by", "section_id",
        "num_questions", "time_limit_minutes", "total_marks", "expiry_date",
        "quiz_code", "password", "is_active",
        "assessment_type", "stage_id", "chapter_lesson", "exam_date",
        "start_date", "end_date", "duration_minutes", "passing_score",
        "is_published", "created_at"
    ]
    QUIZ_QUESTION_COLUMNS = [
        "question_id", "quiz_id", "question_text", "question_type",
        "option1", "option2", "option3", "option4", "correct_answer", "marks"
    ]
    QUIZ_RESULT_COLUMNS = [
        "result_id", "quiz_id", "student_id", "student_name",
        "score", "total_marks", "start_time", "submission_time", "answers", "status"
    ]

    def _normalize_assessment_type(self, value):
        v = str(value or "").strip().lower()
        if v in ("exam", "امتحان"):
            return "exam"
        return "quiz"

    def _ensure_quiz_columns(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=self.QUIZ_COLUMNS)
        work = df.copy()
        for col in self.QUIZ_COLUMNS:
            if col not in work.columns:
                work[col] = ""
        if "assessment_type" not in work.columns:
            work["assessment_type"] = "quiz"
        work["assessment_type"] = work["assessment_type"].apply(self._normalize_assessment_type)
        return work

    def _migrate_legacy_exams_into_quizzes(self):
        """Backfill legacy Exams rows into Quizzes as assessment_type=exam."""
        try:
            quizzes = self._sheet_to_df("Quizzes")
            quizzes = self._ensure_quiz_columns(quizzes)
            existing_ids = set(quizzes["quiz_id"].astype(str).tolist()) if not quizzes.empty else set()
            exams = self._sheet_to_df("Exams")
            if exams.empty:
                return quizzes
            changed = False
            for _, row in exams.iterrows():
                exam_id = str(row.get("exam_id", "")).strip()
                if not exam_id or exam_id in existing_ids:
                    continue
                mapped = {
                    "quiz_id": exam_id,
                    "title": row.get("title", ""),
                    "description": row.get("description", ""),
                    "created_by": row.get("created_by", ""),
                    "section_id": row.get("section_id", ""),
                    "num_questions": "",
                    "time_limit_minutes": row.get("duration_minutes", "30"),
                    "total_marks": row.get("total_marks", "20"),
                    "expiry_date": row.get("end_date", ""),
                    "quiz_code": "",
                    "password": "",
                    "is_active": row.get("is_active", "True"),
                    "assessment_type": "exam",
                    "stage_id": row.get("stage_id", ""),
                    "chapter_lesson": row.get("chapter_lesson", ""),
                    "exam_date": row.get("exam_date", ""),
                    "start_date": row.get("start_date", ""),
                    "end_date": row.get("end_date", ""),
                    "duration_minutes": row.get("duration_minutes", ""),
                    "passing_score": row.get("passing_score", ""),
                    "is_published": row.get("is_published", "False"),
                    "created_at": row.get("created_at", ""),
                }
                quizzes = pd.concat([quizzes, pd.DataFrame([mapped])], ignore_index=True)
                existing_ids.add(exam_id)
                changed = True
            if changed:
                self._df_to_sheet("Quizzes", quizzes, self.QUIZ_COLUMNS)
            return quizzes
        except Exception:
            return self._ensure_quiz_columns(self._sheet_to_df("Quizzes"))

    def get_quizzes(self):
        quizzes = self._migrate_legacy_exams_into_quizzes()
        return self._ensure_quiz_columns(quizzes)

    def add_quiz(self, quiz_data):
        df = self._ensure_quiz_columns(self.get_quizzes())
        quiz_data = {**quiz_data}
        quiz_data["assessment_type"] = self._normalize_assessment_type(quiz_data.get("assessment_type", "quiz"))
        df = pd.concat([df, pd.DataFrame([quiz_data])], ignore_index=True)
        self._df_to_sheet("Quizzes", self._ensure_quiz_columns(df), self.QUIZ_COLUMNS)

    def update_quiz(self, quiz_id, updates):
        df = self._ensure_quiz_columns(self.get_quizzes())
        idx = df[df.quiz_id == quiz_id].index
        if len(idx) > 0:
            for k, v in updates.items():
                if k == "assessment_type":
                    df.at[idx[0], k] = self._normalize_assessment_type(v)
                else:
                    df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("Quizzes", self._ensure_quiz_columns(df), self.QUIZ_COLUMNS)

    def delete_quiz_keep_results(self, quiz_id):
        df = self._ensure_quiz_columns(self.get_quizzes())
        df = df[df.quiz_id != quiz_id]
        self._df_to_sheet("Quizzes", df, self.QUIZ_COLUMNS)
        qdf = self._sheet_to_df("QuizQuestions")
        qdf = qdf[qdf.quiz_id != quiz_id]
        for col in self.QUIZ_QUESTION_COLUMNS:
            if col not in qdf.columns:
                qdf[col] = ""
        self._df_to_sheet("QuizQuestions", qdf, self.QUIZ_QUESTION_COLUMNS)

    def delete_quiz(self, quiz_id):
        self.delete_quiz_keep_results(quiz_id)
        rdf = self._sheet_to_df("QuizResults")
        rdf = rdf[rdf.quiz_id != quiz_id]
        self._df_to_sheet("QuizResults", rdf, self.QUIZ_RESULT_COLUMNS)

    def get_quiz_questions(self, quiz_id):
        df = self._sheet_to_df("QuizQuestions")
        if df.empty:
            return pd.DataFrame()
        for col in self.QUIZ_QUESTION_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[df.quiz_id == quiz_id]

    def add_question(self, q_data):
        df = self._sheet_to_df("QuizQuestions")
        if df.empty:
            df = pd.DataFrame(columns=self.QUIZ_QUESTION_COLUMNS)
        for col in self.QUIZ_QUESTION_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        df = pd.concat([df, pd.DataFrame([q_data])], ignore_index=True)
        self._df_to_sheet("QuizQuestions", df, self.QUIZ_QUESTION_COLUMNS)

    def delete_question(self, question_id):
        df = self._sheet_to_df("QuizQuestions")
        df = df[df.question_id != question_id]
        for col in self.QUIZ_QUESTION_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        self._df_to_sheet("QuizQuestions", df, self.QUIZ_QUESTION_COLUMNS)

    # --- Quiz Results ---
    def get_quiz_results(self, quiz_id=None):
        df = self._sheet_to_df("QuizResults")
        if df.empty:
            return pd.DataFrame()
        for col in self.QUIZ_RESULT_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        if quiz_id:
            return df[df.quiz_id == quiz_id]
        return df

    def start_quiz_attempt(self, quiz_id, student_id, student_name):
        result_id = str(uuid.uuid4())
        now_iso = get_cairo_now().isoformat()
        new_row = {
            "result_id": result_id, "quiz_id": quiz_id, "student_id": student_id,
            "student_name": student_name, "score": "", "total_marks": "20",
            "start_time": now_iso, "submission_time": now_iso, "answers": "{}", "status": "started"
        }
        df = self._sheet_to_df("QuizResults")
        if df.empty:
            df = pd.DataFrame(columns=self.QUIZ_RESULT_COLUMNS)
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        self._df_to_sheet("QuizResults", df, self.QUIZ_RESULT_COLUMNS)
        return result_id

    def save_answers(self, result_id, answers_dict):
        df = self._sheet_to_df("QuizResults")
        idx = df[df.result_id == result_id].index
        if len(idx) > 0:
            df.at[idx[0], "answers"] = json.dumps(answers_dict, ensure_ascii=False)
            self._df_to_sheet("QuizResults", df, self.QUIZ_RESULT_COLUMNS)

    def submit_quiz_attempt(self, result_id, score, answers_json):
        df = self._sheet_to_df("QuizResults")
        idx = df[df.result_id == result_id].index
        if len(idx) > 0:
            df.at[idx[0], "score"] = str(score)
            df.at[idx[0], "answers"] = answers_json
            df.at[idx[0], "submission_time"] = get_cairo_now().isoformat()
            df.at[idx[0], "status"] = "submitted"
            self._df_to_sheet("QuizResults", df, self.QUIZ_RESULT_COLUMNS)

    def delete_quiz_result(self, result_id):
        df = self._sheet_to_df("QuizResults")
        df = df[df.result_id != result_id]
        self._df_to_sheet("QuizResults", df, self.QUIZ_RESULT_COLUMNS)

    # =====================================================================
    # Audit Log - سجل التدقيق الجديد
    # =====================================================================
    def get_audit_log(self):
        """جلب جميع سجلات التدقيق من ورقة AuditLog."""
        return self._sheet_to_df("AuditLog")

    def add_audit_log(self, action, details="", user_info=None, client_info=None):
        """
        إضافة سجل تدقيق جديد إلى ورقة AuditLog.
        
        Parameters:
        - action: نوع العملية (مثل "تسجيل دخول", "إضافة عضو", ...)
        - details: تفاصيل إضافية عن العملية
        - user_info: dict يحتوي على معلومات المستخدم (user_id, username, ...)
        - client_info: dict يحتوي على معلومات العميل (ip, browser, os, ...)
        """
        if user_info is None:
            user_info = st.session_state.get("user", {}) if "user" in st.session_state else {}
        if client_info is None:
            client_info = get_client_info()
        
        username = user_info.get("username", "") if isinstance(user_info, dict) else ""
        user_id = user_info.get("user_id", "") if isinstance(user_info, dict) else ""
        # إذا كان user_info هو مجرد user_id string
        if isinstance(user_info, str):
            user_id = user_info
            username = ""
        elif isinstance(user_info, dict) and not username and user_id:
            # حاول الحصول على اسم المستخدم من users sheet إذا كان متاحاً
            try:
                if 'db_instance' in st.session_state:
                    users_df = st.session_state.db_instance.get_users()
                    if not users_df.empty:
                        match = users_df[users_df.user_id == user_id]
                        if not match.empty:
                            username = match.iloc[0].get("username", "")
            except Exception:
                pass
        
        log_entry = {
            "log_id": str(uuid.uuid4()),
            "timestamp": get_cairo_now().isoformat(),
            "username": username,
            "user_id": user_id,
            "action": action,
            "details": details,
            "ip_address": client_info.get("ip_address", ""),
            "country": client_info.get("country", ""),
            "city": client_info.get("city", ""),
            "browser": client_info.get("browser", ""),
            "os": client_info.get("os", ""),
            "device_type": client_info.get("device_type", ""),
            "screen_size": client_info.get("screen_size", "")
        }
        
        df = self.get_audit_log()
        if df.empty:
            df = pd.DataFrame(columns=AUDIT_LOG_COLUMNS)
        df = pd.concat([df, pd.DataFrame([log_entry])], ignore_index=True)
        self._df_to_sheet("AuditLog", df, AUDIT_LOG_COLUMNS)
        return log_entry["log_id"]

    def delete_audit_log(self, log_id):
        """حذف سجل تدقيق معين."""
        df = self.get_audit_log()
        if df.empty:
            return
        df = df[df.log_id != log_id]
        self._df_to_sheet("AuditLog", df, AUDIT_LOG_COLUMNS)

    # =====================================================================
    # دوال قديمة للتوافق العكسي - تشير إلى AuditLog الجديد
    # =====================================================================
    def get_logs(self):
        """للتوافق مع الكود القديم - يحول إلى AuditLog."""
        return self.get_audit_log()

    def add_log(self, user_id, action, details=""):
        """للتوافق مع الكود القديم - يحول إلى AuditLog."""
        client_info = get_client_info()
        user_info = {"user_id": user_id, "username": ""}
        # محاولة الحصول على اسم المستخدم
        try:
            users_df = self.get_users()
            if not users_df.empty:
                match = users_df[users_df.user_id == user_id]
                if not match.empty:
                    user_info["username"] = match.iloc[0].get("username", "")
                    user_info["full_name"] = match.iloc[0].get("full_name", "")
        except Exception:
            pass
        return self.add_audit_log(action, details, user_info=user_info, client_info=client_info)

    def delete_log(self, log_id):
        """للتوافق مع الكود القديم - يحذف من AuditLog."""
        self.delete_audit_log(log_id)

    # --- Events ---
    EVENT_COLUMNS = ["event_id", "event_name", "event_type", "event_date", "event_time",
                     "location", "max_capacity", "description", "created_by", "status"]

    def get_events(self):
        return self._sheet_to_df("Events")

    def add_event(self, event_data):
        df = self.get_events()
        if df.empty:
            df = pd.DataFrame(columns=self.EVENT_COLUMNS)
        df = pd.concat([df, pd.DataFrame([event_data])], ignore_index=True)
        self._df_to_sheet("Events", df, self.EVENT_COLUMNS)

    def update_event(self, event_id, updates):
        df = self.get_events()
        idx = df[df.event_id == event_id].index
        if len(idx) > 0:
            for k, v in updates.items():
                df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("Events", df, self.EVENT_COLUMNS)

    def delete_event(self, event_id):
        df = self.get_events()
        df = df[df.event_id != event_id]
        self._df_to_sheet("Events", df, self.EVENT_COLUMNS)

    # --- EventRSVP ---
    EVENT_RSVP_COLUMNS = ["rsvp_id", "event_id", "student_id", "student_name", "rsvp_status", "rsvp_date"]

    def get_event_rsvps(self, event_id=None):
        df = self._sheet_to_df("EventRSVP")
        if df.empty or not event_id:
            return df
        return df[df.event_id == event_id]

    def add_event_rsvp(self, rsvp_data):
        df = self.get_event_rsvps()
        if df.empty:
            df = pd.DataFrame(columns=self.EVENT_RSVP_COLUMNS)
        df = pd.concat([df, pd.DataFrame([rsvp_data])], ignore_index=True)
        self._df_to_sheet("EventRSVP", df, self.EVENT_RSVP_COLUMNS)

    def delete_event_rsvp(self, rsvp_id):
        df = self._sheet_to_df("EventRSVP")
        df = df[df.rsvp_id != rsvp_id]
        self._df_to_sheet("EventRSVP", df, self.EVENT_RSVP_COLUMNS)

    # --- EventAttendance ---
    EVENT_ATTENDANCE_COLUMNS = ["record_id", "event_id", "student_id", "status", "notes"]

    def get_event_attendance(self, event_id=None):
        df = self._sheet_to_df("EventAttendance")
        if df.empty or not event_id:
            return df
        return df[df.event_id == event_id]

    def add_event_attendance(self, attendance_data):
        df = self.get_event_attendance()
        if df.empty:
            df = pd.DataFrame(columns=self.EVENT_ATTENDANCE_COLUMNS)
        df = pd.concat([df, pd.DataFrame([attendance_data])], ignore_index=True)
        self._df_to_sheet("EventAttendance", df, self.EVENT_ATTENDANCE_COLUMNS)

    def delete_event_attendance(self, record_id):
        df = self._sheet_to_df("EventAttendance")
        df = df[df.record_id != record_id]
        self._df_to_sheet("EventAttendance", df, self.EVENT_ATTENDANCE_COLUMNS)

    # --- Exams (compatibility wrappers over unified quiz tables) ---
    EXAM_COLUMNS = ["exam_id", "title", "description", "created_by", "stage_id", "section_id", "chapter_lesson", "exam_date", "start_date", "end_date", "duration_minutes", "total_marks", "passing_score", "is_active", "is_published", "created_at"]
    EXAM_QUESTION_COLUMNS = ["question_id", "exam_id", "question_text", "question_type", "option1", "option2", "option3", "option4", "correct_answer", "marks"]
    EXAM_RESULT_COLUMNS = ["result_id", "exam_id", "student_id", "student_name", "score", "total_marks", "start_time", "submission_time", "answers", "status"]

    def get_exams(self):
        quizzes = self.get_quizzes()
        if quizzes.empty:
            return pd.DataFrame(columns=self.EXAM_COLUMNS)
        exams = quizzes[quizzes["assessment_type"] == "exam"].copy() if "assessment_type" in quizzes.columns else pd.DataFrame()
        if exams.empty:
            return pd.DataFrame(columns=self.EXAM_COLUMNS)
        exams["exam_id"] = exams["quiz_id"]
        for col in self.EXAM_COLUMNS:
            if col not in exams.columns:
                exams[col] = ""
        return exams[self.EXAM_COLUMNS]

    def add_exam(self, exam_data):
        mapped = {
            "quiz_id": exam_data.get("exam_id", str(uuid.uuid4())),
            "title": exam_data.get("title", ""),
            "description": exam_data.get("description", ""),
            "created_by": exam_data.get("created_by", ""),
            "section_id": exam_data.get("section_id", ""),
            "num_questions": "",
            "time_limit_minutes": exam_data.get("duration_minutes", "30"),
            "total_marks": exam_data.get("total_marks", "20"),
            "expiry_date": exam_data.get("end_date", ""),
            "quiz_code": "",
            "password": "",
            "is_active": exam_data.get("is_active", "True"),
            "assessment_type": "exam",
            "stage_id": exam_data.get("stage_id", ""),
            "chapter_lesson": exam_data.get("chapter_lesson", ""),
            "exam_date": exam_data.get("exam_date", ""),
            "start_date": exam_data.get("start_date", ""),
            "end_date": exam_data.get("end_date", ""),
            "duration_minutes": exam_data.get("duration_minutes", ""),
            "passing_score": exam_data.get("passing_score", ""),
            "is_published": exam_data.get("is_published", "False"),
            "created_at": exam_data.get("created_at", get_cairo_now().isoformat()),
        }
        self.add_quiz(mapped)

    def update_exam(self, exam_id, updates):
        mapped = {}
        rename_map = {
            "exam_id": "quiz_id",
            "duration_minutes": "time_limit_minutes",
        }
        for k, v in updates.items():
            mapped[rename_map.get(k, k)] = v
        mapped["assessment_type"] = "exam"
        self.update_quiz(exam_id, mapped)

    def delete_exam(self, exam_id):
        self.delete_quiz(exam_id)

    def get_exam_questions(self, exam_id=None):
        if not exam_id:
            all_q = self._sheet_to_df("QuizQuestions")
            for col in self.QUIZ_QUESTION_COLUMNS:
                if col not in all_q.columns:
                    all_q[col] = ""
            all_q["exam_id"] = all_q["quiz_id"]
            return all_q[self.EXAM_QUESTION_COLUMNS]
        qdf = self.get_quiz_questions(exam_id).copy()
        if qdf.empty:
            return pd.DataFrame(columns=self.EXAM_QUESTION_COLUMNS)
        qdf["exam_id"] = qdf["quiz_id"]
        return qdf[self.EXAM_QUESTION_COLUMNS]

    def add_exam_question(self, q_data):
        mapped = {
            "question_id": q_data.get("question_id", str(uuid.uuid4())),
            "quiz_id": q_data.get("exam_id", ""),
            "question_text": q_data.get("question_text", ""),
            "question_type": q_data.get("question_type", ""),
            "option1": q_data.get("option1", ""),
            "option2": q_data.get("option2", ""),
            "option3": q_data.get("option3", ""),
            "option4": q_data.get("option4", ""),
            "correct_answer": q_data.get("correct_answer", ""),
            "marks": q_data.get("marks", "1"),
        }
        self.add_question(mapped)

    def delete_exam_question(self, question_id):
        self.delete_question(question_id)

    def get_exam_results(self, exam_id=None):
        results = self.get_quiz_results(exam_id).copy() if exam_id else self.get_quiz_results().copy()
        if results.empty:
            return pd.DataFrame(columns=self.EXAM_RESULT_COLUMNS)
        results["exam_id"] = results["quiz_id"]
        return results[self.EXAM_RESULT_COLUMNS]

    def start_exam_attempt(self, exam_id, student_id, student_name):
        return self.start_quiz_attempt(exam_id, student_id, student_name)

    def save_exam_answers(self, result_id, answers_dict):
        self.save_answers(result_id, answers_dict)

    def submit_exam_attempt(self, result_id, score, answers_json):
        self.submit_quiz_attempt(result_id, score, answers_json)

    def delete_exam_result(self, result_id):
        self.delete_quiz_result(result_id)

    # --- Exam Engine ---
    def grade_exam_attempt(self, exam_id, answers_dict):
        """
        تصحيح امتحان تلقائياً بناءً على الإجابات.
        returns: (score, total_marks, correct_count, wrong_count)
        """
        questions = self.get_exam_questions(exam_id)
        if questions.empty:
            return 0, 0, 0, 0
        
        correct_count = 0
        wrong_count = 0
        total_marks = 0
        
        for _, q_row in questions.iterrows():
            q = q_row.to_dict()
            q_id = q.get("question_id", "")
            correct = str(q.get("correct_answer", "")).strip().lower()
            student_ans = str(answers_dict.get(q_id, "")).strip().lower()
            marks = float(q.get("marks", 1)) if q.get("marks") else 1
            
            total_marks += marks
            if correct == student_ans:
                correct_count += 1
            else:
                wrong_count += 1
        
        score = correct_count  # Each correct answer gets its marks
        return score, total_marks, correct_count, wrong_count

    def calculate_exam_result(self, result_id):
        """
        حساب النتيجة النهائية لامتحان.
        returns: dict with score, total_marks, percentage, grade
        """
        df = self.get_exam_results()
        idx = df[df.result_id == result_id].index
        if len(idx) == 0:
            return None
        
        row = df.iloc[idx[0]].to_dict()
        exam_id = row.get("exam_id", row.get("quiz_id", ""))
        answers_str = row.get("answers", "{}")
        
        try:
            answers = json.loads(answers_str) if answers_str else {}
        except Exception:
            answers = {}
        
        score, total_marks, correct, wrong = self.grade_exam_attempt(exam_id, answers)
        
        percentage = (score / total_marks * 100) if total_marks > 0 else 0
        
        # Grade calculation
        if percentage >= 90:
            grade = "ممتاز"
        elif percentage >= 80:
            grade = "جيد جداً"
        elif percentage >= 70:
            grade = "جيد"
        elif percentage >= 60:
            grade = "مقبول"
        else:
            grade = "راسب"
        
        return {
            "score": score,
            "total_marks": total_marks,
            "correct_count": correct,
            "wrong_count": wrong,
            "percentage": round(percentage, 1),
            "grade": grade
        }

    def shuffle_questions(self, exam_id):
        """
        خلط أسئلة الامتحان عشوائياً.
        returns: shuffled questions DataFrame
        """
        questions = self.get_exam_questions(exam_id)
        if questions.empty:
            return pd.DataFrame()
        
        shuffled = questions.sample(frac=1).reset_index(drop=True)
        return shuffled

    def shuffle_options(self, question_row):
        """
        خلط خيارات السؤال مع الحفاظ على الإجابة الصحيحة.
        question_row: dict representing a question
        returns: dict with shuffled options and correct answer
        """
        import random
        
        q_type = question_row.get("question_type", "")
        correct = str(question_row.get("correct_answer", "")).strip()
        
        if q_type == "صح وخطأ":
            # For true/false, don't shuffle
            return question_row
        
        # Get options
        options = []
        for i in range(1, 5):
            opt = question_row.get(f"option{i}", "")
            if opt and str(opt).strip():
                options.append(str(opt).strip())
        
        if not options:
            return question_row
        
        # Shuffle options
        random.shuffle(options)
        
        # Create new question row
        new_row = question_row.copy()
        for i, opt in enumerate(options, 1):
            new_row[f"option{i}"] = opt
        
        # Update correct answer to match shuffled position
        if correct in options:
            new_row["correct_answer"] = correct
        
        return new_row

    def get_student_exam_stats(self, student_id):
        """
        جلب إحصائيات الطالبة في الامتحانات.
        returns: dict with stats
        """
        results = self.get_exam_results()
        if results.empty:
            return {
                "total_exams": 0,
                "completed_exams": 0,
                "average_score": 0,
                "highest_score": 0,
                "lowest_score": 0,
                "pass_rate": 0
            }
        
        student_results = results[results["student_id"] == student_id]
        if student_results.empty:
            return {
                "total_exams": 0,
                "completed_exams": 0,
                "average_score": 0,
                "highest_score": 0,
                "lowest_score": 0,
                "pass_rate": 0
            }
        
        submitted = student_results[student_results["status"] == "submitted"]
        total_exams = len(student_results)
        completed_exams = len(submitted)
        
        if completed_exams == 0:
            return {
                "total_exams": total_exams,
                "completed_exams": 0,
                "average_score": 0,
                "highest_score": 0,
                "lowest_score": 0,
                "pass_rate": 0
            }
        
        scores = pd.to_numeric(submitted["score"], errors="coerce").fillna(0)
        total_marks = pd.to_numeric(submitted["total_marks"], errors="coerce").fillna(20)
        
        # Calculate pass/fail (assuming passing is 50% or more)
        passing = scores >= (total_marks * 0.5)
        pass_rate = (passing.sum() / len(scores) * 100) if len(scores) > 0 else 0
        
        return {
            "total_exams": total_exams,
            "completed_exams": completed_exams,
            "average_score": round(scores.mean(), 1),
            "highest_score": round(scores.max(), 1),
            "lowest_score": round(scores.min(), 1),
            "pass_rate": round(pass_rate, 1)
        }

    def get_chapter_tracking(self, exam_id=None):
        """
        تتبع الفصول/المواضيع في الامتحانات.
        returns: DataFrame with chapter tracking info
        """
        questions = self.get_exam_questions()
        if questions.empty:
            return pd.DataFrame()
        
        if exam_id:
            questions = questions[questions["exam_id"] == exam_id]
        
        # Group by question_type (can be extended to actual chapters)
        tracking = questions.groupby("question_type").agg({
            "question_id": "count",
            "correct_answer": lambda x: (x != "").sum()
        }).reset_index()
        
        tracking.columns = ["chapter_type", "total_questions", "has_answers"]
        tracking["completion_rate"] = (tracking["has_answers"] / tracking["total_questions"] * 100).round(1)
        
        return tracking

    # --- Homeworks ---
    HOMEWORK_COLUMNS = ["homework_id", "title", "description", "created_by", "section_id", "subject", "due_date", "total_marks", "is_active", "created_at"]

    def get_homeworks(self):
        return self._sheet_to_df("Homeworks")

    def add_homework(self, homework_data):
        df = self.get_homeworks()
        if df.empty:
            df = pd.DataFrame(columns=self.HOMEWORK_COLUMNS)
        df = pd.concat([df, pd.DataFrame([homework_data])], ignore_index=True)
        self._df_to_sheet("Homeworks", df, self.HOMEWORK_COLUMNS)

    def update_homework(self, homework_id, updates):
        df = self.get_homeworks()
        idx = df[df.homework_id == homework_id].index
        if len(idx) > 0:
            for k, v in updates.items():
                df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("Homeworks", df, self.HOMEWORK_COLUMNS)

    # --- Homework Submissions ---
    HOMEWORK_SUBMISSION_COLUMNS = [
        "submission_id", "homework_id", "student_id", "student_name", "section_id",
        "image_data", "image_name", "submission_note", "status",
        "grade", "feedback", "submitted_at", "reviewed_by", "reviewed_at"
    ]

    def get_homework_submissions(self):
        return self._sheet_to_df("HomeworkSubmissions")

    def add_homework_submission(self, submission_data):
        df = self.get_homework_submissions()
        if df.empty:
            df = pd.DataFrame(columns=self.HOMEWORK_SUBMISSION_COLUMNS)
        df = pd.concat([df, pd.DataFrame([submission_data])], ignore_index=True)
        self._df_to_sheet("HomeworkSubmissions", df, self.HOMEWORK_SUBMISSION_COLUMNS)

    def update_homework_submission(self, submission_id, updates):
        df = self.get_homework_submissions()
        idx = df[df.submission_id == submission_id].index
        if len(idx) > 0:
            for k, v in updates.items():
                df.at[idx[0], k] = self._safe_str(v)
            self._df_to_sheet("HomeworkSubmissions", df, self.HOMEWORK_SUBMISSION_COLUMNS)

    # --- Notifications ---
    NOTIFICATION_COLUMNS = ["notification_id", "user_id", "title", "message", "notification_type", "is_read", "created_at"]

    def get_notifications(self, user_id=None):
        df = self._sheet_to_df("Notifications")
        if df.empty or not user_id:
            return df
        return df[df.user_id == user_id]

    def add_notification(self, notification_data):
        df = self._sheet_to_df("Notifications")
        if df.empty:
            df = pd.DataFrame(columns=self.NOTIFICATION_COLUMNS)
        df = pd.concat([df, pd.DataFrame([notification_data])], ignore_index=True)
        self._df_to_sheet("Notifications", df, self.NOTIFICATION_COLUMNS)

    def mark_notification_read(self, notification_id):
        df = self._sheet_to_df("Notifications")
        idx = df[df.notification_id == notification_id].index
        if len(idx) > 0:
            df.at[idx[0], "is_read"] = "True"
            self._df_to_sheet("Notifications", df, self.NOTIFICATION_COLUMNS)


# =============================================================================
# JWT & Session Helpers
# =============================================================================
def generate_token(user: dict, secret: str) -> str:
    payload = {
        "user_id": user.get("user_id", ""), "role": user.get("role", ""),
        "full_name": user.get("full_name", ""), "section_id": user.get("section_id", ""),
        "status": user.get("status", "active"),
        "exp": datetime.utcnow() + timedelta(hours=SESSION_TIMEOUT_HOURS)
    }
    return jwt.encode(payload, secret, algorithm="HS256")


def generate_quiz_token(quiz_id: str, student_id: str) -> str:
    payload = {"quiz_id": quiz_id, "student_id": student_id, "exp": datetime.utcnow() + timedelta(hours=48)}
    return jwt.encode(payload, QUIZ_JWT_SECRET, algorithm="HS256")


def verify_quiz_token(token: str):
    try:
        return jwt.decode(token, QUIZ_JWT_SECRET, algorithms=["HS256"])
    except Exception:
        return None


def verify_token(token: str, secret: str):
    try:
        return jwt.decode(token, secret, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None
    except Exception:
        return None


def init_session():
    defaults = {
        "authenticated": False, "user": None, "token": None, "last_login_time": None,
        "student_quiz": None, "student_quiz_started": False, "quiz_phase": "enter_name",
        "student_name": "", "student_id": "", "quiz_start_time": None, "quiz_end_time": None,
        "quiz_submit_time": None, "quiz_token": None, "quiz_answers": {}, "quiz_submitted": False,
        "last_score": 0, "menu_choice": "🏠 لوحة التحكم", "show_sidebar": True,
        "open_help_dialog": False, "current_attempt_id": None, "last_saved_answers_str": "",
        "quiz_questions": None, "show_review": False, "data_errors": [], "data_validated": False,
        "quiz_load_failures": 0,
        # Student Dashboard (تسجيل دخول الطالبات)
        "student_logged_in": False, "current_student": None,
        "student_dashboard_page": "🏠 الرئيسية",
        "sidebar_open": False,
        "selected_quiz_id": None, "selected_exam_id": None,
        "selected_assessment_type": None, "selected_assessment_id": None,
        "quiz_interface_started": False,
        "quiz_confirmation_id": None,
        "assessment_confirmation": None,
        "review_result_id": None, "review_result_type": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def logout(db=None):
    if db and st.session_state.user:
        try:
            db.add_log(st.session_state.user.get("user_id", ""), "تسجيل خروج", "تم تسجيل الخروج بنجاح")
        except Exception:
            pass
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()


def student_logout(db=None):
    """تسجيل خروج الطالبة مع الحفاظ على حالة النظام العام."""
    student = st.session_state.get("current_student") or {}
    student_id = student.get("student_id", "")
    if db and student_id:
        try:
            db.add_log(student_id, "تسجيل خروج طالبة", f"تم تسجيل خروج الطالبة: {student.get('full_name', '')}")
        except Exception:
            pass
    student_keys = [
        "student_logged_in", "current_student", "student_dashboard_page", "sidebar_open",
    ] + ASSESSMENT_SESSION_KEYS + [
        "quiz_question_index", "quiz_answers", "quiz_end_time", "quiz_attempt_id",
        "quiz_last_saved_answers", "quiz_confirm_finish", "quiz_start_time",
        "exam_question_index", "exam_answers", "exam_last_saved_answers", "exam_end_time",
        "exam_start_time", "exam_questions", "exam_shuffled_options", "exam_attempt_id",
        "exam_submitted", "exam_submit_time", "exam_result", "exam_confirm_finish",
        "exam_last_save_time", "quiz_questions_list", "quiz_questions_quiz_id",
        "assessment_questions_type",
    ]
    for key in student_keys:
        st.session_state.pop(key, None)
    st.rerun()


def send_telegram_message(message: str) -> bool:
    bot_token, chat_id = get_telegram_config()
    if not bot_token or not chat_id:
        return False
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {"chat_id": chat_id, "text": message, "parse_mode": "HTML"}
    try:
        response = requests.post(url, json=payload, timeout=10)
        return response.status_code == 200
    except Exception:
        return False


def send_telegram_photo(caption: str, file_bytes, filename: str) -> bool:
    bot_token, chat_id = get_telegram_config()
    if not bot_token or not chat_id:
        return False
    url = f"https://api.telegram.org/bot{bot_token}/sendPhoto"
    files = {'photo': (filename, file_bytes)}
    data = {'chat_id': chat_id, 'caption': caption, 'parse_mode': 'HTML'}
    try:
        response = requests.post(url, data=data, files=files, timeout=15)
        return response.status_code == 200
    except Exception:
        return False


# =============================================================================
# مركز المساعدة
# =============================================================================
@st.dialog("🆘 مركز المساعدة والدعم الفني", width="large")
def show_help_dialog():
    hdr_col1, hdr_col2 = st.columns([0.85, 0.15])
    with hdr_col1:
        st.markdown("<h3 style='text-align:center; color:#667eea; margin:0; padding-top:0.5rem;'>📬 تواصل معنا</h3>", unsafe_allow_html=True)
    with hdr_col2:
        if st.button("✕ إغلاق", key="help_dialog_close_btn", use_container_width=True):
            st.session_state.open_help_dialog = False
            st.rerun()
    contact_name, contact_whatsapp = get_support_config()
    if contact_whatsapp:
        st.info(f"📞 للدعم المباشر: {contact_name} - {contact_whatsapp}")
    st.markdown("---")
    with st.form("help_form_enhanced", clear_on_submit=False):
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("الاسم *", placeholder="أدخل اسمك الكامل")
            whatsapp = st.text_input("رقم الواتساب *", placeholder="01xxxxxxxxx")
        with col2:
            issue_type = st.selectbox("نوع المشكلة *", ["مشكلة تقنية", "مشكلة في البيانات", "طلب مساعدة", "اقتراح تحسين", "أخرى"])
            urgency = st.selectbox("الأولوية", ["عادي", "مستعجل", "طارئ جداً"], index=0)
        issue_desc = st.text_area("وصف المشكلة أو الطلب *", placeholder="اشرح المشكلة بالتفصيل...", height=150)
        uploaded_file = st.file_uploader("📎 إرفاق لقطة شاشة (اختياري)", type=["png", "jpg", "jpeg"])
        submitted = st.form_submit_button("🚀 إرسال الطلب", use_container_width=True)
        if submitted:
            if not name or not whatsapp or not issue_desc:
                st.error("⚠️ الرجاء ملء جميع الحقول المطلوبة")
            else:
                urgency_icon = {"عادي": "ℹ️", "مستعجل": "⚠️", "طارئ جداً": "🔴"}
                message = (
                    f"{urgency_icon.get(urgency, '')} بلاغ جديد من مركز المساعدة\n"
                    f"👤 الاسم: {name}\n📱 الواتساب: {whatsapp}\n📂 النوع: {issue_type}\n⚡ الأولوية: {urgency}\n📝 التفاصيل: {issue_desc}"
                )
                success = True
                if uploaded_file is not None:
                    if not send_telegram_photo(message, uploaded_file.getvalue(), uploaded_file.name):
                        success = False
                else:
                    if not send_telegram_message(message):
                        success = False
                if success:
                    st.success("✅ تم إرسال طلبك بنجاح! سنتواصل معك قريباً.")
                    st.balloons()
                else:
                    st.error("❌ فشل الإرسال، يرجى المحاولة لاحقاً أو التواصل مباشرة via الواتساب.")


# =============================================================================
# RBAC
# =============================================================================
VALID_ROLES = ["System Admin", "Father Account", "Service Manager", "Teacher", "Student"]
VALID_STATUSES = ["active", "inactive", "suspended"]
EVENT_TYPES = ["اجتماع", "خدمة", "رحلة", "احتفال"]
RSVP_STATUSES = ["سأحضر", "لن أحضر", "ربما"]

def require_role(required_roles):
    user = st.session_state.get("user")
    if not user:
        return False
    return user.get("role", "") in required_roles

def check_access(required_roles):
    if not require_role(required_roles):
        st.error("🚫 لا تملك الصلاحية للوصول إلى هذه الصفحة")
        st.stop()

def get_user_status(user_row):
    status = user_row.get("status", "active")
    if pd.isna(status) or str(status).strip() == "":
        return "active"
    return str(status).strip().lower()


# =============================================================================
# Helper Functions
# =============================================================================
def get_role_menu(role):
    menus = {
        "System Admin": [
            "🏠 لوحة التحكم", "🔔 الإشعارات", "👥 إدارة الأعضاء", "🏫 إدارة المراحل الدراسية", "📚 إدارة الفصول",
            "📋 الحضور", "💬 الافتقاد", "📷 ماسح QR",
            ADMIN_ASSESSMENTS_PAGE, "📊 التقارير والإحصائيات",
            "📅 إدارة الفعاليات", "📜 سجل العمليات", "🔒 تغيير كلمة المرور"
        ],
        "Father Account": ["🏠 لوحة التحكم", "🔔 الإشعارات", "👥 إدارة الأعضاء", ADMIN_ASSESSMENTS_PAGE, "📊 التقارير والإحصائيات", "🔒 تغيير كلمة المرور"],
        "Service Manager": [
            "🏠 لوحة التحكم", "🔔 الإشعارات", "👥 إدارة الأعضاء", "📋 الحضور", "💬 الافتقاد", "📷 ماسح QR",
            ADMIN_ASSESSMENTS_PAGE, "📅 إدارة الفعاليات", "📊 التقارير والإحصائيات", "🔒 تغيير كلمة المرور"
        ],
        "Teacher": [
            "🏠 لوحة التحكم", "🔔 الإشعارات", "👥 إدارة الأعضاء", "📋 الحضور", "💬 الافتقاد", "📷 ماسح QR",
            ADMIN_ASSESSMENTS_PAGE, "📅 إدارة الفعاليات", "🔒 تغيير كلمة المرور"
        ],
        "Student": ["🏠 لوحة التحكم", "🔔 الإشعارات", ADMIN_ASSESSMENTS_PAGE, "📅 إدارة الفعاليات", "🔒 تغيير كلمة المرور"]
    }
    return menus.get(role, [])


def get_sections_for_supervisor(db, user_id):
    stage_ids = db.get_stages_for_supervisor(user_id)
    if not stage_ids:
        return []
    sections = db.get_sections()
    if sections.empty:
        return []
    section_ids = sections[sections.stage_id.isin(stage_ids)]["section_id"].tolist()
    return section_ids


def filter_students_by_role(students, role, section_id, db=None, user_id=None):
    if role == "Teacher" and section_id:
        return students[students.section_id == section_id] if not students.empty and "section_id" in students.columns else pd.DataFrame()
    elif role == "Service Manager":
        if db and user_id:
            section_ids = get_sections_for_supervisor(db, user_id)
            if section_ids and not students.empty and "section_id" in students.columns:
                return students[students.section_id.isin(section_ids)]
        return students
    else:
        return students


def filter_attendance_by_role(attendance, role, section_id, db=None, user_id=None):
    if role == "Teacher" and section_id:
        return attendance[attendance.section_id == section_id] if not attendance.empty and "section_id" in attendance.columns else pd.DataFrame()
    elif role == "Service Manager":
        if db and user_id:
            section_ids = get_sections_for_supervisor(db, user_id)
            if section_ids and not attendance.empty and "section_id" in attendance.columns:
                return attendance[attendance.section_id.isin(section_ids)]
        return attendance
    return attendance


def clear_quiz_session_keys():
    quiz_keys = [
        "student_quiz", "student_quiz_started", "quiz_phase", "student_name",
        "student_id", "quiz_start_time", "quiz_end_time", "quiz_submit_time",
        "quiz_token", "quiz_answers", "quiz_submitted", "last_score",
        "current_attempt_id", "last_saved_answers_str", "quiz_questions", "show_review"
    ]
    for key in quiz_keys:
        if key in st.session_state:
            del st.session_state[key]


# =============================================================================
# Helper Functions
# =============================================================================
def get_initials(name: str) -> str:
    """Get initials from a name (first letters of first 2 words)."""
    if not name or not isinstance(name, str):
        return "؟"
    parts = name.strip().split()
    if len(parts) >= 2:
        return f"{parts[0][0]}{parts[1][0]}".upper()
    elif len(parts) == 1:
        return parts[0][:2].upper()
    return "؟"


def get_role_css_class(role: str) -> str:
    """Get CSS class for role badge."""
    role_map = {
        "System Admin": "admin",
        "Father Account": "priest",
        "Service Manager": "leader",
        "Teacher": "teacher",
        "Student": "student"
    }
    return role_map.get(role, "admin")


def get_status_css_class(status: str) -> str:
    """Get CSS class for status badge."""
    if not status:
        return "inactive"
    status_str = str(status).strip().lower()
    return "active" if status_str == "active" else "inactive"


# =============================================================================
# Validation
# =============================================================================
def validate_data_integrity(db):
    errors = []
    students = db.get_students()
    sections = db.get_sections()
    if not students.empty and not sections.empty:
        valid_sections = set(sections["section_id"].tolist())
        for _, row in students.iterrows():
            sid = row.get("section_id", "")
            if pd.isna(sid) or str(sid).strip() == "":
                errors.append(f"الطالبة {row.get('full_name', '')} ليس لديها فصل.")
            elif str(sid).strip() not in valid_sections:
                errors.append(f"الطالبة {row.get('full_name', '')} تنتمي لفصل غير موجود ({sid}).")
    return errors


def auto_fix_missing_sections(db):
    students = db.get_students()
    sections = db.get_sections()
    if students.empty:
        return False
    existing_ids = set(sections["section_id"].tolist()) if not sections.empty else set()
    student_section_ids = students["section_id"].dropna().unique().tolist()
    missing = [sid for sid in student_section_ids if sid and str(sid).strip() not in existing_ids]
    if missing:
        for sid in missing:
            db.add_section({"section_id": str(sid), "section_name": f"فصل (معرف {sid[:8]})"})
        return True
    return False


# =============================================================================
# Initialization & Login
# =============================================================================
def show_initialization(db):
    users = db.get_users()
    if users.empty:
        st.markdown("<div class='card'><h2 style='text-align:center;'>🔧 لا يوجد مستخدمون بعد</h2></div>", unsafe_allow_html=True)
        st.markdown("#### يرجى الضغط على الزر التالي لإنشاء مدير النظام الافتراضي:")
        if st.button("🛠️ تهيئة النظام وإنشاء المسؤول الأول", use_container_width=True, key="init_admin_btn"):
            admin_data = {
                "user_id": "admin-001", "username": "admin", "password": "admin123",
                "role": "System Admin", "full_name": "مدير النظام",
                "section_id": "", "phone": "0100000000", "email": "admin@church.com"
            }
            admin_data["password"] = hash_password(admin_data["password"])
            db.add_user(admin_data)
            st.success("✅ تم إنشاء مدير النظام بنجاح! الرجاء تسجيل الدخول باستخدام الحساب الافتراضي.")
            time.sleep(2)
            st.rerun()
        st.stop()


def show_login_page(db, jwt_secret):
    render_login_top_bar()
    # Hero banner for login page
    st.markdown(hero_header("نظام إدارة الكنيسة", "كنيسة الشهيدة دميانة"), unsafe_allow_html=True)
    show_initialization(db)
    tab1, tab2 = st.tabs(["🔐 دخول الخدام", "📝 تسجيل دخول الطالبات"])
    with tab1:
        with st.form("login_form"):
            username = st.text_input("اسم المستخدم", placeholder="أدخل اسم المستخدم").strip()
            password = st.text_input("كلمة المرور", type="password", placeholder="أدخل كلمة المرور").strip()
            if st.form_submit_button("تسجيل الدخول", use_container_width=True):
                if not username or not password:
                    st.error("يرجى إدخال اسم المستخدم وكلمة المرور")
                else:
                    with st.spinner("جاري التحقق..."):
                        users = db.get_users()
                        user_row = users[users.username == username]
                        if user_row.empty:
                            st.error("اسم المستخدم غير موجود")
                        else:
                            user = user_row.iloc[0].to_dict()
                            user_status = get_user_status(user)
                            if user_status != "active":
                                db.add_log(user.get("user_id", ""), "محاولة دخول فاشلة", f"الحساب غير نشط (الحالة: {user_status})")
                                st.error(f"🚫 هذا الحساب {user_status}. يرجى التواصل مع مسؤول النظام.")
                            elif verify_password(password, user.get("password", "")):
                                token = generate_token(user, jwt_secret)
                                st.session_state.token = token
                                st.session_state.user = user
                                st.session_state.authenticated = True
                                st.session_state.last_login_time = get_cairo_now().isoformat()
                                st.session_state.menu_choice = "🏠 لوحة التحكم"
                                st.session_state.show_sidebar = True
                                db.add_log(user["user_id"], "تسجيل دخول", "تم تسجيل الدخول بنجاح")
                                st.success("تم تسجيل الدخول بنجاح!")
                                time.sleep(1)
                                st.rerun()
                            else:
                                db.add_log(user.get("user_id", ""), "محاولة دخول فاشلة", "كلمة مرور خاطئة")
                                st.error("كلمة المرور غير صحيحة")
        with tab2:
            st.subheader("تسجيل دخول الطالبات")
            st.info("أدخلي كود الطالبة وكلمة المرور الخاصة بكِ للدخول إلى حسابك.")
            with st.form("student_login_form"):
                code = st.text_input("كود الطالبة", placeholder="مثال: STU000001").strip()
                passwd = st.text_input("كلمة مرور الطالبة", type="password", placeholder="").strip()
                if st.form_submit_button("تسجيل الدخول", use_container_width=True):
                    if not code or not passwd:
                        st.error("الرجاء إدخال كود الطالبة وكلمة المرور")
                    else:
                        with st.spinner("جاري التحقق..."):
                            # التأكد من وجود عمود profile_edit_used
                            try:
                                db.ensure_student_profile_edit_column()
                            except Exception:
                                pass
                            students = db.get_students()
                            student_match = students[
                                (students.student_code.astype(str).str.strip() == code) & 
                                (students.student_password.astype(str).str.strip() == passwd)
                            ]
                            if student_match.empty:
                                st.error("كود الطالبة أو كلمة المرور غير صحيحة")
                            else:
                                student = student_match.iloc[0].to_dict()
                                student_status = str(student.get("status", "active")).strip().lower()
                                if student_status != "active":
                                    st.error("🚫 هذا الحساب غير نشط. يرجى التواصل مع مسؤول النظام.")
                                else:
                                    st.session_state.student_logged_in = True
                                    st.session_state.current_student = student
                                    st.session_state.student_dashboard_page = "🏠 الرئيسية"
                                    st.session_state.sidebar_open = False
                                    st.session_state.selected_quiz_id = None
                                    st.session_state.quiz_interface_started = False
                                    db.add_log(student.get("student_id", ""), "تسجيل دخول طالبة", f"تم تسجيل دخول الطالبة: {student.get('full_name', '')}")
                                    st.success(f"مرحباً {student.get('full_name', '')}! تم تسجيل الدخول بنجاح.")
                                    time.sleep(1)
                                    st.rerun()
        return 0


def grade_attempt(db, quiz_id, answers_dict):
    """Grade a quiz attempt and return the score."""
    questions = db.get_quiz_questions(quiz_id)
    if questions.empty:
        return 0
    
    correct_count = 0
    for _, q_row in questions.iterrows():
        q = q_row.to_dict()
        correct = str(q.get("correct_answer", "")).strip().lower()
        student_ans = str(answers_dict.get(q.get("question_id", ""), "")).strip().lower()
        if correct == student_ans:
            correct_count += 1
    num_q = len(questions)
    score = round((correct_count / num_q) * 20, 1) if num_q > 0 else 0
    return score


def save_current_answers(db):
    if not st.session_state.current_attempt_id:
        return
    current_answers = json.dumps(st.session_state.quiz_answers, ensure_ascii=False)
    if current_answers != st.session_state.last_saved_answers_str:
        db.save_answers(st.session_state.current_attempt_id, st.session_state.quiz_answers)
        st.session_state.last_saved_answers_str = current_answers


def show_student_quiz(db):
    if st.session_state.quiz_phase in ["taking_quiz", "finished"]:
        if not st.session_state.get("quiz_token"):
            st.error("انتهت جلسة الاختبار. يرجى إعادة الدخول.")
            clear_quiz_session_keys()
            st.stop()
        token_data = verify_quiz_token(st.session_state.quiz_token)
        if token_data is None:
            st.error("انتهت صلاحية جلسة الاختبار. يرجى إعادة الدخول.")
            clear_quiz_session_keys()
            st.stop()

    quiz = st.session_state.student_quiz
    if st.session_state.quiz_phase == "enter_name":
        st.title(f"📝 {quiz.get('title', '')}")
        st.markdown(f"**عدد الأسئلة:** {quiz.get('num_questions', '')} | **الدرجة الكلية:** 20 | **الوقت:** {quiz.get('time_limit_minutes', '')} دقيقة")
        st.markdown("---")
        students_df = db.get_students()
        active_students = students_df[students_df["status"] == "active"] if not students_df.empty else pd.DataFrame()
        if active_students.empty:
            st.warning("لا توجد طالبات مسجلات حالياً. يرجى التواصل مع المسؤول.")
            st.stop()
        active_students = active_students.sort_values("full_name", key=lambda col: col.str.strip().str.lower())
        options_dict = dict(zip(active_students["student_id"], active_students["full_name"]))
        selected_id = st.selectbox(
            "اختر اسمك من القائمة", options=list(options_dict.keys()),
            format_func=lambda x: options_dict[x], index=None, placeholder="اختر اسمك..."
        )
        if selected_id is not None:
            student_row = active_students[active_students.student_id == selected_id].iloc[0]
            sec_id = student_row.get("section_id", "")
            sections_df = db.get_sections()
            section_name = ""
            if not sections_df.empty and sec_id:
                sec_match = sections_df[sections_df.section_id == sec_id]
                if not sec_match.empty:
                    section_name = sec_match.iloc[0].get("section_name", "")
            if not section_name:
                section_name = "لم يتم تعيين فصل"
            st.info(f"أنتِ في فصل: **{section_name}**")
        st.markdown("---")
        st.info("إذا لم تجد اسمك في القائمة، يرجى التواصل مع مشرف الخدمة لإضافتك.")
        if selected_id is not None:
            existing = db.get_quiz_results(quiz.get("quiz_id"))
            if not existing.empty:
                student_attempts = existing[existing["student_id"] == selected_id]
                if not student_attempts.empty:
                    attempt = student_attempts.iloc[0]
                    if attempt.get("status") == "started":
                        answers_str = attempt.get("answers", "{}")
                        try:
                            saved_answers = json.loads(answers_str) if answers_str else {}
                        except Exception:
                            saved_answers = {}
                        score = grade_attempt(db, quiz["quiz_id"], saved_answers)
                        db.submit_quiz_attempt(attempt["result_id"], score, json.dumps(saved_answers, ensure_ascii=False))
                        st.warning("تم تسليم محاولتك السابقة تلقائياً بناءً على ما قمت بحفظه.")
                        st.session_state.last_score = score
                        st.session_state.quiz_submit_time = get_cairo_now()
                        st.session_state.quiz_phase = "finished"
                        st.session_state.quiz_submitted = True
                        st.session_state.quiz_token = generate_quiz_token(quiz["quiz_id"], selected_id)
                        st.rerun()
                    else:
                        st.error("لقد قمت بتسليم هذا الاختبار بالفعل. لا يمكنك الدخول مرة أخرى.")
                        st.stop()
        if st.button("بدء الاختبار", use_container_width=True, disabled=(selected_id is None), key="start_quiz_btn"):
            selected_student = active_students[active_students["student_id"] == selected_id].iloc[0].to_dict()
            st.session_state.student_name = selected_student["full_name"]
            st.session_state.student_id = selected_id
            st.session_state.quiz_start_time = get_cairo_now()
            time_limit_seconds = int(quiz.get('time_limit_minutes', 15)) * 60
            st.session_state.quiz_end_time = st.session_state.quiz_start_time + timedelta(seconds=time_limit_seconds)
            attempt_id = db.start_quiz_attempt(quiz["quiz_id"], selected_id, st.session_state.student_name)
            st.session_state.current_attempt_id = attempt_id
            st.session_state.quiz_answers = {}
            st.session_state.last_saved_answers_str = ""
            st.session_state.quiz_questions = None
            st.session_state.show_review = False
            st.session_state.quiz_load_failures = 0
            st.session_state.quiz_token = generate_quiz_token(quiz["quiz_id"], selected_id)
            st.session_state.quiz_phase = "taking_quiz"
            st.rerun()
        return

    elif st.session_state.quiz_phase == "taking_quiz":
        now = get_cairo_now()
        if now > st.session_state.quiz_end_time:
            st.warning("انتهى الوقت المخصص للامتحان. جاري تسليم إجاباتك تلقائياً...")
            score = grade_attempt(db, quiz["quiz_id"], st.session_state.quiz_answers)
            answers_json = json.dumps(st.session_state.quiz_answers, ensure_ascii=False)
            db.submit_quiz_attempt(st.session_state.current_attempt_id, score, answers_json)
            st.session_state.quiz_submitted = True
            st.session_state.last_score = score
            st.session_state.quiz_submit_time = now
            st.session_state.quiz_phase = "finished"
            st.rerun()

        if not st.session_state.get("quiz_questions"):
            try:
                questions_df = db.get_quiz_questions(quiz["quiz_id"])
                if questions_df.empty:
                    st.warning("لا توجد أسئلة في هذا الاختبار بعد.")
                    return
                st.session_state.quiz_questions = questions_df.to_dict('records')
            except Exception:
                st.error("تعذر تحميل الأسئلة.")
                return
        else:
            questions_df = pd.DataFrame(st.session_state.quiz_questions)

        end_time_iso = st.session_state.quiz_end_time.isoformat()
        countdown_html = f"""
        <!DOCTYPE html>
        <html><head><meta charset="utf-8"><style>
        body {{ font-family: 'Cairo', sans-serif; margin: 0; padding: 0; display: flex; justify-content: center; align-items: center; height: 100%; background: transparent; }}
        #timer {{ font-size: 1.8rem; font-weight: bold; padding: 1rem 2rem; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 15px; box-shadow: 0 4px 12px rgba(102,126,234,0.4); text-align: center; }}
        </style></head><body>
        <div id="timer">⏳ الوقت المتبقي: <span id="time"></span></div>
        <script>
        var endTime = new Date("{end_time_iso}").getTime();
        function update() {{ var now = new Date().getTime(); var dist = endTime - now;
            if (dist <= 0) {{ document.getElementById('time').innerHTML = "00:00"; parent.postMessage({{type: "QUIZ_TIME_UP"}}, "*"); clearInterval(intervalId); return; }}
            var mins = Math.floor((dist % (1000*60*60)) / (1000*60)); var secs = Math.floor((dist % (1000*60)) / 1000);
            document.getElementById('time').innerHTML = (mins<10?'0'+mins:mins) + ":" + (secs<10?'0'+secs:secs); }}
        update(); var intervalId = setInterval(update, 1000);
        </script></body></html>
        """
        st.components.v1.html(countdown_html, height=80, scrolling=False)
        st.title(f"📝 {quiz.get('title', '')}")
        st.markdown(f"الطالبة: **{st.session_state.student_name}** | الدرجة الكلية: 20")
        st.markdown("---")
        for idx, row in questions_df.iterrows():
            q = row.to_dict()
            q_id = q.get("question_id", "")
            st.markdown(f"**سؤال {idx+1}:** {q.get('question_text', '')}")
            q_type = q.get("question_type", "")
            prev_answer = st.session_state.quiz_answers.get(q_id, "")
            if q_type in ["اختيار من متعدد", "صح وخطأ"]:
                options = [q.get("option1", ""), q.get("option2", ""), q.get("option3", ""), q.get("option4", "")] if q_type == "اختيار من متعدد" else ["صح", "خطأ"]
                options = [opt for opt in options if opt and str(opt).strip()]
                if options:
                    current_index = options.index(prev_answer) if prev_answer in options else None
                    ans = st.radio("اختر الإجابة", options, key=f"q_{q_id}", index=current_index)
                    new_answer = ans if ans else ""
            else:
                new_answer = st.text_input("الإجابة", key=f"q_{q_id}", value=prev_answer)
            if new_answer != prev_answer:
                st.session_state.quiz_answers[q_id] = new_answer
                save_current_answers(db)
            st.markdown("---")
        if st.button("تسليم الاختبار", use_container_width=True, key="submit_quiz_btn"):
            score = grade_attempt(db, quiz["quiz_id"], st.session_state.quiz_answers)
            answers_json = json.dumps(st.session_state.quiz_answers, ensure_ascii=False)
            db.submit_quiz_attempt(st.session_state.current_attempt_id, score, answers_json)
            st.session_state.quiz_submitted = True
            st.session_state.last_score = score
            st.session_state.quiz_submit_time = get_cairo_now()
            st.session_state.quiz_phase = "finished"
            st.rerun()
        return

    elif st.session_state.quiz_phase == "finished":
        if not st.session_state.get("show_review", False):
            st.success("تم تسليم الاختبار بنجاح!")
            score = st.session_state.last_score
            score_display = int(score) if score.is_integer() else score
            st.info(f"نتيجتك: {score_display}/20")
            st.markdown("---")
            st.markdown("#### ⏱️ معلومات الوقت")
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                st.markdown("**بداية الامتحان:**")
                st.markdown(format_cairo_time(st.session_state.quiz_start_time))
            with col_t2:
                st.markdown("**نهاية الامتحان (التسليم):**")
                st.markdown(format_cairo_time(st.session_state.quiz_submit_time))
            col_btn, _ = st.columns([2, 3])
            if col_btn.button("عرض الإجابات والأخطاء", use_container_width=True, key="show_review_btn"):
                st.session_state.show_review = True
                st.rerun()
            if st.button("إنهاء والعودة إلى الرئيسية", use_container_width=True, key="finish_no_review_btn"):
                clear_quiz_session_keys()
                st.rerun()
        else:
            st.markdown("## مراجعة الإجابات")
            if not st.session_state.get("quiz_questions"):
                questions_df = db.get_quiz_questions(quiz["quiz_id"])
                if questions_df.empty:
                    st.warning("لا يمكن تحميل الأسئلة للمراجعة.")
                else:
                    st.session_state.quiz_questions = questions_df.to_dict('records')
            if st.session_state.get("quiz_questions"):
                questions_df = pd.DataFrame(st.session_state.quiz_questions)
                student_answers = st.session_state.quiz_answers
                for idx, row in questions_df.iterrows():
                    q = row.to_dict()
                    qid = q.get("question_id", "")
                    correct = str(q.get("correct_answer", "")).strip().lower()
                    student_ans = str(student_answers.get(qid, "")).strip().lower()
                    is_correct = (correct == student_ans)
                    st.markdown(f"**سؤال {idx+1}:** {q.get('question_text', '')}")
                    col1, col2 = st.columns(2)
                    col1.markdown(f"📝 إجابتك: {student_ans if student_ans else 'لم تجب'}")
                    col2.markdown(f"✅ الإجابة الصحيحة: {correct}")
                    if is_correct:
                        st.success("✔️ صحيح")
                    else:
                        st.error("❌ خطأ")
                    st.markdown("---")
                if st.button("إنهاء المراجعة والعودة إلى الرئيسية", use_container_width=True, key="finish_review_btn"):
                    clear_quiz_session_keys()
                    st.rerun()
        return


def show_student_assessment_interface(db):
    """Unified assessment taking experience for quizzes and exams."""
    show_unified_assessment_taking_interface(db)


def show_unified_assessment_taking_interface(db):
    """Single assessment UI — shared timer, save, submit, and navigation."""
    if not st.session_state.get("student_logged_in", False):
        st.error("يجب تسجيل الدخول أولاً.")
        st.session_state.quiz_interface_started = False
        st.rerun()
        return

    student = st.session_state.get("current_student")
    a_type = st.session_state.get("selected_assessment_type")
    a_id = st.session_state.get("selected_assessment_id")
    if not a_type:
        if st.session_state.get("selected_exam_id"):
            a_type, a_id = "exam", st.session_state.selected_exam_id
        elif st.session_state.get("selected_quiz_id"):
            a_type, a_id = "quiz", st.session_state.selected_quiz_id
    if not a_type or not a_id or not student:
        st.session_state.quiz_interface_started = False
        st.rerun()
        return

    st.session_state.selected_assessment_type = a_type
    st.session_state.selected_assessment_id = a_id
    student_id = student.get("student_id", "")
    student_name = student.get("full_name", "طالبة")

    assessment_row = get_assessment_record(db, a_type, a_id)
    if not assessment_row:
        st.error("لم يتم العثور على الاختبار.")
        st.session_state.quiz_interface_started = False
        st.rerun()
        return

    can_access, deny_reason = student_can_access_assessment(db, student, a_type, a_id)
    if not can_access:
        st.warning(deny_reason or "غير مصرح بالدخول إلى هذا الاختبار.")
        st.session_state.quiz_interface_started = False
        if st.button("العودة إلى المسابقات والاختبارات", use_container_width=True):
            clear_assessment_session_state()
            st.session_state.student_dashboard_page = STUDENT_ASSESSMENTS_PAGE
            st.rerun()
        return

    status = _get_assessment_attempt_status(db, student_id, a_type, a_id)
    if status == "submitted":
        st.warning("⚠️ لقد قمتِ بإنجاز هذا الاختبار بالفعل.")
        st.session_state.quiz_interface_started = False
        if st.button("العودة إلى المسابقات والاختبارات", use_container_width=True):
            clear_assessment_session_state()
            st.session_state.student_dashboard_page = STUDENT_ASSESSMENTS_PAGE
            st.rerun()
        return

    assessment_title = assessment_row.get("title", "اختبار")
    duration_minutes = get_assessment_duration_minutes(assessment_row, a_type)

    defaults = {
        "assessment_question_index": 0,
        "assessment_answers": {},
        "assessment_last_saved_answers": "",
        "assessment_attempt_id": None,
        "assessment_submitted": False,
        "assessment_result": None,
        "assessment_confirm_finish": False,
        "assessment_questions": None,
        "assessment_shuffled_options": {},
        "assessment_questions_id": None,
        "assessment_end_time": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    if (
        st.session_state.assessment_questions is None
        or st.session_state.get("assessment_questions_id") != a_id
        or st.session_state.get("assessment_questions_type") != a_type
    ):
        questions, shuffled_options = load_assessment_questions(db, a_type, a_id)
        if not questions:
            st.warning("لا توجد أسئلة في هذا الاختبار.")
            if st.button("🔙 العودة", use_container_width=True):
                st.session_state.quiz_interface_started = False
                st.rerun()
            return
        st.session_state.assessment_questions = questions
        st.session_state.assessment_shuffled_options = shuffled_options
        st.session_state.assessment_questions_id = a_id
        st.session_state.assessment_questions_type = a_type

    questions = st.session_state.assessment_questions
    shuffled_options = st.session_state.assessment_shuffled_options
    total_questions = len(questions)

    if st.session_state.assessment_attempt_id is None:
        existing_id, saved_answers = get_in_progress_attempt(db, student_id, a_type, a_id)
        if existing_id:
            st.session_state.assessment_attempt_id = existing_id
            if saved_answers and not st.session_state.assessment_answers:
                st.session_state.assessment_answers = saved_answers
        else:
            if a_type == "exam":
                attempt_id = db.start_exam_attempt(a_id, student_id, student_name)
            else:
                attempt_id = db.start_quiz_attempt(a_id, student_id, student_name)
            st.session_state.assessment_attempt_id = attempt_id

    attempt_id = st.session_state.assessment_attempt_id
    end_time = get_attempt_deadline(db, a_type, a_id, attempt_id)
    st.session_state.assessment_end_time = end_time

    def auto_save():
        if not attempt_id:
            return
        answers_json = json.dumps(st.session_state.assessment_answers, ensure_ascii=False)
        if answers_json != st.session_state.assessment_last_saved_answers:
            save_assessment_answers(db, a_type, attempt_id, st.session_state.assessment_answers)
            st.session_state.assessment_last_saved_answers = answers_json

    def submit_internal(auto=False):
        if not attempt_id or st.session_state.assessment_submitted:
            return
        auto_save()
        result = submit_assessment_attempt(
            db, a_type, a_id, attempt_id, st.session_state.assessment_answers, auto=auto
        )
        st.session_state.assessment_submitted = True
        st.session_state.assessment_result = result

    now = get_cairo_now()
    if end_time is not None and now >= end_time and not st.session_state.assessment_submitted:
        submit_internal(auto=True)
        st.rerun()

    if st.session_state.assessment_submitted:
        result = st.session_state.assessment_result or {}
        st.success("✅ تم تسليم الاختبار بنجاح!")
        st.info(f"**درجتك:** {result.get('score', 0)} / {result.get('total_marks', 0)}")
        if st.button("🔙 العودة إلى المسابقات والاختبارات", use_container_width=True):
            clear_assessment_session_state()
            st.session_state.student_dashboard_page = STUDENT_ASSESSMENTS_PAGE
            st.rerun()
        return

    st.markdown(
        hero_header(assessment_title, f"⏱️ {duration_minutes} دقيقة | 👤 {student_name}"),
        unsafe_allow_html=True,
    )

    if end_time is not None:
        st.components.v1.html(render_assessment_timer_html(end_time.isoformat()), height=85, scrolling=False)

    current_index = st.session_state.assessment_question_index
    if total_questions > 0:
        st.progress((current_index + 1) / total_questions)
        st.caption(f"📋 السؤال {current_index + 1} من {total_questions}")

    q = questions[current_index]
    q_id = q.get("question_id", "")
    q_text = q.get("question_text", "")
    st.markdown(f"**سؤال {current_index + 1}:** {q_text}")

    options = shuffled_options.get(q_id, [])
    prev_answer = st.session_state.assessment_answers.get(q_id, "")
    if options:
        index = options.index(prev_answer) if prev_answer in options else None
        ans = st.radio("اختر الإجابة", options, index=index, key=f"assess_q_{a_type}_{q_id}")
        if ans is not None and st.session_state.assessment_answers.get(q_id) != ans:
            st.session_state.assessment_answers[q_id] = ans
            auto_save()
    else:
        ans_text = st.text_input("الإجابة", value=prev_answer, key=f"assess_q_{a_type}_{q_id}")
        if ans_text != prev_answer:
            st.session_state.assessment_answers[q_id] = ans_text
            auto_save()

    st.markdown("---")
    col_prev, col_mid, col_next = st.columns([1, 2, 1])
    with col_prev:
        if st.button("⬅️ السابق", use_container_width=True, disabled=current_index == 0, key="assess_prev"):
            st.session_state.assessment_question_index = max(0, current_index - 1)
            st.rerun()
    with col_mid:
        if st.button("🚨 تسليم الاختبار", use_container_width=True, key="assess_finish"):
            st.session_state.assessment_confirm_finish = True
            st.rerun()
    with col_next:
        if st.button("التالي ➡️", use_container_width=True, disabled=current_index >= total_questions - 1, key="assess_next"):
            st.session_state.assessment_question_index = min(total_questions - 1, current_index + 1)
            st.rerun()

    if st.session_state.get("assessment_confirm_finish", False):
        st.warning("⚠️ هل أنت متأكدة من تسليم الاختبار؟ لن تتمكني من تعديل إجاباتك بعد التسليم.")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("✅ نعم، تسليم", use_container_width=True, key="assess_yes"):
                submit_internal(auto=False)
                st.session_state.assessment_confirm_finish = False
                st.rerun()
        with c2:
            if st.button("❌ تراجع", use_container_width=True, key="assess_no"):
                st.session_state.assessment_confirm_finish = False
                st.rerun()


def show_student_exam_interface_for_student(db):
    """Legacy wrapper — redirects to unified assessment interface."""
    show_unified_assessment_taking_interface(db)


def show_student_quiz_interface(db):
    """Legacy wrapper — redirects to unified assessment interface."""
    show_unified_assessment_taking_interface(db)


# =============================================================================
# Student Portal Helpers
# =============================================================================
STUDENT_MENU_ITEMS = [
    "🏠 الرئيسية",
    "👤 ملفي الشخصي",
    STUDENT_ASSESSMENTS_PAGE,
    "🔔 الإشعارات",
    "🚪 تسجيل الخروج",
]

ASSESSMENT_SESSION_KEYS = [
    "selected_quiz_id", "selected_exam_id", "selected_assessment_type", "selected_assessment_id",
    "quiz_interface_started", "quiz_confirmation_id", "assessment_confirmation",
    "review_result_id", "review_result_type", "assessment_tab",
    "assessment_attempt_id", "assessment_questions", "assessment_shuffled_options",
    "assessment_answers", "assessment_question_index", "assessment_last_saved_answers",
    "assessment_submitted", "assessment_result", "assessment_confirm_finish",
    "assessment_questions_id", "assessment_end_time",
]


def is_availability_true(value):
    """تحقق آمن من قيمة التوفر (true/True/TRUE/'true'/1/yes)."""
    if value is None:
        return False
    if isinstance(value, float) and pd.isna(value):
        return False
    return str(value).strip().lower() in ("true", "1", "yes", "نعم", "active")


def get_availability_column(df):
    """إرجاع اسم عمود التوفر من Quizzes أو Exams."""
    if df is None or df.empty:
        return None
    for col in ("is_active", "availability", "is_available"):
        if col in df.columns:
            return col
    return None


def filter_available_rows(df):
    """فلترة الصفوف المتاحة بناءً على عمود التوفر."""
    if df is None or df.empty:
        return pd.DataFrame()
    col = get_availability_column(df)
    if col is None:
        return df.copy()
    mask = df[col].apply(is_availability_true)
    return df[mask].copy()


def get_assessment_question_count(db, assessment_type, assessment_id):
    """عدد أسئلة الاختبار أو الامتحان من QuizQuestions."""
    qdf = db.get_quiz_questions(assessment_id)
    return len(qdf) if not qdf.empty else 0


def build_unified_assessments(db):
    """قائمة موحدة من Quizzes فقط مع type مدمج."""
    items = []
    quizzes = filter_available_rows(db.get_quizzes())
    if not quizzes.empty:
        for _, row in quizzes.iterrows():
            qid = str(row.get("quiz_id", "")).strip()
            if not qid:
                continue
            a_type = "exam" if str(row.get("assessment_type", "quiz")).strip() == "exam" else "quiz"
            num_q = row.get("num_questions", "")
            if not num_q or str(num_q).strip() == "":
                num_q = get_assessment_question_count(db, a_type, qid)
            time_limit = row.get("duration_minutes") if a_type == "exam" else row.get("time_limit_minutes")
            items.append({
                "assessment_type": a_type,
                "assessment_id": qid,
                "title": row.get("title", "اختبار"),
                "description": row.get("description", ""),
                "type_label": "امتحان" if a_type == "exam" else "اختبار",
                "num_questions": num_q,
                "total_marks": row.get("total_marks", "20"),
                "time_limit_minutes": time_limit or row.get("time_limit_minutes", ""),
            })
    return items


def normalize_student_dashboard_page(page):
    """Map legacy student assessment pages to the unified page."""
    return LEGACY_STUDENT_ASSESSMENT_PAGES.get(page, page)


def normalize_admin_menu_choice(choice):
    """Map legacy admin exam menu to unified assessments page."""
    if choice == LEGACY_ADMIN_ASSESSMENTS_PAGE:
        return ADMIN_ASSESSMENTS_PAGE
    return choice


def _parse_assessment_datetime(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        dt = pd.to_datetime(value)
        if pd.isna(dt):
            return None
        if hasattr(dt, "to_pydatetime"):
            dt = dt.to_pydatetime()
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=CAIRO_TZ)
        else:
            dt = dt.astimezone(CAIRO_TZ)
        return dt
    except Exception:
        return None


def get_assessment_record(db, assessment_type, assessment_id):
    """Load assessment row from unified Quizzes table."""
    df = db.get_quizzes()
    id_col = "quiz_id"
    if df.empty or id_col not in df.columns:
        return None
    match = df[df[id_col].astype(str) == str(assessment_id)]
    if match.empty:
        return None
    return match.iloc[0].to_dict()


def get_assessment_duration_minutes(assessment_row, assessment_type):
    if not assessment_row:
        return 15 if assessment_type == "quiz" else 30
    raw = assessment_row.get("duration_minutes", assessment_row.get("time_limit_minutes", 30 if assessment_type == "exam" else 15))
    try:
        return max(1, int(float(raw or (15 if assessment_type == "quiz" else 30))))
    except (TypeError, ValueError):
        return 15 if assessment_type == "quiz" else 30


def student_can_access_assessment(db, student, assessment_type, assessment_id):
    """Backend eligibility: availability, dates, section/stage, publish rules."""
    row = get_assessment_record(db, assessment_type, assessment_id)
    if not row:
        return False, "لم يتم العثور على الاختبار."

    avail_col = get_availability_column(pd.DataFrame([row]))
    if avail_col and not is_availability_true(row.get(avail_col)):
        return False, "هذا الاختبار غير متاح حالياً."

    student_section = str(student.get("section_id", "")).strip()
    now = get_cairo_now()

    if assessment_type == "quiz":
        expiry = _parse_assessment_datetime(row.get("expiry_date"))
        if expiry and now.date() > expiry.date():
            return False, "انتهت صلاحية هذه المسابقة."
        quiz_section = str(row.get("section_id", "")).strip()
        if quiz_section and student_section and quiz_section != student_section:
            return False, "هذه المسابقة غير مخصصة لفصلك."
        return True, ""

    if str(row.get("is_published", "False")).strip() != "True":
        return False, "الامتحان غير منشور بعد."

    start_dt = _parse_assessment_datetime(row.get("start_date") or row.get("exam_date"))
    end_dt = _parse_assessment_datetime(row.get("end_date"))
    if start_dt and now < start_dt:
        return False, "لم يبدأ موعد هذا الامتحان بعد."
    if end_dt and now.date() > end_dt.date():
        return False, "انتهى موعد هذا الامتحان."

    exam_section = str(row.get("section_id", "")).strip()
    if exam_section and student_section and exam_section != student_section:
        return False, "هذا الامتحان غير مخصص لفصلك."

    if not exam_section and row.get("stage_id"):
        sections = db.get_sections()
        if not sections.empty and student_section:
            sec_match = sections[sections["section_id"].astype(str) == student_section]
            if not sec_match.empty:
                student_stage = str(sec_match.iloc[0].get("stage_id", "")).strip()
                if student_stage and str(row.get("stage_id", "")).strip() != student_stage:
                    return False, "هذا الامتحان غير مخصص لمرحلتك."
    return True, ""


def build_unified_assessments_for_student(db, student):
    """Assessments visible to the logged-in student after eligibility rules."""
    eligible = []
    for item in build_unified_assessments(db):
        ok, _reason = student_can_access_assessment(
            db, student, item["assessment_type"], item["assessment_id"]
        )
        if ok:
            eligible.append(item)
    return eligible


def get_assessment_attempt_row(db, assessment_type, attempt_id):
    df = db.get_quiz_results()
    if df.empty or "result_id" not in df.columns:
        return None
    match = df[df["result_id"].astype(str) == str(attempt_id)]
    if match.empty:
        return None
    return match.iloc[0].to_dict()


def get_attempt_deadline(db, assessment_type, assessment_id, attempt_id):
    """Server-side deadline from stored start_time + configured duration."""
    attempt = get_assessment_attempt_row(db, assessment_type, attempt_id)
    if not attempt:
        return None
    start_time = _parse_assessment_datetime(attempt.get("start_time"))
    if not start_time:
        return None
    assessment_row = get_assessment_record(db, assessment_type, assessment_id)
    duration = get_assessment_duration_minutes(assessment_row, assessment_type)
    return start_time + timedelta(minutes=duration)


def grade_assessment_attempt(db, assessment_type, assessment_id, answers_dict):
    """Server-side grading — never trust client-provided scores."""
    if assessment_type == "exam":
        score, total_marks, correct, wrong = db.grade_exam_attempt(assessment_id, answers_dict)
        return {
            "score": score, "total_marks": total_marks,
            "correct_count": correct, "wrong_count": wrong,
        }
    score = grade_attempt(db, assessment_id, answers_dict)
    assessment_row = get_assessment_record(db, assessment_type, assessment_id)
    total_marks = assessment_row.get("total_marks", "20") if assessment_row else "20"
    try:
        total_marks = float(total_marks)
    except (TypeError, ValueError):
        total_marks = 20
    return {"score": score, "total_marks": total_marks, "correct_count": None, "wrong_count": None}


def save_assessment_answers(db, assessment_type, attempt_id, answers_dict):
    if assessment_type == "exam":
        db.save_exam_answers(attempt_id, answers_dict)
    else:
        db.save_answers(attempt_id, answers_dict)


def submit_assessment_attempt(db, assessment_type, assessment_id, attempt_id, answers_dict, auto=False):
    """Grade on server and persist submission."""
    graded = grade_assessment_attempt(db, assessment_type, assessment_id, answers_dict)
    answers_json = json.dumps(answers_dict, ensure_ascii=False)
    if assessment_type == "exam":
        db.submit_exam_attempt(attempt_id, graded["score"], answers_json)
    else:
        db.submit_quiz_attempt(attempt_id, graded["score"], answers_json)
    graded["auto_submitted"] = auto
    return graded


def load_assessment_questions(db, assessment_type, assessment_id):
    """Load questions with optional shuffle for quiz/exam engines."""
    if assessment_type == "exam":
        questions_df = db.shuffle_questions(assessment_id)
    else:
        questions_df = db.get_quiz_questions(assessment_id)
    if questions_df is None or questions_df.empty:
        return [], {}
    questions = questions_df.to_dict("records")
    random.shuffle(questions)
    shuffled_options = {}
    for q in questions:
        q_id = q.get("question_id", "")
        q_type = q.get("question_type", "")
        if str(q_type).strip() == "صح وخطأ":
            shuffled_options[q_id] = ["صح", "خطأ"]
        else:
            opts = [str(q.get(f"option{i}", "")).strip() for i in range(1, 5)]
            opts = [o for o in opts if o]
            random.shuffle(opts)
            shuffled_options[q_id] = opts
    return questions, shuffled_options


def clear_assessment_session_state():
    for key in ASSESSMENT_SESSION_KEYS:
        st.session_state.pop(key, None)


def render_assessment_timer_html(end_time_iso):
    return f"""
    <!DOCTYPE html><html><head><meta charset="utf-8"><style>
    body {{ font-family: 'Cairo', sans-serif; margin: 0; padding: 0; display: flex; justify-content: center; }}
    #timer {{ font-size: 1.6rem; font-weight: bold; padding: 0.8rem 2rem; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 15px; box-shadow: 0 4px 12px rgba(102,126,234,0.4); text-align: center; }}
    </style></head><body>
    <div id="timer">⏳ الوقت المتبقي: <span id="time"></span></div>
    <script>
    var endTime = new Date("{end_time_iso}").getTime();
    function update() {{ var now = new Date().getTime(); var dist = endTime - now;
        if (dist <= 0) {{ document.getElementById('time').innerHTML = "00:00"; setTimeout(function() {{ window.parent.location.reload(); }}, 800); return; }}
        var mins = Math.floor((dist % (1000*60*60)) / (1000*60)); var secs = Math.floor((dist % (1000*60)) / 1000);
        document.getElementById('time').innerHTML = (mins<10?'0'+mins:mins) + ":" + (secs<10?'0'+secs:secs); }}
    update(); setInterval(update, 1000);
    </script></body></html>
    """


def get_in_progress_attempt(db, student_id, assessment_type, assessment_id):
    """Return (attempt_id, saved_answers) for a started-but-not-submitted attempt."""
    results = db.get_quiz_results()
    id_col = "quiz_id"
    if results.empty or id_col not in results.columns or "student_id" not in results.columns:
        return None, {}
    attempts = results[
        (results[id_col].astype(str) == str(assessment_id))
        & (results["student_id"].astype(str) == str(student_id))
    ]
    if attempts.empty or "status" not in attempts.columns:
        return None, {}
    started = attempts[attempts["status"] == "started"]
    if started.empty:
        return None, {}
    row = started.iloc[0].to_dict()
    try:
        saved = json.loads(row.get("answers", "{}") or "{}")
    except (json.JSONDecodeError, TypeError):
        saved = {}
    return row.get("result_id"), saved


def get_student_submitted_results(db, student_id):
    """جمع نتائج المسابقات والامتحانات المسلّمة للطالبة."""
    quiz_results = db.get_quiz_results()
    if quiz_results.empty or "student_id" not in quiz_results.columns:
        return pd.DataFrame(), pd.DataFrame()
    quizzes = db.get_quizzes()
    type_map = {}
    if not quizzes.empty and "quiz_id" in quizzes.columns:
        type_map = quizzes.set_index("quiz_id")["assessment_type"].astype(str).to_dict()
    all_student = quiz_results[quiz_results["student_id"] == student_id].copy()
    if "status" in all_student.columns:
        all_student = all_student[all_student["status"] == "submitted"]
    all_student["assessment_type"] = all_student["quiz_id"].astype(str).map(
        lambda qid: "exam" if str(type_map.get(qid, "quiz")).strip() == "exam" else "quiz"
    )
    student_quiz = all_student[all_student["assessment_type"] == "quiz"].copy()
    student_exam = all_student[all_student["assessment_type"] == "exam"].copy()
    if not student_exam.empty:
        student_exam["exam_id"] = student_exam["quiz_id"]
    return student_quiz, student_exam


def verify_student_owns_result(db, student_id, result_id, result_type):
    """التحقق من أن النتيجة تخص الطالبة الحالية."""
    results_df = db.get_quiz_results()
    if results_df.empty or "result_id" not in results_df.columns:
        return None
    attempt_df = results_df[results_df["result_id"] == result_id]
    if attempt_df.empty:
        return None
    attempt = attempt_df.iloc[0].to_dict()
    if str(attempt.get("student_id", "")).strip() != str(student_id).strip():
        return None
    assessment = get_assessment_record(db, result_type, attempt.get("quiz_id", ""))
    if result_type == "exam" and str((assessment or {}).get("assessment_type", "quiz")).strip() != "exam":
        return None
    if result_type == "quiz" and str((assessment or {}).get("assessment_type", "quiz")).strip() == "exam":
        return None
    return attempt


def inject_student_sidebar_css(sidebar_open):
    """CSS للقائمة الجانبية الكاملة للطالبة."""
    if sidebar_open:
        st.markdown("""
        <style>
        section[data-testid="stSidebar"] {
            position: fixed !important;
            top: 0 !important;
            right: 0 !important;
            left: auto !important;
            width: 100vw !important;
            height: 100vh !important;
            max-width: 100vw !important;
            max-height: 100vh !important;
            z-index: 999999 !important;
            border-radius: 0 !important;
            overflow-y: auto !important;
            overflow-x: hidden !important;
            background: #ffffff !important;
            transform: translateX(0) !important;
            direction: rtl !important;
        }
        section[data-testid="stSidebar"] > div {
            width: 100% !important;
            max-width: 100vw !important;
            direction: rtl !important;
        }
        [data-testid="collapsedControl"],
        button[kind="header"],
        [data-testid="stSidebarCollapseButton"] {
            display: none !important;
            visibility: hidden !important;
        }
        </style>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <style>
        section[data-testid="stSidebar"] {
            display: none !important;
            transform: translateX(100%) !important;
        }
        [data-testid="collapsedControl"],
        button[kind="header"],
        [data-testid="stSidebarCollapseButton"] {
            display: none !important;
            visibility: hidden !important;
        }
        </style>
        """, unsafe_allow_html=True)


def render_student_sidebar(db, student, menu_items, current_page):
    """قائمة جانبية موحدة للطالبة — Streamlit native."""
    full_name = student.get("full_name", "طالبة")
    with st.sidebar:
        col_main, col_close = st.columns([9, 1])
        with col_main:
            st.markdown(f"### 👤 {full_name}")
            st.caption("طالبة")
        

        if st.button("✕ إغلاق", key="student_sidebar_close_text_btn", use_container_width=True):
            st.session_state.sidebar_open = False
            st.rerun()

        st.markdown("---")

        for item in menu_items:
            if item == "🚪 تسجيل الخروج":
                continue
            btn_type = "primary" if item == current_page else "secondary"
            if st.button(item, key=f"student_nav_{item}", use_container_width=True, type=btn_type):
                st.session_state.student_dashboard_page = item
                st.session_state.sidebar_open = False
                st.rerun()

        st.markdown("---")
        if st.button("🚪 تسجيل الخروج", use_container_width=True, key="student_logout_btn"):
            student_logout(db)


# =============================================================================
# Student Dashboard (تسجيل دخول الطالبات)
# =============================================================================
def show_student_dashboard(db):
    """لوحة تحكم الطالبة — Dashboard واحد مع قائمة جانبية موحدة."""
    student = st.session_state.get("current_student")
    if not student:
        st.session_state.student_logged_in = False
        st.rerun()
        return

    students_df = db.get_students()
    if students_df.empty or "student_id" not in students_df.columns:
        st.error("تعذر تحميل بيانات الطالبة.")
        return
    student_row = students_df[students_df["student_id"] == student.get("student_id", "")]
    if student_row.empty:
        st.error("لم يتم العثور على بيانات الطالبة.")
        st.session_state.student_logged_in = False
        st.rerun()
        return
    student = student_row.iloc[0].to_dict()
    st.session_state.current_student = student

    if st.session_state.get("quiz_interface_started", False):
        render_student_top_bar(st.session_state.get("student_dashboard_page", "🏠 الرئيسية"))
        show_student_assessment_interface(db)
        return

    if "sidebar_open" not in st.session_state:
        st.session_state.sidebar_open = False

    menu_items = STUDENT_MENU_ITEMS
    current_page = normalize_student_dashboard_page(
        st.session_state.get("student_dashboard_page", "🏠 الرئيسية")
    )
    if current_page != st.session_state.get("student_dashboard_page"):
        st.session_state.student_dashboard_page = current_page
    if current_page not in menu_items:
        current_page = menu_items[0]
        st.session_state.student_dashboard_page = current_page

    inject_student_sidebar_css(st.session_state.sidebar_open)
    if st.session_state.sidebar_open:
        render_student_sidebar(db, student, menu_items, current_page)

    render_student_top_bar(current_page)

    if current_page == "👤 ملفي الشخصي":
        show_student_profile_tab(db, student)
    elif current_page == STUDENT_ASSESSMENTS_PAGE:
        show_student_assessments_page(db, student)
    elif current_page == "🔔 الإشعارات":
        show_student_notifications_tab(db, student)
    elif current_page == "🚪 تسجيل الخروج":
        student_logout(db)
    else:
        show_student_home_tab(db, student)


def show_student_grades_tab(db, student):
    """درجاتي — نتائج المسابقات والامتحانات للطالبة الحالية."""
    st.markdown(hero_header("درجاتي", "📊 درجات ونتائج الاختبارات"), unsafe_allow_html=True)

    student_id = student.get("student_id", "")
    student_quiz, student_exam = get_student_submitted_results(db, student_id)
    quizzes = db.get_quizzes()
    exams = db.get_exams()

    frames = []
    if not student_quiz.empty:
        sq = student_quiz.copy()
        sq["result_source"] = "quiz"
        frames.append(sq)
    if not student_exam.empty:
        se = student_exam.copy()
        se["result_source"] = "exam"
        frames.append(se)

    if not frames:
        st.info("لا توجد درجات بعد.")
        return

    student_results = pd.concat(frames, ignore_index=True, sort=False)
    if "submission_time" in student_results.columns:
        student_results = student_results.sort_values("submission_time", ascending=False)

    display = student_results.copy()
    display["المسابقة"] = ""
    display["النوع"] = ""
    if not quizzes.empty and "quiz_id" in display.columns:
        quiz_titles = quizzes[["quiz_id", "title"]].rename(columns={"title": "quiz_title"})
        display = display.merge(quiz_titles, on="quiz_id", how="left")
        display.loc[display["result_source"] == "quiz", "المسابقة"] = display["quiz_title"]
        display.loc[display["result_source"] == "quiz", "النوع"] = "مسابقة"
    if not exams.empty and "exam_id" in display.columns:
        exam_titles = exams[["exam_id", "title"]].rename(columns={"title": "exam_title"})
        display = display.merge(exam_titles, on="exam_id", how="left")
        display.loc[display["result_source"] == "exam", "المسابقة"] = display["exam_title"]
        display.loc[display["result_source"] == "exam", "النوع"] = "امتحان"

    if "score" in display.columns:
        display["score"] = pd.to_numeric(display["score"], errors="coerce").fillna(0)
    if "total_marks" in display.columns:
        display["total_marks"] = pd.to_numeric(display["total_marks"], errors="coerce").fillna(20)
    if "score" in display.columns and "total_marks" in display.columns:
        display["النسبة"] = (display["score"] / display["total_marks"] * 100).round(1)

    display_cols = ["النوع", "المسابقة", "score", "total_marks", "النسبة", "submission_time", "status"]
    available_cols = [c for c in display_cols if c in display.columns]
    st.dataframe(
        display[available_cols].rename(columns={
            "score": "الدرجة",
            "total_marks": "الدرجة الكلية",
            "submission_time": "التاريخ",
            "status": "الحالة"
        }),
        use_container_width=True
    )


def show_student_exam_history_tab(db, student):
    """سجل الامتحانات — مراجعة المحاولات المكتملة."""
    st.markdown(hero_header("سجل الامتحانات", "📋 تاريخ الاختبارات والمسابقات"), unsafe_allow_html=True)

    student_id = student.get("student_id", "")
    if st.session_state.get("review_result_id") and st.session_state.get("review_result_type"):
        render_student_attempt_review(db, student, st.session_state.review_result_id, st.session_state.review_result_type)
        return

    student_quiz, student_exam = get_student_submitted_results(db, student_id)
    quizzes = db.get_quizzes()
    exams = db.get_exams()

    history_rows = []
    if not student_quiz.empty:
        for _, row in student_quiz.iterrows():
            title = "—"
            if not quizzes.empty and "quiz_id" in quizzes.columns:
                match = quizzes[quizzes["quiz_id"] == row.get("quiz_id", "")]
                if not match.empty:
                    title = match.iloc[0].get("title", "—")
            history_rows.append({**row.to_dict(), "attempt_type": "quiz", "title": title})
    if not student_exam.empty:
        for _, row in student_exam.iterrows():
            title = "—"
            if not exams.empty and "exam_id" in exams.columns:
                match = exams[exams["exam_id"] == row.get("exam_id", "")]
                if not match.empty:
                    title = match.iloc[0].get("title", "—")
            history_rows.append({**row.to_dict(), "attempt_type": "exam", "title": title})

    if not history_rows:
        st.info("لا توجد سجلات اختبارات بعد.")
        return

    history_rows.sort(key=lambda x: str(x.get("submission_time", "")), reverse=True)
    for attempt in history_rows:
        result_id = attempt.get("result_id", "")
        attempt_type = attempt.get("attempt_type", "quiz")
        title = attempt.get("title", "—")
        score = attempt.get("score", "")
        total_marks = attempt.get("total_marks", "")
        submission_time = attempt.get("submission_time", "")
        percent = None
        try:
            score_val = float(score)
            total_val = float(total_marks) if total_marks else 0
            percent = round((score_val / total_val) * 100, 1) if total_val > 0 else None
        except (TypeError, ValueError):
            percent = None

        date_line, time_line = format_arabic_datetime(attempt.get("submission_time", ""))

        label = f"🎯 {title} — {score}/{total_marks}"
        if percent is not None:
            label += f" — {percent}%"
        if date_line and date_line != "—":
            label += f" — {date_line}"
            if time_line:
                label += f" {time_line}"

        with st.expander(label, expanded=False):
            st.markdown(f"**الحالة:** {attempt.get('status', '')}")
            if st.button("📖 مراجعة المحاولة", key=f"hist_review_{attempt_type}_{result_id}"):
                st.session_state.review_result_id = result_id
                st.session_state.review_result_type = attempt_type
                st.rerun()


def show_student_notifications_tab(db, student):
    """الإشعارات - عرض إشعارات الطالبة."""
    st.markdown(hero_header("الإشعارات", "🔔 الإشعارات والرسائل"), unsafe_allow_html=True)
    
    student_id = student.get("student_id", "")
    notifications = db.get_notifications(student_id)
    
    if notifications.empty:
        st.info("لا توجد إشعارات حالياً.")
        return
    
    # عرض الإشعارات
    for _, notif in notifications.iterrows():
        title = notif.get("title", "")
        message = notif.get("message", "")
        created_at = notif.get("created_at", "")
        is_read = notif.get("is_read", "False")
        
        # تحديد لون الإشعار
        bg_color = "#f8fafc" if is_read == "True" else "#dbeafe"
        border_color = "#e2e8f0" if is_read == "True" else "#2563eb"
        
        st.markdown(f"""
        <div style="background: {bg_color}; border: 1px solid {border_color}; border-radius: 12px; padding: 1rem; margin-bottom: 1rem;">
            <div style="font-weight: 700; color: #0f172a; margin-bottom: 0.5rem;">{title}</div>
            <div style="color: #64748b; margin-bottom: 0.5rem;">{message}</div>
            <div style="font-size: 0.75rem; color: #94a3b8;">{created_at}</div>
        </div>
        """, unsafe_allow_html=True)
    
    # زر تحديد الكل كمقروء
    if st.button("✅ تحديد الكل كمقروء", use_container_width=True, key="mark_all_read_btn"):
        for _, notif in notifications.iterrows():
            if notif.get("is_read", "False") != "True":
                db.mark_notification_read(notif.get("notification_id", ""))
        st.success("تم تحديد جميع الإشعارات كمقروءة")
        st.rerun()


def render_student_attempt_review(db, student, result_id, result_type):
    """مراجعة تفصيلية — بطاقات أسئلة مدمجة."""
    inject_competitions_page_css()

    student_id = student.get("student_id", "")
    attempt = verify_student_owns_result(db, student_id, result_id, result_type)
    if attempt is None:
        st.error("🚫 غير مصرح بمراجعة هذه النتيجة.")
        return

    assessment_id = attempt.get("quiz_id", "")
    assessment_row = get_assessment_record(db, result_type, assessment_id) or {}
    title = assessment_row.get("title", "الامتحان" if result_type == "exam" else "الاختبار")
    questions_df = db.get_quiz_questions(assessment_id)

    score = attempt.get("score", "")
    total_marks = attempt.get("total_marks", "") or "20"
    try:
        score_val = float(score)
        total_val = float(total_marks) if total_marks else 0
        percentage = round((score_val / total_val * 100), 1) if total_val > 0 else None
    except (TypeError, ValueError):
        percentage = None
    date_line, time_line = format_arabic_datetime(attempt.get("submission_time", ""))
    result_label, result_cls = _comp_result_summary(score, total_marks, percentage)
    pct_display = f"{percentage}%" if percentage is not None else "—"

    st.markdown(f"""
    <div class="comp-card">
        <div class="comp-my-title">🎯 {title}</div>
        <div class="comp-stat-grid">
            <div class="comp-stat-box">
                <div class="comp-stat-label">⭐ الدرجة</div>
                <div class="comp-stat-value">{score} / {total_marks}</div>
                <div class="comp-stat-sub">من {total_marks}</div>
            </div>
            <div class="comp-stat-box">
                <div class="comp-stat-label">٪ النسبة</div>
                <div class="comp-stat-value">{pct_display}</div>
            </div>
            <div class="comp-stat-box">
                <div class="comp-stat-label">📅 تاريخ الحل</div>
                <div class="comp-stat-value" style="font-size:0.78rem;">{date_line}</div>
                <div class="comp-stat-sub">{time_line}</div>
            </div>
        </div>
        <div class="comp-result-row">
            <div class="comp-result-label">النتيجة</div>
            <div class="comp-result-badge {result_cls}">{result_label}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<p class="comp-section-title">📝 الأسئلة والإجابات</p>', unsafe_allow_html=True)

    if questions_df.empty:
        st.markdown('<div class="comp-empty">لا توجد أسئلة لهذا الاختبار</div>', unsafe_allow_html=True)
    else:
        try:
            answers_data = json.loads(attempt.get("answers", "{}") or "{}")
        except (json.JSONDecodeError, TypeError):
            answers_data = {}

        for idx, row in questions_df.iterrows():
            q = row.to_dict()
            q_id = q.get("question_id", "")
            q_num = idx + 1
            student_answer = str(answers_data.get(q_id, "") or "").strip()
            correct_answer = str(q.get("correct_answer", "")).strip()
            q_text = q.get("question_text", "")

            if not student_answer:
                result_text = "⚪ لم تتم الإجابة"
                ans_cls = "comp-q-neutral"
            elif student_answer.strip().lower() == correct_answer.strip().lower():
                result_text = "صحيحة ✅"
                ans_cls = "comp-q-correct"
            else:
                result_text = "خاطئة ❌"
                ans_cls = "comp-q-wrong"

            st.markdown(f"""
            <div class="comp-q-card">
                <div class="comp-q-header">
                    <div class="comp-q-num">{q_num}</div>
                    <div class="comp-q-title">السؤال {q_num}</div>
                </div>
                <div class="comp-q-text">{q_text}</div>
                <div class="comp-q-row">
                    <div class="comp-q-cell">
                        <div class="comp-q-cell-label">إجابتك</div>
                        <div class="comp-q-cell-value">{student_answer or '—'}</div>
                    </div>
                    <div class="comp-q-cell">
                        <div class="comp-q-cell-label">الإجابة الصحيحة</div>
                        <div class="comp-q-cell-value">{correct_answer or '—'}</div>
                    </div>
                    <div class="comp-q-cell">
                        <div class="comp-q-cell-label">النتيجة</div>
                        <div class="comp-q-cell-value {ans_cls}">{result_text}</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

    if st.button("⬅️ العودة إلى المسابقات والاختبارات", key=f"back_from_review_{result_id}", use_container_width=True):
        st.session_state.review_result_id = None
        st.session_state.review_result_type = None
        st.rerun()


def show_student_home_tab(db, student):
    """الرئيسية — ترحيب وإحصائيات."""
    full_name = student.get("full_name", "طالبة")
    st.markdown(f"## مرحبًا، {full_name} 👋")

    student_id = student.get("student_id", "")
    student_quiz, student_exam = get_student_submitted_results(db, student_id)
    available_count = len(build_unified_assessments_for_student(db, student))
    completed_count = len(student_quiz) + len(student_exam)

    all_results = []
    for df in (student_quiz, student_exam):
        if not df.empty:
            all_results.append(df)
    latest_score = None
    if all_results:
        combined = pd.concat(all_results, ignore_index=True, sort=False)
        if "submission_time" in combined.columns:
            combined = combined.sort_values("submission_time", ascending=False)
        if not combined.empty and "score" in combined.columns:
            latest_score = pd.to_numeric(combined.iloc[0]["score"], errors="coerce")

    all_scores = []
    for df in (student_quiz, student_exam):
        if not df.empty and "score" in df.columns:
            all_scores.extend(pd.to_numeric(df["score"], errors="coerce").dropna().tolist())
    avg_score = round(sum(all_scores) / len(all_scores), 1) if all_scores else None

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📝 اختبارات متاحة", available_count)
    col2.metric("✅ اختبارات منجزة", completed_count)
    col3.metric("📊 آخر درجة", latest_score if latest_score is not None else "—")
    col4.metric("📈 متوسط الدرجات", avg_score if avg_score is not None else "—")


def show_student_profile_tab(db, student):
    """ملفي الشخصي — عرض وتعديل بيانات الطالبة (مرة واحدة فقط)."""
    st.markdown(hero_header("ملفي الشخصي", "👤 بياناتي الشخصية"), unsafe_allow_html=True)

    students_df = db.get_students()
    student_row = students_df[students_df["student_id"] == student.get("student_id", "")]
    if student_row.empty:
        st.error("لم يتم العثور على بيانات الطالبة.")
        return
    student = student_row.iloc[0].to_dict()
    st.session_state.current_student = student

    sections = db.get_sections()
    section_name = ""
    sec_id = student.get("section_id", "")
    if not sections.empty and sec_id:
        sec_match = sections[sections["section_id"] == sec_id]
        if not sec_match.empty:
            section_name = sec_match.iloc[0].get("section_name", "")

    profile_edit_used = str(student.get("profile_edit_used", "")).strip().lower() in ("true", "1", "yes", "نعم")

    st.markdown("### 📋 بياناتي")
    col1, col2 = st.columns(2)
    col1.markdown(f"**الاسم الكامل:** {student.get('full_name', '—')}")
    col2.markdown(f"**كود الطالبة:** {student.get('student_code', '—')}")
    col1.markdown(f"**الفصل:** {section_name or '—'}")
    col2.markdown(f"**الهاتف:** {student.get('phone', '—') or '—'}")
    col1.markdown(f"**رقم ولي الأمر:** {student.get('parent_phone', '—') or '—'}")
    col2.markdown(f"**تاريخ الميلاد:** {student.get('birthdate', '—') or '—'}")
    col1.markdown(f"**العنوان:** {student.get('address', '—') or '—'}")
    col2.markdown(f"**المدرسة:** {student.get('school', '—') or '—'}")
    if student.get("notes"):
        st.markdown(f"**ملاحظات:** {student.get('notes', '')}")

    st.markdown("---")

    if profile_edit_used:
        st.warning("⚠️ لقد استخدمتِ فرصة تعديل البيانات بالفعل. البيانات أصبحت للقراءة فقط.")
    else:
        st.markdown("### ✏️ تعديل بياناتي (مرة واحدة فقط)")
        st.info("يمكنكِ تعديل بياناتك مرة واحدة فقط. بعد الحفظ لن تتمكني من التعديل مرة أخرى.")
        with st.form("student_profile_edit_form"):
            edit_name = st.text_input("الاسم الكامل*", value=student.get("full_name", ""))
            edit_phone = st.text_input("الهاتف", value=student.get("phone", ""))
            edit_parent_phone = st.text_input("رقم ولي الأمر", value=student.get("parent_phone", ""))
            bd_value = pd.to_datetime(student.get("birthdate", "")).date() if student.get("birthdate") else None
            edit_birthdate = st.date_input("تاريخ الميلاد", value=bd_value)
            edit_address = st.text_input("العنوان", value=student.get("address", ""))
            edit_school = st.text_input("المدرسة", value=student.get("school", ""))
            edit_notes = st.text_area("ملاحظات", value=student.get("notes", ""))
            submitted = st.form_submit_button("💾 حفظ البيانات", use_container_width=True)
            if submitted:
                if not edit_name:
                    st.error("الاسم الكامل مطلوب")
                else:
                    ok = db.update_student(student.get("student_id", ""), {
                        "full_name": edit_name,
                        "phone": edit_phone,
                        "parent_phone": edit_parent_phone,
                        "birthdate": edit_birthdate.strftime("%Y-%m-%d") if edit_birthdate else "",
                        "address": edit_address,
                        "school": edit_school,
                        "notes": edit_notes,
                        "profile_edit_used": "True"
                    })
                    if not ok:
                        st.error("🚫 تم رفض التعديل: لقد استخدمتِ فرصة تعديل البيانات بالفعل.")
                    else:
                        db.add_log(student.get("student_id", ""), "تعديل بيانات الطالبة", f"تم تعديل بيانات الطالبة: {edit_name}")
                        st.success("✅ تم حفظ بياناتك بنجاح! لن تتمكني من التعديل مرة أخرى.")
                        time.sleep(1)
                        st.rerun()


def _get_assessment_attempt_status(db, student_id, assessment_type, assessment_id):
    """حالة محاولة الطالبة: available | started | submitted."""
    if assessment_type == "exam":
        results = db.get_exam_results()
        id_col = "exam_id"
    else:
        results = db.get_quiz_results()
        id_col = "quiz_id"
    if results.empty or id_col not in results.columns or "student_id" not in results.columns:
        return "available"
    attempts = results[(results[id_col].astype(str) == str(assessment_id)) & (results["student_id"].astype(str) == str(student_id))]
    if attempts.empty:
        return "available"
    if "status" in attempts.columns and not attempts[attempts["status"] == "submitted"].empty:
        return "submitted"
    return "started"


def inject_competitions_page_css():
    """CSS for the redesigned competitions page (RTL, mobile-first)."""
    st.markdown("""
    <style>
    .comp-page { direction: rtl; font-family: 'Cairo', sans-serif; max-width: 100%; overflow-x: hidden; }
    .comp-top-header {
        display: flex; align-items: center; justify-content: space-between;
        background: #fff; border-bottom: 1px solid #e5e7eb;
        padding: 0.65rem 0.25rem; margin: -1rem -1rem 1rem -1rem;
        box-shadow: 0 1px 3px rgba(15,23,42,0.06); position: sticky; top: 0; z-index: 100;
    }
    .comp-header-title { font-size: 1rem; font-weight: 800; color: #0f172a; text-align: center; flex: 1; margin: 0 0.5rem; line-height: 1.4; }
    .comp-hero {
        position: relative; border-radius: 16px; overflow: hidden; min-height: 130px;
        margin-bottom: 1.25rem; background-size: cover; background-position: center;
        box-shadow: 0 4px 14px rgba(37,99,235,0.15);
    }
    .comp-hero-overlay {
        background: linear-gradient(135deg, rgba(15,23,42,0.55), rgba(37,99,235,0.45));
        padding: 1.5rem 1.25rem; min-height: 130px; display: flex; flex-direction: column; justify-content: center;
    }
    .comp-hero-title { color: #fff; font-size: 1.5rem; font-weight: 800; margin: 0 0 0.35rem 0; }
    .comp-hero-sub { color: rgba(255,255,255,0.92); font-size: 0.85rem; margin: 0; line-height: 1.5; }
    .comp-section-title {
        font-size: 1rem; font-weight: 800; color: #0f172a; margin: 1.25rem 0 0.75rem 0;
        display: flex; align-items: center; gap: 0.4rem;
    }
    .comp-card {
        background: #fff; border: 1px solid #e2e8f0; border-radius: 14px;
        padding: 1rem 1.1rem; margin-bottom: 0.75rem;
        box-shadow: 0 2px 8px rgba(15,23,42,0.05);
    }
    .comp-available-row {
        display: flex; align-items: center; justify-content: space-between; gap: 0.75rem;
    }
    .comp-available-title { font-size: 0.9rem; font-weight: 700; color: #0f172a; line-height: 1.5; flex: 1; }
    .comp-badge-available {
        display: inline-block; background: #dcfce7; color: #166534;
        font-size: 0.72rem; font-weight: 700; padding: 0.2rem 0.55rem; border-radius: 999px; margin-top: 0.35rem;
    }
    .comp-badge-done { background: #dbeafe; color: #1d4ed8; }
    .comp-badge-started { background: #fef3c7; color: #92400e; }
    .comp-icon-box {
        width: 42px; height: 42px; border-radius: 10px; background: #eff6ff;
        display: flex; align-items: center; justify-content: center; font-size: 1.2rem; flex-shrink: 0;
    }
    .comp-my-title { font-size: 0.92rem; font-weight: 800; color: #0f172a; margin-bottom: 0.85rem; line-height: 1.45; }
    .comp-stat-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 0.5rem; margin-bottom: 0.85rem; }
    .comp-stat-box {
        background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 12px;
        padding: 0.65rem 0.5rem; text-align: center;
    }
    .comp-stat-label { font-size: 0.68rem; color: #64748b; font-weight: 600; margin-bottom: 0.25rem; }
    .comp-stat-value { font-size: 0.95rem; font-weight: 800; color: #2563eb; line-height: 1.3; }
    .comp-stat-sub { font-size: 0.65rem; color: #94a3b8; margin-top: 0.15rem; }
    .comp-result-row { text-align: center; margin: 0.75rem 0; }
    .comp-result-label { font-size: 0.75rem; color: #64748b; font-weight: 600; }
    .comp-result-badge {
        display: inline-block; margin-top: 0.35rem; padding: 0.35rem 0.85rem;
        border-radius: 999px; font-size: 0.8rem; font-weight: 700;
    }
    .comp-result-success { background: #dcfce7; color: #166534; }
    .comp-result-fail { background: #fee2e2; color: #991b1b; }
    .comp-nav-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 0.65rem; margin-top: 0.5rem; }
    .comp-q-card {
        background: #fff; border: 1px solid #e2e8f0; border-radius: 14px;
        padding: 0.85rem 1rem; margin-bottom: 0.65rem;
        box-shadow: 0 1px 4px rgba(15,23,42,0.04);
    }
    .comp-q-header { display: flex; align-items: center; gap: 0.5rem; margin-bottom: 0.5rem; }
    .comp-q-num {
        width: 28px; height: 28px; border-radius: 50%; background: #2563eb; color: #fff;
        font-size: 0.75rem; font-weight: 800; display: flex; align-items: center; justify-content: center; flex-shrink: 0;
    }
    .comp-q-title { font-size: 0.82rem; font-weight: 700; color: #0f172a; }
    .comp-q-text { font-size: 0.85rem; color: #334155; line-height: 1.55; margin-bottom: 0.65rem; }
    .comp-q-row { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 0.4rem; }
    .comp-q-cell { background: #f8fafc; border-radius: 10px; padding: 0.5rem 0.4rem; text-align: center; }
    .comp-q-cell-label { font-size: 0.65rem; color: #64748b; font-weight: 600; margin-bottom: 0.2rem; }
    .comp-q-cell-value { font-size: 0.78rem; font-weight: 700; color: #0f172a; word-break: break-word; }
    .comp-q-correct { color: #166534; }
    .comp-q-wrong { color: #dc2626; }
    .comp-q-neutral { color: #64748b; }
    .comp-empty { text-align: center; color: #64748b; font-size: 0.88rem; padding: 1.5rem 1rem; }
    div[data-testid="stButton"] button.comp-menu-btn {
        background: #2563eb !important; color: #fff !important; border: none !important;
        border-radius: 10px !important; min-width: 44px !important; min-height: 44px !important;
        font-size: 1.2rem !important; padding: 0.4rem 0.65rem !important;
    }
    div[data-testid="stButton"] button.comp-help-btn {
        background: #eff6ff !important; color: #2563eb !important; border: 1px solid #bfdbfe !important;
        border-radius: 10px !important; font-size: 0.78rem !important; font-weight: 700 !important;
        padding: 0.45rem 0.65rem !important;
    }
    div[data-testid="stButton"] button.comp-nav-btn {
        background: #fff !important; color: #2563eb !important; border: 2px solid #2563eb !important;
        border-radius: 12px !important; font-weight: 700 !important; font-size: 0.82rem !important;
    }
    div[data-testid="stButton"] button.comp-start-btn {
        background: #2563eb !important; color: #fff !important; border: none !important;
        border-radius: 10px !important; font-weight: 700 !important; font-size: 0.82rem !important;
    }
    div[data-testid="stButton"] button.comp-review-btn {
        background: #eff6ff !important; color: #1d4ed8 !important; border: 1px solid #bfdbfe !important;
        border-radius: 10px !important; font-weight: 700 !important; font-size: 0.82rem !important;
    }
    @media (max-width: 480px) {
        .comp-stat-grid { grid-template-columns: 1fr; }
        .comp-q-row { grid-template-columns: 1fr; }
        .comp-header-title { font-size: 0.88rem; }
        .comp-hero-title { font-size: 1.25rem; }
        .comp-nav-grid { grid-template-columns: 1fr; }
    }
    </style>
    """, unsafe_allow_html=True)


def render_competitions_hero():
    """Hero banner with background image."""
    bg = f"url('data:image/jpeg;base64,{BG_IMAGE_BASE64}')" if BG_IMAGE_BASE64 else "linear-gradient(135deg, #2563eb, #7c3aed)"
    st.markdown(f"""
    <div class="comp-hero" style="background-image: {bg};">
        <div class="comp-hero-overlay">
            <h2 class="comp-hero-title">المسابقات والاختبارات</h2>
            <p class="comp-hero-sub">تابع إنجازاتك واطلع على مسابقاتك واختباراتك في مكان واحد</p>
        </div>
    </div>
    """, unsafe_allow_html=True)


def _comp_result_summary(score, total_marks, percent):
    """Human-readable overall result label."""
    if percent is None:
        return "—", "comp-result-fail"
    if percent >= 50:
        return "إجابة صحيحة ✅", "comp-result-success"
    return "تحتاج مراجعة ❌", "comp-result-fail"


def show_student_assessments_page(db, student):
    """Unified student page: available assessments, grades, and attempt history."""
    inject_competitions_page_css()
    st.markdown('<div class="comp-page">', unsafe_allow_html=True)
    render_competitions_hero()

    if st.session_state.get("review_result_id") and st.session_state.get("review_result_type"):
        render_student_attempt_review(db, student, st.session_state.review_result_id, st.session_state.review_result_type)
        st.markdown('</div>', unsafe_allow_html=True)
        return

    confirmation = st.session_state.get("assessment_confirmation")
    if confirmation:
        _render_assessment_start_confirmation(db, student, confirmation)
        st.markdown('</div>', unsafe_allow_html=True)
        return

    tab_labels = ["📋 المتاحة", "📊 درجاتي", "🗂️ السجل"]
    tab_keys = ["available", "grades", "history"]
    if "assessment_tab" not in st.session_state:
        st.session_state.assessment_tab = "available"
    selected_tab = st.radio(
        "أقسام الصفحة",
        tab_keys,
        format_func=lambda x: tab_labels[tab_keys.index(x)],
        horizontal=True,
        key="assessment_tab_radio",
        index=tab_keys.index(st.session_state.assessment_tab) if st.session_state.assessment_tab in tab_keys else 0,
        label_visibility="collapsed",
    )
    st.session_state.assessment_tab = selected_tab
    st.markdown("---")

    if selected_tab == "available":
        _render_student_available_assessments(db, student)
    elif selected_tab == "grades":
        show_student_grades_tab(db, student)
    else:
        show_student_exam_history_tab(db, student)

    st.markdown('</div>', unsafe_allow_html=True)


def show_student_competitions_tab(db, student):
    """Legacy alias — redirects to unified assessments page."""
    show_student_assessments_page(db, student)


def _render_assessment_start_confirmation(db, student, confirmation):
    a_type = confirmation.get("type", "quiz")
    a_id = confirmation.get("id", "")
    assessment_row = get_assessment_record(db, a_type, a_id)
    if not assessment_row:
        st.warning("تعذر العثور على بيانات الاختبار.")
        st.session_state.assessment_confirmation = None
        st.rerun()
        return

    can_access, deny_reason = student_can_access_assessment(db, student, a_type, a_id)
    if not can_access:
        st.warning(deny_reason or "غير مصرح بالدخول إلى هذا الاختبار.")
        st.session_state.assessment_confirmation = None
        return

    num_q = get_assessment_question_count(db, a_type, a_id)
    total_marks = assessment_row.get("total_marks", "20")
    time_limit = assessment_row.get("time_limit_minutes") or assessment_row.get("duration_minutes", "—")
    type_label = "امتحان" if a_type == "exam" else "مسابقة"

    st.markdown('<div class="comp-card">', unsafe_allow_html=True)
    st.markdown("## تأكيد دخول الاختبار")
    st.markdown(f"**النوع:** {type_label}")
    st.markdown(f"**اسم الاختبار:** {assessment_row.get('title', '')}")
    st.markdown(f"**عدد الأسئلة:** {num_q or '—'}")
    st.markdown(f"**الدرجة الكلية:** {total_marks}")
    st.markdown(f"**الوقت المحدد:** {time_limit} دقيقة")
    st.markdown("---")
    st.markdown("## تعليمات مهمة قبل بدء الاختبار")
    instructions = [
        "يجب قراءة التعليمات جيدًا قبل البدء.",
        "بمجرد بدء الاختبار يبدأ احتساب الوقت.",
        "يجب الالتزام بالوقت المحدد.",
        "لا تقومي بإغلاق صفحة الاختبار.",
        "لا تقومي بالخروج من صفحة الاختبار.",
        "لا تقومي بإعادة تحميل الصفحة.",
        "لا تقومي بمغادرة الاختبار أثناء أدائه.",
        "سيتم التعامل مع انتهاء الوقت حسب نظام التسليم التلقائي.",
        "تأكدي من إجاباتك قبل تسليم الاختبار.",
    ]
    for line in instructions:
        st.markdown(f"- {line}")
    st.markdown("---")

    col_cancel, col_confirm = st.columns(2)
    with col_cancel:
        if st.button("إلغاء", use_container_width=True, key="cancel_assessment_start"):
            st.session_state.assessment_confirmation = None
            st.rerun()
    with col_confirm:
        if st.button("أوافق وأبدأ الاختبار", use_container_width=True, key="confirm_assessment_start"):
            clear_assessment_session_state()
            st.session_state.selected_assessment_type = a_type
            st.session_state.selected_assessment_id = a_id
            st.session_state.assessment_confirmation = None
            st.session_state.quiz_interface_started = True
            if a_type == "quiz":
                st.session_state.selected_quiz_id = a_id
            else:
                st.session_state.selected_exam_id = a_id
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


def _render_student_available_assessments(db, student):
    student_id = student.get("student_id", "")
    quizzes = db.get_quizzes()
    exams = db.get_exams()

    st.markdown('<p class="comp-section-title">⚽ الاختبارات المتاحة</p>', unsafe_allow_html=True)
    assessments = build_unified_assessments_for_student(db, student)
    available_items = []
    for item in assessments:
        a_type = item["assessment_type"]
        a_id = item["assessment_id"]
        status = _get_assessment_attempt_status(db, student_id, a_type, a_id)
        if status != "submitted":
            available_items.append((item, status))

    if not available_items:
        st.markdown('<div class="comp-empty">لا توجد اختبارات متاحة حالياً</div>', unsafe_allow_html=True)
    else:
        for item, status in available_items:
            a_type = item["assessment_type"]
            a_id = item["assessment_id"]
            type_badge = "امتحان" if a_type == "exam" else "مسابقة"
            if status == "started":
                badge = '<span class="comp-badge-available comp-badge-started">الحالة: بدأت — يمكنك المتابعة</span>'
                btn_label = "متابعة الاختبار"
            else:
                badge = f'<span class="comp-badge-available">النوع: {type_badge} | متاحة</span>'
                btn_label = "بدء الاختبار"
            st.markdown(f"""
            <div class="comp-card">
                <div class="comp-available-row">
                    <div class="comp-icon-box">{"📝" if a_type == "exam" else "⚽"}</div>
                    <div style="flex:1;">
                        <div class="comp-available-title">{item['title']}</div>
                        {badge}
                    </div>
                    <div style="color:#2563eb;font-size:1.2rem;">◀</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button(btn_label, key=f"start_{a_type}_{a_id}", use_container_width=True):
                st.session_state.assessment_confirmation = {"type": a_type, "id": a_id}
                st.rerun()

    st.markdown('<p class="comp-section-title">📋 نتائجي الأخيرة</p>', unsafe_allow_html=True)
    student_quiz, student_exam = get_student_submitted_results(db, student_id)
    history_rows = []
    if not student_quiz.empty:
        for _, row in student_quiz.iterrows():
            title = "—"
            if not quizzes.empty:
                m = quizzes[quizzes["quiz_id"] == row.get("quiz_id", "")]
                if not m.empty:
                    title = m.iloc[0].get("title", "—")
            history_rows.append({**row.to_dict(), "attempt_type": "quiz", "title": title})
    if not student_exam.empty:
        for _, row in student_exam.iterrows():
            title = "—"
            if not exams.empty:
                m = exams[exams["exam_id"] == row.get("exam_id", "")]
                if not m.empty:
                    title = m.iloc[0].get("title", "—")
            history_rows.append({**row.to_dict(), "attempt_type": "exam", "title": title})

    if not history_rows:
        st.markdown('<div class="comp-empty">لا توجد نتائج بعد</div>', unsafe_allow_html=True)
    else:
        history_rows.sort(key=lambda x: str(x.get("submission_time", "")), reverse=True)
        for attempt in history_rows[:5]:
            result_id = attempt.get("result_id", "")
            attempt_type = attempt.get("attempt_type", "quiz")
            title = attempt.get("title", "—")
            score = attempt.get("score", "")
            total_marks = attempt.get("total_marks", "") or "20"
            percent = None
            try:
                sv = float(score)
                tv = float(total_marks) if total_marks else 0
                percent = round((sv / tv) * 100, 1) if tv > 0 else None
            except (TypeError, ValueError):
                pass
            date_line, time_line = format_arabic_datetime(attempt.get("submission_time", ""))
            result_label, result_cls = _comp_result_summary(score, total_marks, percent)
            pct_display = f"{percent}%" if percent is not None else "—"

            st.markdown(f"""
            <div class="comp-card">
                <div class="comp-my-title">🎯 {title}</div>
                <div class="comp-stat-grid">
                    <div class="comp-stat-box">
                        <div class="comp-stat-label">⭐ الدرجة</div>
                        <div class="comp-stat-value">{score} / {total_marks}</div>
                    </div>
                    <div class="comp-stat-box">
                        <div class="comp-stat-label">٪ النسبة</div>
                        <div class="comp-stat-value">{pct_display}</div>
                    </div>
                    <div class="comp-stat-box">
                        <div class="comp-stat-label">📅 تاريخ الحل</div>
                        <div class="comp-stat-value" style="font-size:0.78rem;">{date_line}</div>
                        <div class="comp-stat-sub">{time_line}</div>
                    </div>
                </div>
                <div class="comp-result-row">
                    <div class="comp-result-label">النتيجة</div>
                    <div class="comp-result-badge {result_cls}">{result_label}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("📖 مراجعة الأسئلة والإجابات", key=f"comp_review_{attempt_type}_{result_id}", use_container_width=True):
                st.session_state.review_result_id = result_id
                st.session_state.review_result_type = attempt_type
                st.rerun()


# =============================================================================
# Sidebar Navigation
# =============================================================================
def show_sidebar_navigation(db):
    with st.sidebar:
        user = st.session_state.user
        role = user.get("role", "")
        menu_items = get_role_menu(role)
        if not menu_items:
            st.warning("صلاحية غير معروفة")
            return None

        # ===== Premium Sidebar Header =====
        st.markdown("""
        <div class='sidebar-brand'>
            <div class='brand-logo'>⛪</div>
            <div class='brand-text'>
                <h3>كنيسة الشهيدة دميانة</h3>
                <small>نظام الإدارة المتكامل</small>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ===== User Card =====
        full_name = user.get('full_name', '')
        role_label = {
            "System Admin": "مدير النظام",
            "Father Account": "أب كاهن",
            "Service Manager": "أمين الخدمة",
            "Teacher": "مدرسة",
            "Student": "طالبة"
        }.get(role, role)
        initials = get_initials(full_name)
        st.markdown(f"""
        <div class='sidebar-user'>
            <div class='user-avatar-lg'>{initials}</div>
            <div class='user-info'>
                <strong>{full_name}</strong>
                <span>{role_label}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ===== Notification Badge =====
        try:
            unread_count = get_unread_notification_count(db, user.get("user_id", ""))
            if unread_count > 0:
                st.markdown(f"""
                <div style="background:linear-gradient(135deg,#fee2e2,#fef3c7); border:1px solid #fecaca; border-radius:12px; padding:0.6rem 1rem; margin-bottom:0.8rem; display:flex; align-items:center; gap:0.6rem;">
                    <span style="font-size:1.2rem;">🔔</span>
                    <div style="flex:1;">
                        <div style="font-size:0.8rem; font-weight:700; color:#991b1b;">لديك {unread_count} إشعار غير مقروء</div>
                        <div style="font-size:0.7rem; color:#6b7280;">اضغط على الإشعارات في القائمة</div>
                    </div>
                    <span style="background:#dc2626; color:white; border-radius:9999px; padding:0.15rem 0.6rem; font-size:0.75rem; font-weight:700;">{unread_count}</span>
                </div>
                """, unsafe_allow_html=True)
        except Exception:
            pass

        # ===== Collapse button =====
        if st.button("إخفاء القائمة", key="hide_sidebar_btn", use_container_width=True):
            st.session_state.show_sidebar = False
            st.rerun()

        # ===== Menu items with Arabic icons only =====
        current_choice = normalize_admin_menu_choice(st.session_state.get("menu_choice", menu_items[0]))
        if current_choice not in menu_items:
            current_choice = menu_items[0]
        if current_choice != st.session_state.get("menu_choice"):
            st.session_state.menu_choice = current_choice

        st.markdown('<div class="sidebar-nav nav-btn-container">', unsafe_allow_html=True)
        for item in menu_items:
            btn_type = "primary" if item == current_choice else "secondary"
            if st.button(item, key=f"nav_btn_{item}", use_container_width=True, type=btn_type):
                if item != current_choice:
                    st.session_state.menu_choice = item
                st.session_state.show_sidebar = False
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        # ===== Sidebar footer =====
        st.markdown('<div class="sidebar-footer">', unsafe_allow_html=True)
        if st.button("تسجيل الخروج", use_container_width=True, key="logout_btn"):
            logout(db)
        st.markdown('</div>', unsafe_allow_html=True)
    return current_choice


# =============================================================================
# Dashboard
# =============================================================================
def show_dashboard(db):
    user = st.session_state.user
    role = user.get("role", "")
    section_id = user.get("section_id", "")
    # Hero banner for dashboard
    st.markdown(hero_header("لوحة التحكم", "مرحباً بك في نظام إدارة الكنيسة"), unsafe_allow_html=True)
    if role in ["System Admin", "Service Manager"] and st.session_state.get("data_errors"):
        with st.expander("⚠️ تنبيهات هامة - أخطاء في البيانات", expanded=True):
            for err in st.session_state.data_errors:
                st.warning(err)
            if st.button("🔧 إصلاح تلقائي (إنشاء الفصول الناقصة)", key="auto_fix_btn"):
                if auto_fix_missing_sections(db):
                    st.session_state.data_errors = validate_data_integrity(db)
                    st.success("تم إنشاء الفصول الناقصة. سيتم تحديث الصفحة...")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.info("لا توجد فصول ناقصة لإصلاحها.")
    students = db.get_students()
    attendance = db.get_attendance()
    followup = db.get_followup()
    if role in ["Teacher", "Service Manager"] and section_id:
        if not students.empty and "section_id" in students.columns:
            students = students[students.section_id == section_id]
        if not attendance.empty and "section_id" in attendance.columns:
            attendance = attendance[attendance.section_id == section_id]
        if not followup.empty and not students.empty and "student_id" in followup.columns and "student_id" in students.columns:
            followup = followup[followup.student_id.isin(students["student_id"])]
    if not attendance.empty and "date" in attendance.columns:
        attendance["date"] = pd.to_datetime(attendance["date"], errors="coerce")
    total_students = len(students)
    today_str = get_cairo_now().strftime("%Y-%m-%d")
    present_today = len(attendance[(attendance.date == today_str) & (attendance.status == "حاضر")]) if not attendance.empty and "status" in attendance.columns else 0
    absent_today = len(attendance[(attendance.date == today_str) & (attendance.status == "غائب")]) if not attendance.empty and "status" in attendance.columns else 0
    need_follow = len(followup[followup.regularity_status == "منقطع"]) if not followup.empty and "regularity_status" in followup.columns else 0
    # Statistics cards with improved styling
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
        <div style="background: white; padding: 1.5rem; border-radius: 16px; border: 1px solid #e2e8f0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: right;">
            <div style="font-size: 0.85rem; color: #64748b; font-weight: 600; margin-bottom: 0.5rem;">عدد الطالبات</div>
            <div style="font-size: 2rem; font-weight: 800; color: #2563eb;">{total_students}</div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div style="background: white; padding: 1.5rem; border-radius: 16px; border: 1px solid #e2e8f0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: right;">
            <div style="font-size: 0.85rem; color: #64748b; font-weight: 600; margin-bottom: 0.5rem;">الحضور اليوم</div>
            <div style="font-size: 2rem; font-weight: 800; color: #059669;">{present_today}</div>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div style="background: white; padding: 1.5rem; border-radius: 16px; border: 1px solid #e2e8f0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: right;">
            <div style="font-size: 0.85rem; color: #64748b; font-weight: 600; margin-bottom: 0.5rem;">الغياب اليوم</div>
            <div style="font-size: 2rem; font-weight: 800; color: #dc2626;">{absent_today}</div>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div style="background: white; padding: 1.5rem; border-radius: 16px; border: 1px solid #e2e8f0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: right;">
            <div style="font-size: 0.85rem; color: #64748b; font-weight: 600; margin-bottom: 0.5rem;">منقطعات</div>
            <div style="font-size: 2rem; font-weight: 800; color: #d97706;">{need_follow}</div>
        </div>
        """, unsafe_allow_html=True)
    st.markdown("#### 📈 الحضور الأسبوعي")
    if not attendance.empty and "date" in attendance.columns and "status" in attendance.columns:
        last_week = get_cairo_now().replace(tzinfo=None) - timedelta(days=7)
        recent = attendance[attendance.date >= last_week]
        if not recent.empty:
            fig = px.histogram(recent, x="date", color="status", barmode="group")
            fig.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("لا توجد بيانات حضور للأيام الماضية.")
    else:
        st.info("لا توجد بيانات حضور بعد.")
    st.markdown("#### 🏅 أكثر 5 طالبات غياباً هذا الشهر")
    if not attendance.empty and "date" in attendance.columns and "status" in attendance.columns:
        month_start = get_cairo_now().replace(day=1).strftime("%Y-%m-%d")
        month_att = attendance[(attendance.date >= month_start) & (attendance.status == "غائب")]
        if not month_att.empty:
            absent_counts = month_att.groupby("student_id").size().reset_index(name="أيام الغياب")
            absent_counts = absent_counts.sort_values("أيام الغياب", ascending=False).head(5)
            if not students.empty and "student_id" in students.columns and "full_name" in students.columns:
                absent_counts = absent_counts.merge(students[["student_id", "full_name"]], on="student_id", how="left")
            st.dataframe(absent_counts[["full_name", "أيام الغياب"]], use_container_width=True)
        else:
            st.info("لا يوجد غياب هذا الشهر.")
    st.markdown("#### 🔔 بنات بحاجة لافتقاد عاجل")
    urgent = followup[followup.regularity_status.isin(["منقطع", "متقطع"])] if not followup.empty and "regularity_status" in followup.columns else pd.DataFrame()
    if not urgent.empty:
        if not students.empty and "student_id" in students.columns and "full_name" in students.columns:
            urgent = urgent.merge(students[["student_id", "full_name"]], on="student_id", how="left")
        st.dataframe(urgent[["full_name", "followup_date", "notes"]], use_container_width=True)
    else:
        st.info("كل البنات منتظمات.")
    if role in ["System Admin", "Father Account", "Service Manager"]:
        st.markdown("---")
        st.subheader("🏆 أفضل فصل درجات في المسابقات")
        results = db.get_quiz_results()
        students_all = db.get_students()
        sections_all = db.get_sections()
        if not results.empty and "status" in results.columns and not students_all.empty and not sections_all.empty:
            submitted = results[results.status == "submitted"]
            if not submitted.empty:
                merged = submitted.merge(students_all[["student_id", "section_id"]], on="student_id", how="left")
                merged["score"] = pd.to_numeric(merged["score"], errors="coerce").fillna(0)
                if "section_id" in merged.columns:
                    section_scores = merged.groupby("section_id")["score"].mean().reset_index()
                    section_scores = section_scores.merge(sections_all[["section_id", "section_name"]], on="section_id", how="left")
                    if not section_scores.empty:
                        top_section = section_scores.sort_values("score", ascending=False).iloc[0]
                        st.metric(f"أفضل فصل: {top_section.get('section_name', '')}", f"{top_section.get('score', 0):.1f} / 20 متوسط")
                        st.dataframe(section_scores.rename(columns={"section_name": "الفصل", "score": "متوسط الدرجات"}).set_index("الفصل"), use_container_width=True)


# =============================================================================
# Members Cards Page (Unified)
# =============================================================================
def show_members_cards_page(db):
    inject_user_cards_css()
    st.markdown(hero_header("إدارة الأعضاء", "👥 إدارة جميع الأعضاء والطالبات"), unsafe_allow_html=True)
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    section_id = user.get("section_id", "")

    if role not in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
        st.error("🚫 غير مصرح")
        return

    # تشغيل الترحيل التلقائي لكود وكلمة مرور الطالبات
    try:
        migrated_count, migrate_msg = db.migrate_student_codes_and_passwords()
        if migrated_count and migrated_count > 0:
            st.success(f"✅ {migrate_msg}")
    except Exception:
        pass

    users = db.get_users()
    students = db.get_students()
    sections = db.get_sections()
    stages = db.get_stages()

    # Build unified members list (exclude System Admin and Father Account)
    members = []
    if not users.empty:
        for _, u in users.iterrows():
            u_role = u.get("role", "")
            if u_role in ["System Admin", "Father Account"]:
                continue
            # Teacher cannot see other teachers
            if role == "Teacher" and u_role == "Teacher" and u.get("user_id", "") != user_id:
                continue
            members.append({
                "member_id": u.get("user_id", ""),
                "full_name": u.get("full_name", ""),
                "role": u_role,
                "section_id": u.get("section_id", ""),
                "phone": u.get("phone", ""),
                "email": u.get("email", ""),
                "status": u.get("status", "active"),
                "type": "user",
                "created_by": u.get("created_by", "")
            })
    if not students.empty:
        for _, s in students.iterrows():
            # Teacher can only see students in their section
            if role == "Teacher" and section_id:
                if s.get("section_id", "") != section_id:
                    continue
            members.append({
                "member_id": s.get("student_id", ""),
                "full_name": s.get("full_name", ""),
                "role": "Student",
                "section_id": s.get("section_id", ""),
                "phone": s.get("phone", ""),
                "email": "",
                "status": s.get("status", "active"),
                "type": "student",
                "parent_phone": s.get("parent_phone", ""),
                "birthdate": s.get("birthdate", ""),
                "address": s.get("address", ""),
                "school": s.get("school", ""),
                "notes": s.get("notes", ""),
                "created_by": s.get("created_by", ""),
                "student_code": s.get("student_code", ""),
                "student_password": s.get("student_password", ""),
                "stage_id": s.get("stage_id", "")
            })

    members_df = pd.DataFrame(members) if members else pd.DataFrame(columns=["member_id", "full_name", "role", "section_id", "phone", "email", "status", "type"])

    # RBAC filtering based on stages for Service Manager
    if role == "Teacher" and user.get("section_id"):
        if not members_df.empty:
            members_df = members_df[members_df["section_id"] == user.get("section_id")]
    elif role == "Service Manager":
        section_ids = get_sections_for_supervisor(db, user_id)
        if section_ids and not members_df.empty:
            members_df = members_df[members_df["section_id"].isin(section_ids)]

    # Search
    search_term = st.text_input("🔍 بحث", placeholder="ابحث في الاسم والتليفون...", label_visibility="collapsed")

    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        role_options = ["الكل", "طالبة", "أمين خدمة", "مدرسة"]
        role_filter = st.selectbox("نوع العضو", role_options)
    with col2:
        status_options = ["الكل", "نشط", "غير نشط"]
        status_filter = st.selectbox("الحالة", status_options)
    with col3:
        section_options = ["الكل"] + sections["section_id"].tolist() if not sections.empty else ["الكل"]
        section_filter = st.selectbox("الفصل", section_options, format_func=lambda x: "الكل" if x == "الكل" else sections[sections.section_id == x]["section_name"].values[0] if not sections.empty else x)

    filtered = members_df.copy()
    if search_term and not filtered.empty:
        mask = pd.Series(False, index=filtered.index)
        for col in ["full_name", "phone"]:
            if col in filtered.columns:
                mask |= filtered[col].astype(str).str.contains(search_term, na=False, case=False)
        filtered = filtered[mask]
    # Translate Arabic filter values back to English for database filtering
    role_filter_map = {"الكل": "الكل", "طالبة": "Student", "أمين خدمة": "Service Manager", "مدرسة": "Teacher"}
    status_filter_map = {"الكل": "الكل", "نشط": "active", "غير نشط": "inactive"}
    
    if role_filter != "الكل" and not filtered.empty and "role" in filtered.columns:
        filtered = filtered[filtered["role"] == role_filter_map.get(role_filter, role_filter)]
    if status_filter != "الكل" and not filtered.empty and "status" in filtered.columns:
        filtered = filtered[filtered["status"] == status_filter_map.get(status_filter, status_filter)]
    if section_filter != "الكل" and not filtered.empty and "section_id" in filtered.columns:
        filtered = filtered[filtered["section_id"] == section_filter]

    st.markdown(f"<p style='text-align:left; color:#666;'>عدد الأعضاء: {len(filtered)}</p>", unsafe_allow_html=True)

    if not filtered.empty:
        # Bulk QR download button for admins
        if role == "System Admin":
            st.markdown("### 📷 بطاقات QR للحضور")
            if st.button("📥 تحميل بطاقات QR للفصل", use_container_width=True, key="bulk_qr_btn"):
                # Generate QR codes for all students in filtered results
                students_for_qr = []
                for _, m in filtered.iterrows():
                    if m.get("type") == "student":
                        students_for_qr.append({
                            "full_name": m.get("full_name", ""),
                            "student_code": m.get("student_code", ""),
                            "student_password": m.get("student_password", ""),
                            "section_id": m.get("section_id", "")
                        })
                
                if students_for_qr:
                    # Get section name
                    section_name = ""
                    if section_filter != "الكل" and not sections.empty:
                        sec_match = sections[sections.section_id == section_filter]
                        if not sec_match.empty:
                            section_name = sec_match.iloc[0].get("section_name", "")
                    
                    # Generate A4 QR page
                    qr_page_bytes = generate_a4_qr_printable_page(students_for_qr[:6], section_name)
                    st.download_button(
                        label="💾 تحميل صفحة A4 ببطاقات QR",
                        data=qr_page_bytes,
                        file_name=f"QR_cards_{section_name or 'students'}_{get_cairo_now().strftime('%Y-%m-%d')}.png",
                        mime="image/png",
                        use_container_width=True,
                        key="download_qr_a4"
                    )
            st.markdown("---")
        
        cols = st.columns(3)
        for idx, (_, m) in enumerate(filtered.iterrows()):
            col = cols[idx % 3]
            with col:
                mid = m.get("member_id", "")
                full_name = m.get("full_name", "غير معروف")
                member_role = m.get("role", "")
                sec_id = m.get("section_id", "")
                phone = m.get("phone", "")
                status = m.get("status", "active")
                member_type = m.get("type", "user")
                initials = get_initials(full_name)
                role_class = get_role_css_class(member_role)
                status_class = get_status_css_class(status)

                section_name = ""
                if not sections.empty and sec_id:
                    sec_match = sections[sections["section_id"] == sec_id]
                    if not sec_match.empty:
                        section_name = sec_match.iloc[0].get("section_name", "")

                role_label = {"Service Manager": "أمين خدمة", "Teacher": "مدرسة", "Student": "طالبة", "System Admin": "مدير نظام"}.get(member_role, member_role)
                status_label = {"active": "نشط", "inactive": "غير نشط"}.get(status, "نشط")

                if member_type == "student":
                    # Student card - show name, phone, status, section, student code, password
                    parent_phone = m.get("parent_phone", "")
                    birthdate = m.get("birthdate", "")
                    address = m.get("address", "")
                    school = m.get("school", "")
                    student_notes = m.get("notes", "")
                    student_code = m.get("student_code", "")
                    student_password = m.get("student_password", "")
                    # الحصول على المرحلة من الفصل
                    stage_name_card = ""
                    if not stages.empty:
                        sec_for_stage = sections[sections["section_id"] == sec_id] if not sections.empty and sec_id else pd.DataFrame()
                        if not sec_for_stage.empty:
                            stg_id = sec_for_stage.iloc[0].get("stage_id", "")
                            if stg_id:
                                stg_match = stages[stages["stage_id"] == stg_id]
                                if not stg_match.empty:
                                    stage_name_card = stg_match.iloc[0].get("stage_name", "")
                    
                    st.markdown(f"""
                    <div class='user-card' id='card-{mid}'>
                        <div class='card-badge {status_class}'>{status_label}</div>
                        <div style='display:flex; gap:1rem; align-items:center;'>
                            <div class='user-avatar'>{initials}</div>
                            <div>
                                <h3 style='margin:0;'>{full_name}</h3>
                                <span class='role-badge {role_class}'>{role_label}</span>
                                <span class='status-badge {status_class}'>{status_label}</span>
                            </div>
                        </div>
                        <div class='student-info-row' style='margin-top:0.8rem;'>
                            <span>🏫 {section_name if section_name else '—'} {('| 📚 ' + stage_name_card) if stage_name_card else ''}</span>
                        </div>
                        <div class='student-info-row'>
                            <span>🆔 الكود: <strong dir='ltr'>{student_code if student_code else '—'}</strong></span>
                        </div>
                        <div class='student-info-row'>
                            <span>🔑 كلمة المرور: <strong dir='ltr'>{student_password if student_password else '—'}</strong></span>
                        </div>
                        <div class='student-info-row'>
                            <span>📱 {phone if phone else '—'}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # QR download button for student card
                    if student_code:
                        qr_bytes = generate_student_id_card({
                            "full_name": full_name,
                            "student_code": student_code,
                            "student_password": student_password
                        }, section_name)
                        st.download_button(
                            label="📷 تحميل بطاقة QR",
                            data=qr_bytes,
                            file_name=f"QR_{full_name}_{student_code}.png",
                            mime="image/png",
                            use_container_width=True,
                            key=f"qr_dl_{mid}"
                        )
                else:
                    # User card - show name, phone, section
                    email = m.get("email", "")
                    st.markdown(f"""
                    <div class='user-card' id='card-{mid}'>
                        <div class='card-badge {status_class}'>{status_label}</div>
                        <div style='display:flex; gap:1rem; align-items:center;'>
                            <div class='user-avatar'>{initials}</div>
                            <div>
                                <h3 style='margin:0;'>{full_name}</h3>
                                <span class='role-badge {role_class}'>{role_label}</span>
                                <span class='status-badge {status_class}'>{status_label}</span>
                            </div>
                        </div>
                        <div class='student-info-row' style='margin-top:0.8rem;'>
                            <span>📱 {phone if phone else '—'}</span>
                        </div>
                        <div class='student-info-row'>
                            <span>🏫 {section_name if section_name else '—'}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                # Action buttons
                action_cols = st.columns(5)
                with action_cols[0]:
                    if st.button("📋", key=f"view_{mid}"):
                        st.session_state.profile_user_id = mid
                        st.rerun()
                
                # Check if teacher can edit/delete
                can_edit_delete = True
                is_own_data = True
                if role == "Teacher":
                    if member_type == "student":
                        created_by = m.get("created_by", "")
                        is_own_data = (created_by == user_id)
                        can_edit_delete = is_own_data
                    else:
                        # Teacher cannot edit other teachers
                        can_edit_delete = False
                        st.warning("⛔ لا يمكنك التعديل على بيانات شخص آخر. هذه البيانات تخص مستخدم آخر.")
                
                if role == "System Admin" and can_edit_delete:
                    with action_cols[1]:
                        if status == "active":
                            if st.button("⏸️", key=f"deact_{mid}"):
                                if member_type == "student":
                                    db.update_student(mid, {"status": "inactive"})
                                else:
                                    db.update_user(mid, {"status": "inactive"})
                                db.add_log(user.get("user_id", ""), "تعطيل عضو", f"تم تعطيل {full_name}")
                                st.success("✅ تم التعطيل")
                                time.sleep(1)
                                st.rerun()
                        else:
                            if st.button("▶️", key=f"act_{mid}"):
                                if member_type == "student":
                                    db.update_student(mid, {"status": "active"})
                                else:
                                    db.update_user(mid, {"status": "active"})
                                db.add_log(user.get("user_id", ""), "تفعيل عضو", f"تم تفعيل {full_name}")
                                st.success("✅ تم التفعيل")
                                time.sleep(1)
                                st.rerun()
                    with action_cols[2]:
                        if st.button("✏️", key=f"edit_{mid}"):
                            st.session_state[f"edit_mode_{mid}"] = True
                    with action_cols[3]:
                        if st.button("🗑️", key=f"del_{mid}"):
                            confirm = st.checkbox(f"تأكيد الحذف؟", key=f"confirm_del_{mid}")
                            if confirm:
                                if member_type == "student":
                                    db.delete_student(mid)
                                else:
                                    db.delete_user(mid)
                                db.add_log(user.get("user_id", ""), "حذف عضو", f"تم حذف {full_name}")
                                st.success("✅ تم الحذف")
                                time.sleep(1)
                                st.rerun()
                elif role == "Teacher" and can_edit_delete and member_type == "student":
                    with action_cols[2]:
                        if st.button("✏️", key=f"edit_{mid}"):
                            st.session_state[f"edit_mode_{mid}"] = True
                    with action_cols[3]:
                        if st.button("🗑️", key=f"del_{mid}"):
                            confirm = st.checkbox(f"تأكيد الحذف؟", key=f"confirm_del_{mid}")
                            if confirm:
                                db.delete_student(mid)
                                db.add_log(user.get("user_id", ""), "حذف طالبة", f"تم حذف {full_name}")
                                st.success("✅ تم الحذف")
                                time.sleep(1)
                                st.rerun()

                # Edit form
                if st.session_state.get(f"edit_mode_{mid}", False):
                    with st.expander("✏️ تعديل البيانات", expanded=True):
                        with st.form(f"edit_member_form_{mid}"):
                            edit_name = st.text_input("الاسم الكامل*", value=full_name)
                            edit_phone = st.text_input("الهاتف", value=phone)
                            edit_email = st.text_input("البريد الإلكتروني", value=phone if member_role in ["Service Manager", "Teacher"] else "")
                            sec_options = sections["section_id"].tolist() if not sections.empty else []
                            current_sec = sec_id if sec_id in sec_options else (sec_options[0] if sec_options else "")
                            edit_section = st.selectbox("الفصل", sec_options, index=sec_options.index(current_sec) if current_sec in sec_options else 0, format_func=lambda x: sections[sections.section_id == x]["section_name"].values[0]) if sec_options else ""
                            
                            # Student specific fields
                            edit_parent_phone = ""
                            edit_birthdate = None
                            edit_address = ""
                            edit_school = ""
                            edit_notes = ""
                            edit_status = "active"
                            if member_type == "student":
                                edit_parent_phone = st.text_input("رقم ولي الأمر", value=m.get("parent_phone", ""))
                                bd_value = pd.to_datetime(m.get("birthdate", "")).date() if m.get("birthdate") else None
                                edit_birthdate = st.date_input("تاريخ الميلاد", value=bd_value)
                                edit_address = st.text_input("العنوان", value=m.get("address", ""))
                                edit_school = st.text_input("المدرسة", value=m.get("school", ""))
                                edit_notes = st.text_area("ملاحظات", value=m.get("notes", ""))
                                edit_status = st.selectbox("الحالة", ["نشطة", "غير نشطة"], index=0 if m.get("status", "active") == "active" else 1)
                            
                            if st.form_submit_button("💾 حفظ"):
                                if member_type == "student":
                                    db.update_student(mid, {
                                        "full_name": edit_name,
                                        "phone": edit_phone,
                                        "section_id": edit_section,
                                        "parent_phone": edit_parent_phone,
                                        "birthdate": edit_birthdate.strftime("%Y-%m-%d") if edit_birthdate else "",
                                        "address": edit_address,
                                        "school": edit_school,
                                        "notes": edit_notes,
                                        "status": "active" if edit_status == "نشطة" else "inactive"
                                    })
                                else:
                                    db.update_user(mid, {"full_name": edit_name, "phone": edit_phone, "email": edit_email, "section_id": edit_section})
                                db.add_log(user.get("user_id", ""), "تعديل عضو", f"تم تعديل {edit_name}")
                                st.session_state[f"edit_mode_{mid}"] = False
                                st.success("✅ تم التحديث")
                                time.sleep(1)
                                st.rerun()
                st.markdown("---")

    # Add new member
    with st.expander("➕ إضافة عضو جديد"):
        with st.form("add_member_form"):
            member_type = st.selectbox("نوع العضو", ["طالبة", "أمين خدمة", "مدرسة"])
            new_name = st.text_input("الاسم الكامل*")
            new_phone = st.text_input("الهاتف")
            new_section_id = ""
            if not sections.empty:
                new_section_id = st.selectbox("الفصل", sections["section_id"], format_func=lambda x: sections[sections.section_id == x]["section_name"].values[0])
            
            # Student specific fields
            new_parent_phone = ""
            new_birthdate = None
            new_address = ""
            new_school = ""
            new_notes = ""
            new_status = "نشطة"
            if member_type == "طالبة":
                new_parent_phone = st.text_input("رقم ولي الأمر")
                new_birthdate = st.date_input("تاريخ الميلاد", value=None)
                new_address = st.text_input("العنوان")
                new_school = st.text_input("المدرسة")
                new_notes = st.text_area("ملاحظات")
                new_status = st.selectbox("الحالة", ["نشطة", "غير نشطة"], index=0)

            submitted = st.form_submit_button("إضافة عضو")
            if submitted:
                if not new_name:
                    st.error("الاسم الكامل مطلوب")
                else:
                    if member_type == "طالبة":
                        student_data = {
                            "student_id": str(uuid.uuid4()),
                            "full_name": new_name.strip(),
                            "section_id": new_section_id,
                            "phone": new_phone,
                            "parent_phone": new_parent_phone,
                            "birthdate": new_birthdate.strftime("%Y-%m-%d") if new_birthdate else "",
                            "address": new_address,
                            "school": new_school,
                            "notes": new_notes,
                            "status": "active" if new_status == "نشطة" else "inactive",
                            "created_by": user_id
                        }
                        db.add_student(student_data)
                    else:
                        role_map = {"أمين خدمة": "Service Manager", "مدرسة": "Teacher"}
                        db.add_user({
                            "user_id": str(uuid.uuid4()),
                            "username": new_name.strip(),
                            "password": hash_password("1234"),
                            "role": role_map.get(member_type, "Teacher"),
                            "full_name": new_name.strip(),
                            "section_id": new_section_id,
                            "phone": new_phone,
                            "email": ""
                        })
                    db.add_log(user.get("user_id", ""), "إضافة عضو", f"تمت إضافة {new_name}")
                    st.success("✅ تمت الإضافة بنجاح")
                    time.sleep(1)
                    st.rerun()


# =============================================================================
# User Management (Deprecated)
# =============================================================================
# DEPRECATED - replaced by show_members_cards_page
def show_user_management(db):
    st.warning("⚠️ هذه الصفحة تم استبدالها بصفحة إدارة الأعضاء الموحدة")
    st.info("يرجى استخدام صفحة 👥 إدارة الأعضاء من القائمة الجانبية")


# =============================================================================
# Students Management (Cards) (Deprecated)
# =============================================================================
# DEPRECATED - replaced by show_members_cards_page
def show_students_cards_page(db):
    st.warning("⚠️ هذه الصفحة تم استبدالها بصفحة إدارة الأعضاء الموحدة")
    st.info("يرجى استخدام صفحة 👥 إدارة الأعضاء من القائمة الجانبية")


# =============================================================================
# Stages Management Page (Standalone)
# =============================================================================
def show_stages_page(db):
    inject_user_cards_css()
    st.markdown(hero_header("إدارة المراحل الدراسية", "🏫 إدارة مراحل الدراسة والملاحظات"), unsafe_allow_html=True)
    user = st.session_state.user
    role = user.get("role", "")
    users = db.get_users()
    stages = db.get_stages()
    sections = db.get_sections()
    students = db.get_students()

    if role == "Father Account":
        st.info("👁️ وضع العرض فقط - يمكنك مراجعة المراحل")
    elif role not in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
        st.error("🚫 غير مصرح")
        return

    if not stages.empty:
        stage_view = stages.copy()
        if "stage_id" in stage_view.columns:
            stage_view["num_sections"] = stage_view["stage_id"].apply(
                lambda sid: len(sections[sections["stage_id"] == sid]) if not sections.empty and "stage_id" in sections.columns else 0
            )
            stage_view["num_students"] = stage_view["stage_id"].apply(
                lambda sid: len(students[students["section_id"].isin(sections[sections["stage_id"] == sid]["section_id"])]) if not sections.empty and not students.empty and "stage_id" in sections.columns else 0
            )
        else:
            stage_view["num_sections"] = 0
            stage_view["num_students"] = 0
        if not users.empty:
            stage_view["supervisors"] = stage_view["stage_id"].apply(
                lambda sid: ", ".join(db.get_supervisor_names_for_stage(sid, users)) if not stages.empty else ""
            )
        else:
            stage_view["supervisors"] = ""
        view_cols = ["stage_name", "status", "num_sections", "num_students", "supervisors"]
        available_cols = [c for c in view_cols if c in stage_view.columns]
        st.dataframe(stage_view[available_cols].rename(columns={"stage_name": "المرحلة", "status": "الحالة", "num_sections": "عدد الفصول", "num_students": "عدد الطلاب", "supervisors": "المشرفون"}), use_container_width=True)
    else:
        st.info("لا توجد مراحل مسجلة.")

    with st.expander("➕ إضافة مرحلة جديدة"):
        with st.form("add_stage_page_form"):
            new_name = st.text_input("اسم المرحلة*")
            new_desc = st.text_area("الوصف")
            new_order = st.number_input("ترتيب العرض", 1, 100, 1)
            new_status = st.selectbox("الحالة", ["active", "inactive"], index=0)
            new_notes = st.text_area("ملاحظات")
            if st.form_submit_button("إضافة المرحلة"):
                if not new_name:
                    st.error("اسم المرحلة مطلوب")
                elif not stages.empty and "stage_name" in stages.columns and not stages[stages.stage_name == new_name.strip()].empty:
                    st.error("⛔ اسم المرحلة موجود مسبقاً!")
                else:
                    db.add_stage({
                        "stage_id": str(uuid.uuid4()), "stage_name": new_name.strip(),
                        "description": new_desc, "display_order": new_order,
                        "status": new_status, "created_date": get_cairo_now().strftime("%Y-%m-%d"),
                        "created_by": user.get("user_id", ""), "manager_user_id": "",
                        "notes": new_notes
                    })
                    st.success("✅ تمت إضافة المرحلة بنجاح")
                    time.sleep(1)
                    st.rerun()

    if not stages.empty:
        with st.expander("✏️ تعديل / حذف مرحلة"):
            stage_sel = st.selectbox("اختر مرحلة", stages["stage_id"],
                                     format_func=lambda x: stages[stages.stage_id == x]["stage_name"].values[0])
            stage_row = stages[stages.stage_id == stage_sel].iloc[0].to_dict()
            edit_name = st.text_input("اسم المرحلة", value=stage_row.get("stage_name", ""), key="page_edit_stage_name")
            edit_desc = st.text_area("الوصف", value=stage_row.get("description", ""), key="page_edit_stage_desc")
            edit_order = st.number_input("ترتيب العرض", 1, 100, int(stage_row.get("display_order", 1) or 1), key="page_edit_stage_order")
            edit_status = st.selectbox("الحالة", ["active", "inactive"], index=0 if stage_row.get("status", "active") == "active" else 1, key="page_edit_stage_status")
            edit_notes = st.text_area("ملاحظات", value=stage_row.get("notes", ""), key="page_edit_stage_notes")

            st.markdown("#### 👥 المشرفون على المرحلة")
            current_supervisors = db.get_supervisors_for_stage(stage_sel)
            eligible_users = users[users.role.isin(["Service Manager", "Teacher", "Father Account", "System Admin"])] if not users.empty else pd.DataFrame()
            if not eligible_users.empty:
                supervisor_options = eligible_users["user_id"].tolist()
                # Filter default to only include IDs that exist in options
                valid_supervisors = [s for s in current_supervisors if s in supervisor_options]
                selected_supervisors = st.multiselect("اختر المشرفين", supervisor_options,
                                                      default=valid_supervisors,
                                                      format_func=lambda x: eligible_users[eligible_users.user_id == x]["full_name"].values[0] if not eligible_users.empty and x in eligible_users["user_id"].values else str(x),
                                                      key="page_stage_supervisors")
            else:
                selected_supervisors = []
                st.info("لا يوجد مستخدمون مؤهلون")

            col1, col2 = st.columns(2)
            if col1.button("تحديث المرحلة"):
                db.update_stage(stage_sel, {"stage_name": edit_name, "description": edit_desc,
                                            "display_order": edit_order, "status": edit_status, "notes": edit_notes})
                db.clear_stage_supervisors(stage_sel)
                for sup_id in selected_supervisors:
                    db.add_stage_supervisor(stage_sel, sup_id)
                st.success("تم التحديث")
                time.sleep(1)
                st.rerun()
            if col2.button("حذف المرحلة"):
                sections_in_stage = sections[sections["stage_id"] == stage_sel] if not sections.empty and "stage_id" in sections.columns else pd.DataFrame()
                if not sections_in_stage.empty:
                    st.error("❌ لا يمكن حذف المرحلة لأنها تحتوي على فصول. قم بحذف الفصول أولاً.")
                else:
                    db.delete_stage(stage_sel)
                    db.clear_stage_supervisors(stage_sel)
                    st.success("تم حذف المرحلة")
                    time.sleep(1)
                    st.rerun()


# =============================================================================
# Sections Management Page (Standalone)
# =============================================================================
def show_sections_page(db):
    inject_user_cards_css()
    st.markdown(hero_header("إدارة الفصول", "📚 إدارة بيانات الفصول الدراسية"), unsafe_allow_html=True)
    user = st.session_state.user
    role = user.get("role", "")
    users = db.get_users()
    stages = db.get_stages()
    sections = db.get_sections()
    students = db.get_students()

    if role == "Father Account":
        st.info("👁️ وضع العرض فقط - يمكنك مراجعة الفصول")
    elif role in ["Service Manager", "Teacher"] and not sections.empty:
        assigned = db.get_sections_by_teacher(user.get("user_id", "")) if role == "Teacher" else db.get_sections_by_leader(user.get("user_id", ""))
        if assigned.empty:
            assigned = db.get_sections_by_teacher(user.get("user_id", ""))
        sections = assigned if not assigned.empty else sections
    elif role not in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
        st.error("🚫 غير مصرح")
        return

    def get_names_from_ids(id_string, users_df, role_name):
        if not id_string or users_df.empty:
            return f"غير محدد{role_name}"
        ids = [x.strip() for x in str(id_string).split(",") if x.strip()]
        names = []
        for uid in ids:
            match = users_df[users_df["user_id"] == uid]
            if not match.empty:
                names.append(match.iloc[0].get("full_name", uid))
        return "، ".join(names) if names else f"غير محدد{role_name}"

    search_term = st.text_input("🔍 بحث", placeholder="ابحث باسم الفصل...", label_visibility="collapsed")
    col1, col2, col3 = st.columns(3)
    with col1:
        stage_options = ["الكل"] + stages["stage_id"].tolist() if not stages.empty else ["الكل"]
        stage_filter = st.selectbox("المرحلة", stage_options,
                                    format_func=lambda x: "الكل" if x == "الكل" else stages[stages.stage_id == x]["stage_name"].values[0] if not stages.empty else x)
    with col2:
        teacher_options = ["الكل"] + users["user_id"].tolist() if not users.empty else ["الكل"]
        teacher_filter = st.selectbox("المدرس", teacher_options,
                                      format_func=lambda x: "الكل" if x == "الكل" else users[users.user_id == x]["full_name"].values[0] if not users.empty else x)
    with col3:
        status_filter = st.selectbox("الحالة", ["الكل", "active", "inactive"])

    filtered = sections.copy() if not sections.empty else pd.DataFrame()
    if search_term and not filtered.empty:
        mask = pd.Series(False, index=filtered.index)
        for col in ["section_name"]:
            if col in filtered.columns:
                mask |= filtered[col].astype(str).str.contains(search_term, na=False, case=False)
        filtered = filtered[mask]
    if stage_filter != "الكل" and not filtered.empty and "stage_id" in filtered.columns:
        filtered = filtered[filtered["stage_id"] == stage_filter]
    if teacher_filter != "الكل" and not filtered.empty and "teacher_id" in filtered.columns:
        filtered = filtered[filtered["teacher_id"] == teacher_filter]
    if status_filter != "الكل" and not filtered.empty and "status" in filtered.columns:
        filtered = filtered[filtered["status"] == status_filter]

    if not filtered.empty:
        for _, sec in filtered.iterrows():
            sec_id = sec.get("section_id", "")
            sec_name = sec.get("section_name", "")
            sec_stage = sec.get("stage_id", "")
            sec_teacher = sec.get("teacher_id", "")
            sec_leader = sec.get("leader_id", "")
            sec_status = sec.get("status", "active")
            sec_notes = sec.get("notes", "")

            if not stages.empty and sec_stage:
                matched_stage = stages[stages["stage_id"] == sec_stage]["stage_name"]
                stage_name = matched_stage.values[0] if not matched_stage.empty else "—"
            else:
                stage_name = "—"
            teacher_name = get_names_from_ids(sec_teacher, users, "")
            leader_name = get_names_from_ids(sec_leader, users, "ة")
            student_count = db.get_section_student_count(sec_id)

            with st.expander(f"🏫 {sec_name} ({stage_name}) - {student_count} طالبة"):
                st.markdown(f"**المرحلة:** {stage_name} | **الحالة:** {sec_status}")
                st.markdown(f"**المدرس:** {teacher_name} | **أمين الخدمة:** {leader_name}")
                st.markdown(f"**عدد الطلاب:** {student_count}")
                st.markdown(f"**ملاحظات:** {sec_notes or '—'}")

                if role == "System Admin":
                    st.markdown("#### 🏫 نقل الفصل بين المراحل")
                    if not stages.empty:
                        new_stage = st.selectbox("نقل إلى مرحلة", stages["stage_id"],
                                                 format_func=lambda x: "—" if x not in stages["stage_id"].values else stages[stages.stage_id == x]["stage_name"].values[0],
                                                 key=f"move_stage_{sec_id}")
                        if st.button("نقل الفصل", key=f"move_sec_{sec_id}"):
                            db.update_section(sec_id, {"stage_id": new_stage})
                            st.success("✅ تم نقل الفصل")
                            time.sleep(1)
                            st.rerun()

                    st.markdown("#### 🔗 التعيينات")
                    eligible_teachers = users[users.role == "Teacher"] if not users.empty else pd.DataFrame()
                    eligible_leaders = users[users.role == "Service Manager"] if not users.empty else pd.DataFrame()
                    current_teachers_raw = str(sec_teacher).split(",") if sec_teacher else []
                    current_leaders_raw = str(sec_leader).split(",") if sec_leader else []
                    valid_teacher_ids = eligible_teachers["user_id"].tolist() if not eligible_teachers.empty else []
                    valid_leader_ids = eligible_leaders["user_id"].tolist() if not eligible_leaders.empty else []
                    # Filter defaults to only include IDs that still exist in the options
                    current_teachers = [t.strip() for t in current_teachers_raw if t.strip() and t.strip() in valid_teacher_ids]
                    current_leaders = [l.strip() for l in current_leaders_raw if l.strip() and l.strip() in valid_leader_ids]

                    selected_teachers = st.multiselect("المدرسات", valid_teacher_ids,
                                                       default=current_teachers,
                                                       format_func=lambda x: eligible_teachers[eligible_teachers.user_id == x]["full_name"].values[0] if not eligible_teachers.empty and x in eligible_teachers["user_id"].values else str(x),
                                                       key=f"teacher_{sec_id}")
                    selected_leaders = st.multiselect("أمناء الخدمة", valid_leader_ids,
                                                      default=current_leaders,
                                                      format_func=lambda x: eligible_leaders[eligible_leaders.user_id == x]["full_name"].values[0] if not eligible_leaders.empty and x in eligible_leaders["user_id"].values else str(x),
                                                      key=f"leader_{sec_id}")
                    if st.button("💾 حفظ التعيينات", key=f"save_assign_{sec_id}"):
                        db.update_section(sec_id, {"teacher_id": ",".join(selected_teachers), "leader_id": ",".join(selected_leaders)})
                        st.success("✅ تم تحديث التعيينات")
                        time.sleep(1)
                        st.rerun()

                    if role == "System Admin":
                        st.markdown("#### ✏️ تعديل / حذف الفصل")
                        new_name = st.text_input("اسم الفصل", value=sec_name, key=f"edit_name_{sec_id}")
                        new_status = st.selectbox("الحالة", ["active", "inactive"], index=0 if sec_status == "active" else 1, key=f"edit_status_{sec_id}")
                        new_notes = st.text_area("ملاحظات", value=sec_notes, key=f"edit_notes_{sec_id}")
                        c1, c2 = st.columns(2)
                        if c1.button("تحديث الفصل", key=f"update_sec_{sec_id}"):
                            db.update_section(sec_id, {"section_name": new_name, "notes": new_notes, "status": new_status})
                            st.success("✅ تم التحديث")
                            time.sleep(1)
                            st.rerun()
                        if c2.button("حذف الفصل", key=f"delete_sec_{sec_id}"):
                            section_students = students[students["section_id"] == sec_id] if not students.empty and "section_id" in students.columns else pd.DataFrame()
                            if not section_students.empty:
                                st.error("❌ لا يمكن حذف الفصل لأنه يحتوي على طالبات. قم بنقل الطالبات أولاً.")
                            else:
                                db.delete_section(sec_id)
                                st.success("✅ تم الحذف")
                                time.sleep(1)
                                st.rerun()
                    else:
                        st.info("👁️ يمكنك فقط مشاهدة بيانات الفصل - لا يمكنك تعديلها أو حذفها")
    else:
        st.info("🔍 لا توجد فصول مطابقة.")

    if role == "System Admin":
        with st.expander("➕ إضافة فصل جديد"):
            with st.form("add_section_page_form"):
                sec_name = st.text_input("اسم الفصل*")
                sec_stage = st.selectbox("المرحلة", stages["stage_id"],
                                         format_func=lambda x: "—" if not stages.empty and x not in stages["stage_id"].values else stages[stages.stage_id == x]["stage_name"].values[0]) if not stages.empty else ""
                eligible_teachers = users[users.role == "Teacher"] if not users.empty else pd.DataFrame()
                eligible_leaders = users[users.role == "Service Manager"] if not users.empty else pd.DataFrame()
                teacher_opts = eligible_teachers["user_id"].tolist() if not eligible_teachers.empty else []
                selected_teachers = st.multiselect("المدرسات", teacher_opts,
                                                   format_func=lambda x: eligible_teachers[eligible_teachers.user_id == x]["full_name"].values[0] if not eligible_teachers.empty and x in eligible_teachers["user_id"].values and not eligible_teachers[eligible_teachers.user_id == x].empty else str(x)) if teacher_opts else []
                leader_opts = eligible_leaders["user_id"].tolist() if not eligible_leaders.empty else []
                selected_leaders = st.multiselect("أمناء الخدمة", leader_opts,
                                                  format_func=lambda x: eligible_leaders[eligible_leaders.user_id == x]["full_name"].values[0] if not eligible_leaders.empty and x in eligible_leaders["user_id"].values and not eligible_leaders[eligible_leaders.user_id == x].empty else str(x)) if leader_opts else []
                sec_notes = st.text_area("ملاحظات")
                if st.form_submit_button("إضافة الفصل"):
                    if not sec_name:
                        st.error("اسم الفصل مطلوب")
                    elif not sections.empty and "section_name" in sections.columns and sec_stage and not sections[sections.section_name == sec_name.strip()].empty:
                        st.error("⛔ اسم الفصل موجود مسبقاً!")
                    else:
                        teacher_ids_str = ",".join(selected_teachers) if selected_teachers else ""
                        leader_ids_str = ",".join(selected_leaders) if selected_leaders else ""
                        db.add_section({
                            "section_id": str(uuid.uuid4()), "section_name": sec_name.strip(),
                            "stage_id": sec_stage, "teacher_id": teacher_ids_str, "leader_id": leader_ids_str,
                            "max_students": "", "room": "",
                            "meeting_day": "", "meeting_time": "",
                            "status": "active", "notes": sec_notes,
                            "manager_user_id": user.get("user_id", "")
                        })
                        st.success("✅ تمت إضافة الفصل بنجاح")
                        time.sleep(1)
                        st.rerun()


# =============================================================================
# Attendance
# =============================================================================
def show_attendance(db):
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    st.markdown(hero_header("تسجيل الحضور", "📋 تسجيل ومتابعة حضور الطالبات"), unsafe_allow_html=True)
    
    # Service Manager can view attendance for their sections but not edit
    if role == "Service Manager":
        section_ids = get_sections_for_supervisor(db, user_id)
        if not section_ids:
            st.info("لا توجد فصول معينة لك.")
            return
        sections = db.get_sections()
        if sections.empty:
            st.warning("لا توجد فصول.")
            return
        supervised_sections = sections[sections.section_id.isin(section_ids)]
        if supervised_sections.empty:
            st.info("لا توجد فصول معينة لك.")
            return
        
        st.subheader("📊 عرض الحضور (للقراءة فقط)")
        selected_section = st.selectbox("اختر الفصل", supervised_sections["section_id"],
                                        format_func=lambda x: "—" if x not in supervised_sections["section_id"].values else supervised_sections[supervised_sections.section_id == x]["section_name"].values[0])
        date = st.date_input("التاريخ", get_cairo_now().date())
        date_str = date.strftime("%Y-%m-%d")
        students = db.get_students()
        section_students = students[students.section_id == selected_section] if not students.empty and "section_id" in students.columns else pd.DataFrame()
        if section_students.empty:
            st.info("لا توجد طالبات في هذا الفصل.")
            return
        existing = db.get_attendance_by_date_section(date_str, selected_section)
        if existing.empty:
            st.info("لا يوجد سجل حضور لهذا اليوم.")
            return
        # Merge student names
        display = existing.merge(section_students[["student_id", "full_name"]], on="student_id", how="left")
        st.dataframe(display[["full_name", "status", "notes"]], use_container_width=True)
        return
    
    # Teacher and System Admin flow continues below
    sections = db.get_sections()
    if sections.empty:
        st.warning("لا توجد فصول.")
        return
    section_id = user.get("section_id", "")
    if role == "Teacher" and section_id:
        selected_section = section_id
        section_name = sections[sections.section_id == section_id]["section_name"].values[0] if not sections.empty else section_id
        st.markdown(f"**الفصل:** {section_name}")
    else:
        selected_section = st.selectbox("اختر الفصل", sections["section_id"],
                                        format_func=lambda x: sections[sections.section_id == x]["section_name"].values[0])
    date = st.date_input("التاريخ", get_cairo_now().date())
    date_str = date.strftime("%Y-%m-%d")
    students = db.get_students()
    section_students = students[students.section_id == selected_section] if not students.empty and "section_id" in students.columns else pd.DataFrame()
    if section_students.empty:
        st.info("لا توجد طالبات في هذا الفصل.")
        return
    existing = db.get_attendance_by_date_section(date_str, selected_section)
    already_filled = not existing.empty
    if already_filled:
        st.warning("⚠️ يوجد تسجيل حضور سابق.")
    statuses = {}
    notes_dict = {}
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    for _, s in section_students.iterrows():
        sid = s["student_id"]
        sname = s["full_name"]
        prev = existing[existing.student_id == sid] if already_filled else pd.DataFrame()
        prev_status = prev.iloc[0]["status"] if not prev.empty else "حاضر"
        prev_notes = prev.iloc[0]["notes"] if not prev.empty else ""
        cols = st.columns([3, 2, 2])
        cols[0].markdown(f"**{sname}**")
        status_list = ["حاضر", "غائب", "متأخر"]
        status_index = status_list.index(prev_status) if prev_status in status_list else 0
        status = cols[1].radio("الحالة", status_list, index=status_index, key=f"att_{sid}", horizontal=True)
        notes = cols[2].text_input("ملاحظة", value=prev_notes, key=f"note_{sid}", label_visibility="collapsed")
        statuses[sid] = status
        notes_dict[sid] = notes
    st.markdown("</div>", unsafe_allow_html=True)
    if st.button("💾 حفظ الحضور", use_container_width=True, key="save_attendance_btn"):
        with st.spinner("جاري حفظ الحضور..."):
            records = []
            for sid, status in statuses.items():
                prev_record = existing[existing.student_id == sid] if already_filled else pd.DataFrame()
                record_id = prev_record.iloc[0]["record_id"] if not prev_record.empty else str(uuid.uuid4())
                records.append({
                    "record_id": record_id, "date": date_str, "student_id": sid,
                    "status": status, "notes": notes_dict.get(sid, ""),
                    "recorded_by": user.get("user_id", ""), "section_id": selected_section
                })
            db.batch_add_attendance(records)
            db.add_log(user.get("user_id", ""), f"تسجيل حضور فصل {selected_section} ليوم {date_str}")
            st.success("✅ تم تسجيل الحضور بنجاح")
            time.sleep(1)
            st.rerun()
    if not existing.empty:
        st.markdown("---")
        st.subheader("🗑️ إدارة سجلات الحضور السابقة")
        rec = existing.copy()
        rec["student_name"] = rec["student_id"].apply(
            lambda sid: section_students[section_students.student_id == sid]["full_name"].values[0]
            if sid in section_students["student_id"].values else sid
        )
        rec = rec[["record_id", "student_name", "status", "notes"]]
        st.dataframe(rec, use_container_width=True)
        
        # Teacher can only delete attendance records they created
        can_delete_attendance = True
        if role == "Teacher":
            can_delete_attendance = False
            st.warning("⛔ لا يمكنك حذف سجل حضور سجلته مدرسة أخرى.")
        
        if role == "System Admin" and can_delete_attendance:
            del_id = st.selectbox("اختر سجل حضور لحذفه", rec["record_id"], key="del_att_sel")
            if st.button("حذف سجل الحضور"):
                db.delete_attendance_record(del_id)
                st.success("تم الحذف")
                time.sleep(1)
                st.rerun()


# =============================================================================
# Follow-up
# =============================================================================
def show_followup(db):
    st.markdown(hero_header("متابعة الافتقاد", "💬 متابعة حالة الطالبات المنتظمات"), unsafe_allow_html=True)
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    students = db.get_students()
    followup = db.get_followup()
    
    # Service Manager filtering by stages
    if role == "Service Manager" and db and user_id:
        section_ids = get_sections_for_supervisor(db, user_id)
        if section_ids:
            responsible = students[students.section_id.isin(section_ids)] if not students.empty and "section_id" in students.columns else pd.DataFrame()
        else:
            responsible = pd.DataFrame()
    else:
        section_id = user.get("section_id", "")
        responsible = filter_students_by_role(students, role, section_id)
    if responsible.empty:
        st.info("لا توجد طالبات مسؤولات عنك.")
        return
    if not followup.empty and "student_id" in followup.columns and "regularity_status" in followup.columns:
        student_ids = responsible["student_id"].tolist() if "student_id" in responsible.columns else []
        fups = followup[followup.student_id.isin(student_ids)]
        regular = len(fups[fups.regularity_status == "منتظم"])
        intermittent = len(fups[fups.regularity_status == "متقطع"])
        disconnected = len(fups[fups.regularity_status == "منقطع"])
    else:
        regular = intermittent = disconnected = 0
    col1, col2, col3 = st.columns(3)
    col1.metric("منتظمات", regular)
    col2.metric("متقطعات", intermittent)
    col3.metric("منقطعات", disconnected)
    st.markdown("---")
    st.subheader("⚠️ بنات بحاجة إلى افتقاد")
    if not followup.empty and "regularity_status" in followup.columns and "student_id" in followup.columns:
        urgent = followup[(followup.regularity_status.isin(["متقطع", "منقطع"])) & (followup.student_id.isin(responsible["student_id"]))]
        if not urgent.empty:
            urgent_display = urgent.merge(responsible[["student_id", "full_name"]], on="student_id", how="left")
            st.dataframe(urgent_display[["full_name", "followup_date", "notes"]], use_container_width=True)
        else:
            st.info("كل البنات منتظمات حالياً.")
    else:
        st.info("لا توجد متابعات سابقة.")
    st.markdown("---")
    st.subheader("➕ إضافة متابعة جديدة")
    if "student_id" in responsible.columns:
        student = st.selectbox("اختر الطالبة", responsible["student_id"],
                               format_func=lambda x: responsible[responsible.student_id == x]["full_name"].values[0], key="followup_student")
        with st.form("followup_form"):
            ftype = st.selectbox("نوع الافتقاد", ["زيارة", "اتصال هاتفي", "رسالة", "لقاء شخصي"])
            notes = st.text_area("ملاحظات")
            regularity = st.selectbox("حالة الانتظام", ["منتظم", "متقطع", "منقطع"])
            if st.form_submit_button("حفظ المتابعة"):
                try:
                    db.add_followup_record({
                        "record_id": str(uuid.uuid4()), "student_id": student,
                        "teacher_id": user.get("user_id", ""), "followup_date": get_cairo_now().strftime("%Y-%m-%d"),
                        "followup_type": ftype, "notes": notes, "regularity_status": regularity
                    })
                    st.success("✅ تم تسجيل الافتقاد بنجاح")
                    time.sleep(1)
                    st.rerun()
                except ValueError as e:
                    st.error(str(e))
    
    # Show existing followup records with delete option for admin
    if not followup.empty and role == "System Admin":
        st.markdown("---")
        st.subheader("🗑️ إدارة سجلات الافتقاد")
        followup_display = followup.merge(
            students[["student_id", "full_name"]] if not students.empty else pd.DataFrame(),
            on="student_id", how="left"
        )
        if not followup_display.empty:
            followup_display = followup_display[["record_id", "full_name", "followup_type", "regularity_status"]]
            st.dataframe(followup_display, use_container_width=True)
            del_followup_id = st.selectbox("اختر سجل افتقاد لحذفه", followup_display["record_id"], key="del_followup_sel")
            if st.button("حذف سجل الافتقاد"):
                db.delete_followup_record(del_followup_id)
                st.success("تم الحذف")
                time.sleep(1)
                st.rerun()


# =============================================================================
# Class Competition Scores
# =============================================================================
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    
    # Service Manager can view competition scores for their stages
    if role == "Service Manager":
        section_ids = get_sections_for_supervisor(db, user_id)
        if not section_ids:
            st.info("لا توجد فصول معينة لك.")
            return
        students = db.get_students()
        quizzes = db.get_quizzes()
        results = db.get_quiz_results()
        section_students = students[students.section_id.isin(section_ids)] if not students.empty and "section_id" in students.columns else pd.DataFrame()
    elif role == "Teacher":
        section_id = user.get("section_id", "")
        if not section_id:
            st.error("🚫 لم يتم تعيين فصل لك.")
            return
        students = db.get_students()
        quizzes = db.get_quizzes()
        results = db.get_quiz_results()
        section_students = students[students.section_id == section_id] if not students.empty and "section_id" in students.columns else pd.DataFrame()
    else:
        st.error("🚫 هذه الصفحة متاحة للمدرسات وأمناء الخدمة فقط.")
        return
    if section_students.empty:
        st.info("لا توجد طالبات مسجلات في فصلك.")
        return
    section_student_ids = section_students["student_id"].tolist()
    if not results.empty and "student_id" in results.columns:
        class_results = results[results["student_id"].isin(section_student_ids)]
        if "status" in class_results.columns:
            class_results = class_results[class_results["status"] == "submitted"]
    else:
        class_results = pd.DataFrame()
    if not quizzes.empty and not class_results.empty:
        class_results = class_results.merge(quizzes[["quiz_id", "title"]], on="quiz_id", how="left")
        class_results = class_results.rename(columns={"title": "اسم المسابقة"})
    else:
        class_results["اسم المسابقة"] = ""
    if not section_students.empty and not class_results.empty:
        class_results = class_results.merge(section_students[["student_id", "full_name"]], on="student_id", how="left")
        class_results = class_results.rename(columns={"full_name": "اسم الطالبة"})
    else:
        class_results["اسم الطالبة"] = ""
    if class_results.empty:
        st.info("لا توجد نتائج مسابقات مسجلة لطالبات فصلك بعد.")
        return
    display_cols = ["اسم المسابقة", "اسم الطالبة", "score", "total_marks", "submission_time"]
    available_cols = [c for c in display_cols if c in class_results.columns]
    display_df = class_results[available_cols].copy()
    if "score" in display_df.columns:
        display_df["score"] = pd.to_numeric(display_df["score"], errors="coerce").fillna(0)
    if "total_marks" in display_df.columns:
        display_df["total_marks"] = pd.to_numeric(display_df["total_marks"], errors="coerce").fillna(20)
    st.markdown("---")
    st.subheader("🔍 بحث وتصفية")
    search_term = st.text_input("ابحث باسم الطالبة أو المسابقة", placeholder="اكتب اسم الطالبة أو المسابقة...")
    if "اسم المسابقة" in display_df.columns:
        quiz_names = ["الكل"] + sorted(display_df["اسم المسابقة"].dropna().unique().tolist())
        filter_quiz = st.selectbox("تصفية حسب المسابقة", quiz_names)
    else:
        filter_quiz = "الكل"
    sort_by = st.selectbox("ترتيب حسب", ["التاريخ", "الدرجة (تنازلي)", "الدرجة (تصاعدي)", "اسم الطالبة"])
    filtered_df = display_df.copy()
    if search_term:
        mask = pd.Series(False, index=filtered_df.index)
        if "اسم الطالبة" in filtered_df.columns:
            mask = mask | filtered_df["اسم الطالبة"].astype(str).str.contains(search_term, na=False, case=False)
        if "اسم المسابقة" in filtered_df.columns:
            mask = mask | filtered_df["اسم المسابقة"].astype(str).str.contains(search_term, na=False, case=False)
        filtered_df = filtered_df[mask]
    if filter_quiz != "الكل" and "اسم المسابقة" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["اسم المسابقة"] == filter_quiz]
    if sort_by == "التاريخ" and "submission_time" in filtered_df.columns:
        filtered_df = filtered_df.sort_values("submission_time", ascending=False)
    elif sort_by == "الدرجة (تنازلي)" and "score" in filtered_df.columns:
        filtered_df = filtered_df.sort_values("score", ascending=False)
    elif sort_by == "الدرجة (تصاعدي)" and "score" in filtered_df.columns:
        filtered_df = filtered_df.sort_values("score", ascending=True)
    elif sort_by == "اسم الطالبة" and "اسم الطالبة" in filtered_df.columns:
        filtered_df = filtered_df.sort_values("اسم الطالبة", ascending=True)
    st.markdown("---")
    st.subheader("📋 النتائج")
    if not filtered_df.empty:
        filtered_df = filtered_df.reset_index(drop=True)
        filtered_df.index = filtered_df.index + 1
        # Rename English columns to Arabic
        display_final = filtered_df.rename(columns={
            "score": "الدرجة",
            "total_marks": "الدرجة الكلية",
            "submission_time": "وقت التسليم"
        })
        available_cols = [c for c in display_final.columns if c in ["اسم المسابقة", "اسم الطالبة", "الدرجة", "الدرجة الكلية", "وقت التسليم"]]
        st.dataframe(display_final[available_cols], use_container_width=True)
        if "score" in filtered_df.columns and "total_marks" in filtered_df.columns:
            st.markdown("---")
            st.subheader("📊 إحصائيات الفصل")
            avg_score = filtered_df["score"].mean()
            max_score = filtered_df["score"].max()
            min_score = filtered_df["score"].min()
            c1, c2, c3 = st.columns(3)
            c1.metric("متوسط الدرجات", f"{avg_score:.1f}")
            c2.metric("أعلى درجة", f"{max_score:.1f}")
            c3.metric("أقل درجة", f"{min_score:.1f}")
            if "اسم الطالبة" in filtered_df.columns:
                st.markdown("---")
                st.subheader("🏆 ترتيب الطالبات")
                ranking = filtered_df.groupby("اسم الطالبة")["score"].sum().reset_index().sort_values("score", ascending=False)
                ranking.index = range(1, len(ranking) + 1)
                st.dataframe(ranking.rename(columns={"score": "المجموع"}), use_container_width=True)
    else:
        st.info("لا توجد نتائج مطابقة للبحث.")


# =============================================================================
# Unified Admin — Competitions & Exams
# =============================================================================
def show_unified_assessments_admin(db):
    """Unified admin page: one workflow for tests and exams."""
    st.markdown(
        hero_header("المسابقات والاختبارات", "📝 إنشاء وإدارة المسابقات والامتحانات في مكان واحد"),
        unsafe_allow_html=True,
    )
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    section_id = user.get("section_id", "")
    if role not in ["System Admin", "Father Account", "Service Manager", "Teacher", "Student"]:
        st.error("🚫 غير مصرح")
        return

    quizzes = db.get_quizzes()
    sections = db.get_sections()
    stages = db.get_stages()
    students = db.get_students()

    def _type_label(a_type):
        return "امتحان" if a_type == "exam" else "اختبار"

    def _can_manage_row(row):
        created_by = str(row.get("created_by", "")).strip()
        if role == "Teacher":
            return created_by == str(user_id).strip()
        return role in ["System Admin", "Service Manager"]

    if role in ["System Admin", "Service Manager"]:
        st.subheader("➕ إنشاء جديد")
        with st.form("unified_assessment_create_form"):
            a_type_ui = st.selectbox("النوع*", ["اختبار", "امتحان"])
            a_type = "exam" if a_type_ui == "امتحان" else "quiz"
            col1, col2 = st.columns(2)
            with col1:
                title = st.text_input("العنوان*")
                description = st.text_area("الوصف")
                total_marks = st.number_input("الدرجة الكلية", min_value=1, max_value=500, value=20)
            with col2:
                duration = st.number_input("الوقت (بالدقائق)", min_value=1, max_value=240, value=30 if a_type == "exam" else 15)
                selected_section = st.selectbox(
                    "الفصل (اختياري)",
                    [""] + (sections["section_id"].tolist() if not sections.empty else []),
                    format_func=lambda x: "كل الفصول" if x == "" else (
                        sections[sections["section_id"] == x]["section_name"].values[0]
                        if not sections.empty and x in sections["section_id"].values else x
                    )
                )

            num_questions = ""
            expiry_date = ""
            stage_id = ""
            chapter_lesson = ""
            start_date = ""
            end_date = ""
            passing_score = ""
            is_published = "False"
            if a_type == "quiz":
                num_questions = st.selectbox("عدد الأسئلة", [10, 20, 30], index=1)
                expiry_date = st.date_input("تاريخ الانتهاء", get_cairo_now().date() + timedelta(days=7)).strftime("%Y-%m-%d")
            else:
                c3, c4 = st.columns(2)
                with c3:
                    stage_options = stages["stage_id"].tolist() if not stages.empty else []
                    stage_id = st.selectbox(
                        "المرحلة المستهدفة*",
                        stage_options,
                        format_func=lambda x: stages[stages.stage_id == x]["stage_name"].values[0] if not stages.empty else x
                    ) if stage_options else ""
                    chapter_lesson = st.text_input("الأصحاح أو الدرس")
                    passing_score = str(st.number_input("درجة النجاح", min_value=1, max_value=500, value=50))
                with c4:
                    sd = st.date_input("تاريخ البداية", get_cairo_now().date())
                    ed = st.date_input("تاريخ النهاية", get_cairo_now().date() + timedelta(days=7))
                    start_date = sd.strftime("%Y-%m-%d")
                    end_date = ed.strftime("%Y-%m-%d")
                    expiry_date = end_date
                    is_published = "True" if st.checkbox("منشور", value=False) else "False"

            if st.form_submit_button("إنشاء", use_container_width=True):
                if not title.strip():
                    st.error("العنوان مطلوب.")
                elif a_type == "exam" and not stage_id:
                    st.error("المرحلة المستهدفة مطلوبة للامتحان.")
                else:
                    quiz_id = str(uuid.uuid4())
                    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6)) if a_type == "quiz" else ""
                    pwd = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5)) if a_type == "quiz" else ""
                    db.add_quiz({
                        "quiz_id": quiz_id,
                        "title": title.strip(),
                        "description": description.strip(),
                        "created_by": user_id,
                        "section_id": selected_section,
                        "num_questions": str(num_questions) if num_questions != "" else "",
                        "time_limit_minutes": str(duration),
                        "total_marks": str(total_marks),
                        "expiry_date": expiry_date,
                        "quiz_code": code,
                        "password": pwd,
                        "is_active": "True",
                        "assessment_type": a_type,
                        "stage_id": stage_id,
                        "chapter_lesson": chapter_lesson.strip(),
                        "exam_date": start_date or "",
                        "start_date": start_date or "",
                        "end_date": end_date or "",
                        "duration_minutes": str(duration),
                        "passing_score": passing_score,
                        "is_published": is_published,
                        "created_at": get_cairo_now().isoformat(),
                    })
                    st.success(f"✅ تم إنشاء {_type_label(a_type)} بنجاح.")
                    st.rerun()

    st.markdown("---")
    st.subheader("📝 إدارة الأسئلة")
    if quizzes.empty:
        st.info("لا توجد اختبارات/امتحانات بعد.")
    else:
        managed = quizzes.copy()
        managed["assessment_type"] = managed["assessment_type"].apply(lambda x: "exam" if str(x).strip() == "exam" else "quiz")
        type_filter = st.selectbox("فلتر النوع", ["الكل", "اختبار", "امتحان"], key="questions_type_filter")
        if type_filter == "اختبار":
            managed = managed[managed["assessment_type"] == "quiz"]
        elif type_filter == "امتحان":
            managed = managed[managed["assessment_type"] == "exam"]
        if managed.empty:
            st.info("لا عناصر مطابقة.")
        else:
            pick_id = st.selectbox("اختر العنصر لإدارة أسئلته", managed["quiz_id"], format_func=lambda x: managed[managed.quiz_id == x]["title"].values[0])
            picked_row = managed[managed["quiz_id"] == pick_id].iloc[0].to_dict()
            picked_type = "exam" if str(picked_row.get("assessment_type", "quiz")) == "exam" else "quiz"
            questions = db.get_quiz_questions(pick_id)
            st.markdown(f"**النوع:** {_type_label(picked_type)} | **عدد الأسئلة:** {len(questions)}")
            if not questions.empty:
                st.dataframe(questions[[c for c in ["question_text", "question_type", "correct_answer", "marks"] if c in questions.columns]], use_container_width=True)
            with st.form("unified_add_question_form"):
                qtext = st.text_area("نص السؤال*")
                qtype = st.selectbox("نوع السؤال", ["اختيار من متعدد", "صح وخطأ", "أكمل", "إجابة قصيرة"])
                opts = {"option1": "", "option2": "", "option3": "", "option4": ""}
                if qtype == "اختيار من متعدد":
                    c1, c2, c3, c4 = st.columns(4)
                    opts["option1"] = c1.text_input("الخيار 1")
                    opts["option2"] = c2.text_input("الخيار 2")
                    opts["option3"] = c3.text_input("الخيار 3")
                    opts["option4"] = c4.text_input("الخيار 4")
                elif qtype == "صح وخطأ":
                    opts["option1"], opts["option2"] = "صح", "خطأ"
                correct = st.text_input("الإجابة الصحيحة*")
                marks = st.number_input("درجة السؤال", min_value=1, max_value=100, value=5 if picked_type == "exam" else 1)
                if st.form_submit_button("إضافة سؤال", use_container_width=True):
                    if not qtext.strip() or not correct.strip():
                        st.error("نص السؤال والإجابة الصحيحة مطلوبان.")
                    else:
                        db.add_question({
                            "question_id": str(uuid.uuid4()),
                            "quiz_id": pick_id,
                            "question_text": qtext.strip(),
                            "question_type": qtype,
                            "option1": opts["option1"],
                            "option2": opts["option2"],
                            "option3": opts["option3"],
                            "option4": opts["option4"],
                            "correct_answer": correct.strip(),
                            "marks": str(marks),
                        })
                        st.success("✅ تمت إضافة السؤال.")
                        st.rerun()
            if not questions.empty:
                del_q = st.selectbox("اختر سؤالًا للحذف", questions["question_id"])
                if st.button("حذف السؤال", key="unified_del_q_btn"):
                    db.delete_question(del_q)
                    st.success("✅ تم حذف السؤال.")
                    st.rerun()

    st.markdown("---")
    st.subheader("📋 القائمة الموحدة")
    if quizzes.empty:
        st.info("لا توجد عناصر بعد.")
    else:
        display = quizzes.copy()
        display["assessment_type"] = display["assessment_type"].apply(lambda x: "exam" if str(x).strip() == "exam" else "quiz")
        c1, c2 = st.columns(2)
        with c1:
            list_type = st.selectbox("فلتر النوع", ["الكل", "اختبار", "امتحان"], key="list_type_filter")
        with c2:
            search_term = st.text_input("بحث بالعنوان", key="list_search_assessment")
        if list_type == "اختبار":
            display = display[display["assessment_type"] == "quiz"]
        elif list_type == "امتحان":
            display = display[display["assessment_type"] == "exam"]
        if search_term:
            display = display[display["title"].astype(str).str.contains(search_term, na=False, case=False)]
        for _, row in display.iterrows():
            a_id = row.get("quiz_id", "")
            a_type = "exam" if str(row.get("assessment_type", "quiz")) == "exam" else "quiz"
            can_manage = _can_manage_row(row)
            sec_name = ""
            if not sections.empty and str(row.get("section_id", "")).strip():
                m = sections[sections["section_id"] == row.get("section_id")]
                if not m.empty:
                    sec_name = m.iloc[0].get("section_name", "")
            st.markdown(f"**{row.get('title', '')}** | النوع: {_type_label(a_type)} | الوقت: {row.get('duration_minutes') or row.get('time_limit_minutes') or '—'} دقيقة | الدرجة: {row.get('total_marks', '—')}")
            if sec_name:
                st.caption(f"الفصل: {sec_name}")
            act = st.columns(5)
            if can_manage:
                active = str(row.get("is_active", "True")).strip() == "True"
                if act[0].button("إغلاق" if active else "تفعيل", key=f"u_toggle_{a_id}"):
                    db.update_quiz(a_id, {"is_active": "False" if active else "True"})
                    st.rerun()
                if a_type == "exam":
                    published = str(row.get("is_published", "False")).strip() == "True"
                    if act[1].button("إلغاء النشر" if published else "نشر", key=f"u_pub_{a_id}"):
                        db.update_quiz(a_id, {"is_published": "False" if published else "True"})
                        st.rerun()
                if act[2].button("تعديل", key=f"u_edit_{a_id}"):
                    st.session_state[f"u_edit_open_{a_id}"] = not st.session_state.get(f"u_edit_open_{a_id}", False)
                if role == "System Admin" and act[3].button("حذف (النتائج تبقى)", key=f"u_del_keep_{a_id}"):
                    db.delete_quiz_keep_results(a_id)
                    st.rerun()
            else:
                st.warning("⛔ لا يمكنك تعديل عنصر أنشأه شخص آخر.")

            if st.session_state.get(f"u_edit_open_{a_id}", False) and can_manage:
                with st.form(f"u_edit_form_{a_id}"):
                    e_title = st.text_input("العنوان", value=row.get("title", ""))
                    e_desc = st.text_area("الوصف", value=row.get("description", ""))
                    e_marks = st.number_input("الدرجة الكلية", min_value=1, max_value=500, value=int(float(row.get("total_marks", "20") or 20)))
                    e_duration = st.number_input("الوقت (بالدقائق)", min_value=1, max_value=240, value=int(float((row.get("duration_minutes") or row.get("time_limit_minutes") or "15"))))
                    e_is_active = st.checkbox("نشط", value=str(row.get("is_active", "True")).strip() == "True")
                    updates = {
                        "title": e_title.strip(),
                        "description": e_desc.strip(),
                        "total_marks": str(e_marks),
                        "time_limit_minutes": str(e_duration),
                        "duration_minutes": str(e_duration),
                        "is_active": "True" if e_is_active else "False",
                    }
                    if a_type == "quiz":
                        e_num_q = st.number_input("عدد الأسئلة", min_value=1, max_value=300, value=int(float(row.get("num_questions", "20") or 20)))
                        e_expiry = st.date_input("تاريخ الانتهاء", value=pd.to_datetime(row.get("expiry_date")).date() if str(row.get("expiry_date", "")).strip() else get_cairo_now().date())
                        updates["num_questions"] = str(e_num_q)
                        updates["expiry_date"] = e_expiry.strftime("%Y-%m-%d")
                    else:
                        e_chapter = st.text_input("الأصحاح أو الدرس", value=row.get("chapter_lesson", ""))
                        e_pass = st.number_input("درجة النجاح", min_value=1, max_value=500, value=int(float(row.get("passing_score", "50") or 50)))
                        e_pub = st.checkbox("منشور", value=str(row.get("is_published", "False")).strip() == "True")
                        e_start = st.date_input("تاريخ البداية", value=pd.to_datetime(row.get("start_date")).date() if str(row.get("start_date", "")).strip() else get_cairo_now().date(), key=f"start_{a_id}")
                        e_end = st.date_input("تاريخ النهاية", value=pd.to_datetime(row.get("end_date")).date() if str(row.get("end_date", "")).strip() else get_cairo_now().date(), key=f"end_{a_id}")
                        updates.update({
                            "chapter_lesson": e_chapter.strip(),
                            "passing_score": str(e_pass),
                            "is_published": "True" if e_pub else "False",
                            "start_date": e_start.strftime("%Y-%m-%d"),
                            "end_date": e_end.strftime("%Y-%m-%d"),
                            "exam_date": e_start.strftime("%Y-%m-%d"),
                            "expiry_date": e_end.strftime("%Y-%m-%d"),
                        })
                    if st.form_submit_button("💾 حفظ"):
                        db.update_quiz(a_id, updates)
                        st.session_state.pop(f"u_edit_open_{a_id}", None)
                        st.rerun()
            st.markdown("---")

    st.subheader("📊 النتائج الموحدة")
    results = db.get_quiz_results()
    if results.empty:
        st.info("لا توجد نتائج بعد.")
        return
    if "status" in results.columns:
        results = results[results["status"] == "submitted"]
    if results.empty:
        st.info("لا توجد نتائج مسلّمة بعد.")
        return
    if role == "Teacher" and section_id and not students.empty and "student_id" in results.columns:
        section_student_ids = students[students.section_id == section_id]["student_id"].tolist()
        results = results[results.student_id.isin(section_student_ids)]
    elif role == "Service Manager" and not students.empty and "student_id" in results.columns:
        section_ids = get_sections_for_supervisor(db, user_id)
        if section_ids:
            section_student_ids = students[students.section_id.isin(section_ids)]["student_id"].tolist()
            results = results[results.student_id.isin(section_student_ids)]
    if results.empty:
        st.info("لا توجد نتائج ضمن صلاحياتك.")
        return
    q_map = quizzes[["quiz_id", "title", "assessment_type"]].copy() if not quizzes.empty else pd.DataFrame(columns=["quiz_id", "title", "assessment_type"])
    if not q_map.empty:
        results = results.merge(q_map, on="quiz_id", how="left")
    results["assessment_type"] = results["assessment_type"].apply(lambda x: "exam" if str(x).strip() == "exam" else "quiz")
    if not students.empty:
        results = results.merge(students[["student_id", "full_name", "section_id"]], on="student_id", how="left")
    if not sections.empty and "section_id" in results.columns:
        results = results.merge(sections[["section_id", "section_name"]], on="section_id", how="left")
    t_filter = st.selectbox("فلتر نتائج النوع", ["الكل", "اختبار", "امتحان"], key="results_type_filter")
    if t_filter == "اختبار":
        results = results[results["assessment_type"] == "quiz"]
    elif t_filter == "امتحان":
        results = results[results["assessment_type"] == "exam"]
    if results.empty:
        st.info("لا توجد نتائج مطابقة للتصفية.")
        return
    results["النوع"] = results["assessment_type"].apply(_type_label)
    show_cols = [c for c in ["النوع", "title", "full_name", "section_name", "score", "total_marks", "submission_time"] if c in results.columns]
    st.dataframe(results[show_cols].rename(columns={
        "title": "الاختبار",
        "full_name": "اسم الطالبة",
        "section_name": "الفصل",
        "score": "الدرجة",
        "total_marks": "الدرجة الكلية",
        "submission_time": "وقت التسليم",
    }), use_container_width=True)
    if "score" in results.columns:
        results["score"] = pd.to_numeric(results["score"], errors="coerce").fillna(0)
        if "full_name" in results.columns and st.button("🏆 ترتيب الطالبات حسب المجموع", key="unified_rank_btn"):
            ranking = results.groupby("full_name")["score"].sum().reset_index().sort_values("score", ascending=False)
            st.dataframe(ranking.rename(columns={"full_name": "اسم الطالبة", "score": "المجموع"}), use_container_width=True)


# =============================================================================
# Quizzes
# =============================================================================
def show_quizzes(db, embedded=False):
    if not embedded:
        st.markdown(hero_header("المسابقات والاختبارات", "📝 إنشاء وإدارة الاختبارات والمسابقات"), unsafe_allow_html=True)
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    section_id = user.get("section_id", "")
    quizzes = db.get_quizzes()
    if role in ["System Admin", "Service Manager"]:
        st.subheader("➕ إنشاء اختبار جديد")
        with st.form("quiz_form"):
            title = st.text_input("عنوان الاختبار*")
            num_questions = st.selectbox("عدد الأسئلة", [10, 20, 30], index=1)
            time_limit = st.number_input("الوقت (بالدقائق)", 1, 180, 15)
            expiry = st.date_input("تاريخ الانتهاء", get_cairo_now().date() + timedelta(days=7))
            if st.form_submit_button("إنشاء"):
                if not title:
                    st.error("يرجى إدخال عنوان الاختبار")
                else:
                    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
                    pwd = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
                    quiz_id = str(uuid.uuid4())
                    db.add_quiz({
                        "quiz_id": quiz_id, "title": title, "description": "",
                        "created_by": user.get("user_id", ""), "section_id": "",
                        "num_questions": num_questions, "time_limit_minutes": time_limit,
                        "total_marks": 20, "expiry_date": expiry.strftime("%Y-%m-%d"),
                        "quiz_code": code, "password": pwd, "is_active": "True"
                    })
                    st.success(f"✅ تم إنشاء الاختبار! الكود: {code}")
                    time.sleep(2)
                    st.rerun()
        st.markdown("---")
        st.subheader("📝 إدارة الأسئلة")
        if not quizzes.empty and "is_active" in quizzes.columns:
            active_quizzes = quizzes[quizzes.is_active == "True"]
            if not active_quizzes.empty:
                quiz_choice = st.selectbox("اختر اختباراً لإدارة أسئلته", active_quizzes["quiz_id"],
                                           format_func=lambda x: active_quizzes[active_quizzes.quiz_id == x]["title"].values[0])
                if quiz_choice:
                    questions = db.get_quiz_questions(quiz_choice)
                    st.markdown(f"**عدد الأسئلة:** {len(questions)}")
                    if not questions.empty:
                        display_cols = [c for c in ["question_text", "question_type", "correct_answer"] if c in questions.columns]
                        st.dataframe(questions[display_cols], use_container_width=True)
                    with st.form("add_question_form"):
                        qtext = st.text_area("نص السؤال*")
                        qtype = st.selectbox("نوع السؤال", ["اختيار من متعدد", "صح وخطأ", "أكمل", "إجابة قصيرة"])
                        opts = {}
                        if qtype == "اختيار من متعدد":
                            cols = st.columns(4)
                            opts["option1"] = cols[0].text_input("الخيار 1")
                            opts["option2"] = cols[1].text_input("الخيار 2")
                            opts["option3"] = cols[2].text_input("الخيار 3")
                            opts["option4"] = cols[3].text_input("الخيار 4")
                        elif qtype == "صح وخطأ":
                            opts["option1"] = "صح"
                            opts["option2"] = "خطأ"
                        else:
                            opts["option1"] = opts["option2"] = opts["option3"] = opts["option4"] = ""
                        correct = st.text_input("الإجابة الصحيحة*")
                        if st.form_submit_button("إضافة سؤال"):
                            if not qtext or not correct:
                                st.error("نص السؤال والإجابة الصحيحة مطلوبان")
                            else:
                                db.add_question({
                                    "question_id": str(uuid.uuid4()), "quiz_id": quiz_choice,
                                    "question_text": qtext, "question_type": qtype,
                                    "option1": opts.get("option1", ""), "option2": opts.get("option2", ""),
                                    "option3": opts.get("option3", ""), "option4": opts.get("option4", ""),
                                    "correct_answer": correct
                                })
                                st.success("✅ تمت إضافة السؤال")
                                time.sleep(1)
                                st.rerun()
                    if not questions.empty:
                        del_q = st.selectbox("اختر سؤالاً لحذفه", questions["question_id"])
                        if st.button("حذف السؤال"):
                            db.delete_question(del_q)
                            st.success("تم الحذف")
                            time.sleep(1)
                            st.rerun()
        st.markdown("---")
        st.subheader("📋 إدارة الاختبارات")
        if quizzes.empty:
            st.info("لا توجد اختبارات بعد.")
        else:
            for _, q in quizzes.iterrows():
                qid = q.get("quiz_id", "")
                title = q.get("title", "")
                active = q.get("is_active", "True") == "True"
                code = q.get("quiz_code", "")
                expiry = q.get("expiry_date", "")
                created_by = q.get("created_by", "")
                col1, col2, col3, col4 = st.columns([3, 2, 2, 2])
                col1.markdown(f"**{title}**")
                col2.markdown(f"الكود: {code}")
                col3.markdown("حالة: " + ("🟢 نشط" if active else "🔴 مغلق"))
                col4.markdown(f"ينتهي: {expiry}")
                
                # Check if teacher can modify this quiz
                can_manage_quiz = True
                if role == "Teacher" and created_by != user_id:
                    can_manage_quiz = False
                
                col_actions = st.columns(4)
                if can_manage_quiz:
                    if active:
                        if col_actions[0].button("إغلاق", key=f"deact_{qid}"):
                            db.update_quiz(qid, {"is_active": "False"})
                            st.success(f"تم إغلاق الاختبار: {title}")
                            time.sleep(1)
                            st.rerun()
                    else:
                        if col_actions[0].button("تفعيل", key=f"act_{qid}"):
                            db.update_quiz(qid, {"is_active": "True"})
                            st.success(f"تم تفعيل الاختبار: {title}")
                            time.sleep(1)
                            st.rerun()
                if role == "System Admin" and can_manage_quiz:
                    if col_actions[1].button("حذف (النتائج تبقى)", key=f"del_keep_{qid}"):
                        db.delete_quiz_keep_results(qid)
                        st.success(f"تم حذف الاختبار '{title}' مع الاحتفاظ بالنتائج.")
                        time.sleep(1)
                        st.rerun()
                
                if role == "Teacher" and not can_manage_quiz:
                    st.warning("⛔ لا يمكنك التعديل على مسابقة أنشأها شخص آخر.")
                
                st.markdown("---")
    st.markdown("### 📊 نتائج الاختبارات")
    results = db.get_quiz_results()
    students = db.get_students()
    sections_all = db.get_sections()
    if not results.empty:
        if "status" in results.columns:
            results = results[results["status"] == "submitted"]
        if role == "Teacher" and section_id and not students.empty and "student_id" in results.columns and "section_id" in students.columns:
            section_student_ids = students[students.section_id == section_id]["student_id"].tolist()
            results = results[results.student_id.isin(section_student_ids)]
        elif role == "Service Manager" and not students.empty and "student_id" in results.columns and "section_id" in students.columns:
            section_ids = get_sections_for_supervisor(db, user_id)
            if section_ids:
                section_student_ids = students[students.section_id.isin(section_ids)]["student_id"].tolist()
                results = results[results.student_id.isin(section_student_ids)]
        if not students.empty:
            if "student_id" in results.columns:
                results = results.merge(students[["student_id", "full_name", "section_id"]], on="student_id", how="left")
                results.rename(columns={"full_name": "اسم الطالبة"}, inplace=True)
        if not sections_all.empty:
            if "section_id" in results.columns:
                results = results.merge(sections_all[["section_id", "section_name"]], on="section_id", how="left")
                results.rename(columns={"section_name": "الفصل"}, inplace=True)
        else:
            results["الفصل"] = ""
        if not quizzes.empty:
            if "quiz_id" in results.columns:
                results = results.merge(quizzes[["quiz_id", "title"]], on="quiz_id", how="left")
                results.rename(columns={"title": "المسابقة"}, inplace=True)
        if "score" in results.columns:
            results["score"] = pd.to_numeric(results["score"], errors="coerce").fillna(0)
        if "quiz_id" in results.columns:
            quiz_ids = results["quiz_id"].unique().tolist()
            if quiz_ids and not quizzes.empty:
                quiz_titles = quizzes[quizzes["quiz_id"].isin(quiz_ids)][["quiz_id", "title"]].drop_duplicates()
                quiz_options = ["الكل"] + quiz_titles["quiz_id"].tolist()
                selected_quiz_filter = st.selectbox("اختر الاختبار لعرض نتائجه فقط", quiz_options,
                                                    format_func=lambda x: "الكل" if x == "الكل" else quiz_titles[quiz_titles.quiz_id == x]["title"].values[0])
                if selected_quiz_filter != "الكل":
                    results = results[results.quiz_id == selected_quiz_filter]
        if results.empty:
            st.info("لا توجد نتائج مطابقة للاختبار المحدد.")
        else:
            base_cols = ["اسم الطالبة", "الفصل", "المسابقة", "score", "total_marks"]
            if "submission_time" in results.columns:
                base_cols.append("submission_time")
            if st.session_state.user.get("role") == "System Admin":
                time_cols = []
                if "start_time" in results.columns:
                    try:
                        results["بداية الاختبار"] = pd.to_datetime(results["start_time"]).apply(
                            lambda x: format_cairo_time(x.replace(tzinfo=CAIRO_TZ)) if pd.notna(x) else ""
                        )
                        time_cols.append("بداية الاختبار")
                    except Exception:
                        pass
                if "submission_time" in results.columns:
                    try:
                        results["تسليم الاختبار"] = pd.to_datetime(results["submission_time"]).apply(
                            lambda x: format_cairo_time(x.replace(tzinfo=CAIRO_TZ)) if pd.notna(x) else ""
                        )
                        time_cols.append("تسليم الاختبار")
                    except Exception:
                        pass
                display_cols = base_cols + time_cols
            else:
                display_cols = base_cols
            display_cols = list(dict.fromkeys(display_cols))
            available = [c for c in display_cols if c in results.columns]
            st.dataframe(results[available], use_container_width=True)
            if st.button("🏆 ترتيب الطالبات حسب المجموع") and "اسم الطالبة" in results.columns and "score" in results.columns:
                top = results.groupby("اسم الطالبة")["score"].sum().reset_index().sort_values("score", ascending=False)
                st.dataframe(top, use_container_width=True)
    else:
        st.info("لا توجد نتائج بعد.")


# =============================================================================
# PHASE 2 — EXAMS PORTAL (Future Feature — NOT yet active)
# =============================================================================
# This module is RESERVED for Phase 2 implementation.
# The scaffolding (Database methods, sheet columns, menu wiring) is in place,
# but the full Exams Portal features (question bank, scheduling, essay/manual
# grading, certificates, anti-cheating, analytics) are NOT yet implemented.
# See: FUTURE_FEATURES_ROADMAP.md → "PHASE 2 — MODULE B: EXAMS PORTAL"
# =============================================================================
def show_exams_management(db, embedded=False):
    if not embedded:
        st.markdown(hero_header("إدارة الامتحانات", "📝 إنشاء وإدارة الامتحانات وأسئلتها"), unsafe_allow_html=True)
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    section_id = user.get("section_id", "")

    if role not in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
        st.error("🚫 غير مصرح")
        return

    if role not in ["System Admin", "Service Manager"]:
        st.error("ليس لديك صلاحية لإدارة الامتحانات.")
        return

    exams = db.get_exams()
    sections = db.get_sections()

    # ============ إنشاء امتحان جديد ============
    if role in ["System Admin", "Service Manager"]:
        st.subheader("➕ إنشاء امتحان جديد")
        stages = db.get_stages()
        with st.form("add_exam_form"):
            col1, col2 = st.columns(2)
            with col1:
                title = st.text_input("عنوان الامتحان*")
                chapter_lesson = st.text_input("الأصحاح أو الدرس")
                description = st.text_area("الوصف")
            with col2:
                start_date = st.date_input("تاريخ البداية", get_cairo_now().date())
                end_date = st.date_input("تاريخ النهاية", get_cairo_now().date() + timedelta(days=7))
                duration = st.number_input("المدة (بالدقائق)", min_value=5, max_value=240, value=45)
            stage_options = stages["stage_id"].tolist() if not stages.empty else []
            selected_stage = st.selectbox(
                "المرحلة المستهدفة*", stage_options,
                format_func=lambda x: "—" if not stages.empty and x not in stages["stage_id"].values else (
                    stages[stages.stage_id == x]["stage_name"].values[0] if not stages.empty else x
                )
            ) if stage_options else ""
            # الفصول داخل المرحلة المختارة
            stage_sections = sections[sections["stage_id"] == selected_stage] if not sections.empty and selected_stage else pd.DataFrame()
            sec_options = stage_sections["section_id"].tolist() if not stage_sections.empty else []
            all_sections_option = "ALL_SECTIONS"
            section_choices = [all_sections_option] + sec_options
            selected_section = st.selectbox(
                "الفصل (اختياري - اختر كل الفصول للمرحلة بأكملها)",
                section_choices,
                format_func=lambda x: "كل فصول المرحلة" if x == all_sections_option else (
                    sections[sections.section_id == x]["section_name"].values[0] if not sections.empty and x in sections["section_id"].values else x
                )
            ) if section_choices else all_sections_option
            if selected_section == all_sections_option:
                selected_section = ""
            total_marks = st.number_input("الدرجة الكلية", min_value=1, max_value=500, value=100)
            passing_score = st.number_input("درجة النجاح", min_value=1, max_value=500, value=50)
            submitted = st.form_submit_button("إنشاء الامتحان", use_container_width=True)
            if submitted:
                if not title:
                    st.error("عنوان الامتحان مطلوب")
                elif not selected_stage:
                    st.error("المرحلة المستهدفة مطلوبة")
                else:
                    exam_id = str(uuid.uuid4())
                    db.add_exam({
                        "exam_id": exam_id,
                        "title": title.strip(),
                        "description": description.strip(),
                        "created_by": user_id,
                        "stage_id": selected_stage,
                        "section_id": selected_section,
                        "chapter_lesson": chapter_lesson.strip(),
                        "exam_date": start_date.strftime("%Y-%m-%d"),
                        "start_date": start_date.strftime("%Y-%m-%d"),
                        "end_date": end_date.strftime("%Y-%m-%d"),
                        "duration_minutes": str(duration),
                        "total_marks": str(total_marks),
                        "passing_score": str(passing_score),
                        "is_active": "True",
                        "is_published": "False",
                        "created_at": get_cairo_now().isoformat()
                    })
                    db.add_log(user_id, "إنشاء امتحان", f"تم إنشاء امتحان: {title}")
                    st.success(f"✅ تم إنشاء الامتحان: {title}")
                    time.sleep(1)
                    st.rerun()

        st.markdown("---")

    # ============ إدارة أسئلة الامتحانات ============
    st.subheader("📝 إدارة أسئلة الامتحانات")
    if exams.empty:
        st.info("لا توجد امتحانات مسجلة. قم بإنشاء امتحان أولاً.")
    else:
        exam_choice = st.selectbox(
            "اختر الامتحان لإدارة أسئلته", exams["exam_id"],
            format_func=lambda x: exams[exams.exam_id == x]["title"].values[0],
            key="exam_q_choice"
        )
        if exam_choice:
            exam_row = exams[exams.exam_id == exam_choice].iloc[0].to_dict()
            questions = db.get_exam_questions(exam_choice)
            st.markdown(f"**عدد الأسئلة:** {len(questions)}")

            if not questions.empty:
                display_cols = [c for c in ["question_text", "question_type", "correct_answer", "marks"] if c in questions.columns]
                st.dataframe(questions[display_cols].rename(columns={
                    "question_text": "نص السؤال", "question_type": "النوع",
                    "correct_answer": "الإجابة الصحيحة", "marks": "الدرجة"
                }), use_container_width=True)

            # إضافة سؤال
            with st.expander("➕ إضافة سؤال جديد"):
                with st.form("add_exam_question_form"):
                    qtext = st.text_area("نص السؤال*")
                    qtype = st.selectbox("نوع السؤال", ["اختيار من متعدد", "صح وخطأ", "أكمل", "إجابة قصيرة"])
                    opts = {}
                    if qtype == "اختيار من متعدد":
                        cols = st.columns(4)
                        opts["option1"] = cols[0].text_input("الخيار 1")
                        opts["option2"] = cols[1].text_input("الخيار 2")
                        opts["option3"] = cols[2].text_input("الخيار 3")
                        opts["option4"] = cols[3].text_input("الخيار 4")
                    elif qtype == "صح وخطأ":
                        opts["option1"] = "صح"
                        opts["option2"] = "خطأ"
                    else:
                        opts["option1"] = opts["option2"] = opts["option3"] = opts["option4"] = ""
                    correct = st.text_input("الإجابة الصحيحة*")
                    marks = st.number_input("درجة السؤال", min_value=1, max_value=100, value=5)
                    if st.form_submit_button("إضافة السؤال"):
                        if not qtext or not correct:
                            st.error("نص السؤال والإجابة الصحيحة مطلوبان")
                        else:
                            db.add_exam_question({
                                "question_id": str(uuid.uuid4()),
                                "exam_id": exam_choice,
                                "question_text": qtext,
                                "question_type": qtype,
                                "option1": opts.get("option1", ""),
                                "option2": opts.get("option2", ""),
                                "option3": opts.get("option3", ""),
                                "option4": opts.get("option4", ""),
                                "correct_answer": correct,
                                "marks": str(marks)
                            })
                            db.add_log(user_id, "إضافة سؤال امتحان", f"تمت إضافة سؤال لامتحان: {exam_row.get('title', '')}")
                            st.success("✅ تمت إضافة السؤال")
                            time.sleep(1)
                            st.rerun()

            # حذف سؤال
            if not questions.empty:
                with st.expander("🗑️ حذف سؤال"):
                    del_q = st.selectbox(
                        "اختر سؤالاً لحذفه", questions["question_id"],
                        format_func=lambda x: str(questions[questions.question_id == x]["question_text"].values[0])[:60] + "..."
                        if not questions[questions.question_id == x].empty else x
                    )
                    if st.button("حذف السؤال", key="del_exam_q_btn"):
                        db.delete_exam_question(del_q)
                        db.add_log(user_id, "حذف سؤال امتحان", f"تم حذف سؤال من امتحان: {exam_row.get('title', '')}")
                        st.success("✅ تم حذف السؤال")
                        time.sleep(1)
                        st.rerun()

        st.markdown("---")

    # ============ إدارة الامتحانات ============
    st.subheader("📋 إدارة الامتحانات")

    # فلاتر
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        search_term = st.text_input("🔍 بحث", placeholder="ابحث باسم الامتحان...")
    with col_f2:
        status_filter = st.selectbox("الحالة", ["الكل", "نشط", "غير نشط"])
    with col_f3:
        pub_filter = st.selectbox("النشر", ["الكل", "منشور", "غير منشور"])

    filtered_exams = exams.copy() if not exams.empty else pd.DataFrame()
    if search_term and not filtered_exams.empty:
        mask = pd.Series(False, index=filtered_exams.index)
        for col in ["title", "subject"]:
            if col in filtered_exams.columns:
                mask |= filtered_exams[col].astype(str).str.contains(search_term, na=False, case=False)
        filtered_exams = filtered_exams[mask]
    if status_filter != "الكل" and not filtered_exams.empty and "is_active" in filtered_exams.columns:
        status_val = "True" if status_filter == "نشط" else "False"
        filtered_exams = filtered_exams[filtered_exams["is_active"].astype(str) == status_val]
    if pub_filter != "الكل" and not filtered_exams.empty and "is_published" in filtered_exams.columns:
        pub_val = "True" if pub_filter == "منشور" else "False"
        filtered_exams = filtered_exams[filtered_exams["is_published"].astype(str) == pub_val]

    if filtered_exams.empty:
        st.info("لا توجد امتحانات مطابقة للتصفية.")
    else:
        for _, ex in filtered_exams.iterrows():
            ex_id = ex.get("exam_id", "")
            ex_title = ex.get("title", "")
            ex_chapter = ex.get("chapter_lesson", "")
            ex_date = ex.get("exam_date", "")
            ex_duration = ex.get("duration_minutes", "")
            ex_marks = ex.get("total_marks", "")
            ex_active = str(ex.get("is_active", "True")).strip() == "True"
            ex_published = str(ex.get("is_published", "False")).strip() == "True"
            ex_created_by = ex.get("created_by", "")
            ex_section = ex.get("section_id", "")

            section_name = ""
            if not sections.empty and ex_section:
                sec_match = sections[sections["section_id"] == ex_section]
                if not sec_match.empty:
                    section_name = sec_match.iloc[0].get("section_name", "")

            # معلومات الامتحان
            col1, col2, col3, col4 = st.columns([3, 2, 2, 2])
            col1.markdown(f"**📝 {ex_title}**")
            col2.markdown(f"📖 {ex_chapter}" if ex_chapter else "📖 —")
            col3.markdown(f"📅 {ex_date}")
            col4.markdown(f"⏱️ {ex_duration} دقيقة")

            # شارات الحالة
            status_cols = st.columns(2)
            status_cols[0].markdown(
                "🟢 **نشط**" if ex_active else "🔴 **موقوف**"
            )
            status_cols[1].markdown(
                "📢 **منشور**" if ex_published else "📭 **غير منشور**"
            )
            if section_name:
                st.markdown(f"🏫 **الفصل:** {section_name}")
            st.markdown(f"🎯 **الدرجة الكلية:** {ex_marks}")

            # أزرار الإجراءات
            can_manage = True
            if role == "Teacher" and ex_created_by != user_id:
                can_manage = False

            if can_manage and role in ["System Admin", "Service Manager"]:
                st.markdown("##### الإجراءات")
                act_cols = st.columns(7)

                # تفعيل / إيقاف
                if ex_active:
                    if act_cols[0].button("⏸️ إيقاف", key=f"stop_exam_{ex_id}"):
                        db.update_exam(ex_id, {"is_active": "False"})
                        db.add_log(user_id, "إيقاف امتحان", f"تم إيقاف الامتحان: {ex_title}")
                        st.success(f"✅ تم إيقاف الامتحان: {ex_title}")
                        time.sleep(1)
                        st.rerun()
                else:
                    if act_cols[0].button("▶️ تفعيل", key=f"act_exam_{ex_id}"):
                        db.update_exam(ex_id, {"is_active": "True"})
                        db.add_log(user_id, "تفعيل امتحان", f"تم تفعيل الامتحان: {ex_title}")
                        st.success(f"✅ تم تفعيل الامتحان: {ex_title}")
                        time.sleep(1)
                        st.rerun()

                # نشر / إلغاء نشر
                if ex_published:
                    if act_cols[1].button("📭 إلغاء النشر", key=f"unpub_exam_{ex_id}"):
                        db.update_exam(ex_id, {"is_published": "False"})
                        db.add_log(user_id, "إلغاء نشر امتحان", f"تم إلغاء نشر الامتحان: {ex_title}")
                        st.success(f"✅ تم إلغاء نشر الامتحان: {ex_title}")
                        time.sleep(1)
                        st.rerun()
                else:
                    if act_cols[2].button("📢 نشر", key=f"pub_exam_{ex_id}"):
                        db.update_exam(ex_id, {"is_published": "True"})
                        db.add_log(user_id, "نشر امتحان", f"تم نشر الامتحان: {ex_title}")
                        st.success(f"✅ تم نشر الامتحان: {ex_title}")
                        time.sleep(1)
                        st.rerun()

                # معاينة
                if act_cols[3].button("👁️ معاينة", key=f"preview_exam_{ex_id}"):
                    st.session_state.selected_exam_preview = ex_id

                # تعديل
                if act_cols[4].button("✏️ تعديل", key=f"edit_exam_{ex_id}"):
                    st.session_state[f"edit_exam_{ex_id}"] = not st.session_state.get(f"edit_exam_{ex_id}", False)

                # حذف
                if act_cols[5].button("🗑️ حذف", key=f"del_exam_{ex_id}"):
                    st.session_state[f"confirm_del_exam_{ex_id}"] = True

                # حذف نهائي
                if act_cols[6].button("🗑️ تأكيد الحذف", key=f"confirm_del_exam_btn_{ex_id}",
                                      disabled=not st.session_state.get(f"confirm_del_exam_{ex_id}", False)):
                    db.delete_exam(ex_id)
                    db.add_log(user_id, "حذف امتحان", f"تم حذف الامتحان: {ex_title}")
                    st.success(f"✅ تم حذف الامتحان: {ex_title}")
                    st.session_state.pop(f"confirm_del_exam_{ex_id}", None)
                    time.sleep(1)
                    st.rerun()

                # ===== معاينة الامتحان =====
                if st.session_state.get("selected_exam_preview") == ex_id:
                    with st.expander(f"👁️ معاينة: {ex_title}", expanded=True):
                        st.markdown(f"**الوصف:** {ex.get('description', '') or '—'}")
                        st.markdown(f"**تاريخ الامتحان:** {ex_date}")
                        st.markdown(f"**المدة:** {ex_duration} دقيقة")
                        st.markdown(f"**الدرجة الكلية:** {ex_marks}")
                        st.markdown(f"**الحالة:** {'نشط' if ex_active else 'موقوف'} | **النشر:** {'منشور' if ex_published else 'غير منشور'}")
                        st.markdown("---")
                        preview_qs = db.get_exam_questions(ex_id)
                        if preview_qs.empty:
                            st.info("لا توجد أسئلة في هذا الامتحان بعد.")
                        else:
                            st.markdown(f"#### الأسئلة ({len(preview_qs)})")
                            for q_idx, (_, q_row) in enumerate(preview_qs.iterrows()):
                                q = q_row.to_dict()
                                st.markdown(f"**سؤال {q_idx+1}:** {q.get('question_text', '')}")
                                st.markdown(f"📝 النوع: {q.get('question_type', '')} | 🎯 الدرجة: {q.get('marks', '')}")
                                col_p1, col_p2 = st.columns(2)
                                col_p1.markdown(f"**الإجابة الصحيحة:** {q.get('correct_answer', '')}")
                                st.markdown("---")
                        if st.button("إغلاق المعاينة", key=f"close_preview_{ex_id}"):
                            st.session_state.pop("selected_exam_preview", None)
                            st.rerun()

                # ===== تعديل الامتحان =====
                if st.session_state.get(f"edit_exam_{ex_id}", False):
                    with st.expander(f"✏️ تعديل: {ex_title}", expanded=True):
                        with st.form(f"edit_exam_form_{ex_id}"):
                            edit_title = st.text_input("عنوان الامتحان", value=ex_title)
                            edit_chapter = st.text_input("الأصحاح أو الدرس", value=ex_chapter)
                            edit_desc = st.text_area("الوصف", value=ex.get("description", ""))
                            edit_date = st.date_input(
                                "تاريخ الامتحان",
                                value=pd.to_datetime(ex_date).date() if ex_date else get_cairo_now().date()
                            )
                            edit_duration = st.number_input("المدة (بالدقائق)", min_value=5, max_value=240, value=int(ex_duration or 45))
                            edit_marks = st.number_input("الدرجة الكلية", min_value=1, max_value=500, value=int(ex_marks or 100))
                            edit_section = st.selectbox(
                                "الفصل", sec_options,
                                index=sec_options.index(ex_section) if ex_section in sec_options else 0,
                                format_func=lambda x: "—" if not sections.empty and x not in sections["section_id"].values else (
                                    sections[sections.section_id == x]["section_name"].values[0] if not sections.empty else x
                                )
                            ) if sec_options else ""
                            if st.form_submit_button("💾 حفظ التعديلات"):
                                db.update_exam(ex_id, {
                                    "title": edit_title.strip(),
                                    "chapter_lesson": edit_chapter.strip(),
                                    "description": edit_desc.strip(),
                                    "exam_date": edit_date.strftime("%Y-%m-%d"),
                                    "duration_minutes": str(edit_duration),
                                    "total_marks": str(edit_marks),
                                    "section_id": edit_section
                                })
                                db.add_log(user_id, "تعديل امتحان", f"تم تعديل الامتحان: {edit_title}")
                                st.success("✅ تم حفظ التعديلات")
                                st.session_state.pop(f"edit_exam_{ex_id}", None)
                                time.sleep(1)
                                st.rerun()
            else:
                st.warning("⛔ لا يمكنك التعديل على امتحان أنشأه شخص آخر.")

            st.markdown("---")

    # ============ نتائج الامتحانات ============
    st.markdown("### 📊 نتائج الامتحانات")
    results = db.get_exam_results()
    students = db.get_students()

    if results.empty:
        st.info("لا توجد نتائج امتحانات بعد.")
    else:
        if "status" in results.columns:
            results = results[results["status"] == "submitted"]

        # تصفية حسب الدور
        if role == "Teacher" and section_id and not students.empty and "student_id" in results.columns:
            section_student_ids = students[students.section_id == section_id]["student_id"].tolist()
            results = results[results.student_id.isin(section_student_ids)]
        elif role == "Service Manager" and not students.empty and "student_id" in results.columns:
            section_ids = get_sections_for_supervisor(db, user_id)
            if section_ids:
                section_student_ids = students[students.section_id.isin(section_ids)]["student_id"].tolist()
                results = results[results.student_id.isin(section_student_ids)]

        # دمج بيانات الطالبات
        if not students.empty and "student_id" in results.columns and "full_name" in students.columns:
            results = results.merge(students[["student_id", "full_name", "section_id"]], on="student_id", how="left")
            results.rename(columns={"full_name": "اسم الطالبة"}, inplace=True)

        # دمج أسماء الامتحانات
        if not exams.empty and "exam_id" in results.columns:
            results = results.merge(exams[["exam_id", "title"]], on="exam_id", how="left")
            results.rename(columns={"title": "الامتحان"}, inplace=True)

        # دمج أسماء الفصول
        if not sections.empty and "section_id" in results.columns:
            results = results.merge(sections[["section_id", "section_name"]], on="section_id", how="left")
            results.rename(columns={"section_name": "الفصل"}, inplace=True)
            # Convert section_id to section_name for display
            if "section_id" in results.columns and "الفصل" in results.columns:
                results["section_id_display"] = results["الفصل"].fillna(results["section_id"])
                results.drop(columns=["section_id"], inplace=True)
                results.rename(columns={"section_id_display": "section_id"}, inplace=True)

        if results.empty:
            st.info("لا توجد نتائج مطابقة.")
        else:
            display_cols = []
            if "اسم الطالبة" in results.columns:
                display_cols.append("اسم الطالبة")
            if "الامتحان" in results.columns:
                display_cols.append("الامتحان")
            if "الفصل" in results.columns:
                display_cols.append("الفصل")
            if "score" in results.columns:
                display_cols.append("score")
            if "total_marks" in results.columns:
                display_cols.append("total_marks")
            if "submission_time" in results.columns:
                display_cols.append("submission_time")

            display_df = results[display_cols].copy()
            display_df = display_df.rename(columns={
                "score": "الدرجة",
                "total_marks": "الدرجة الكلية",
                "submission_time": "وقت التسليم"
            })

            st.dataframe(display_df, use_container_width=True)

            if st.button("🏆 ترتيب الطالبات حسب المجموع") and "اسم الطالبة" in results.columns and "score" in results.columns:
                top = results.groupby("اسم الطالبة")["score"].sum().reset_index().sort_values("score", ascending=False)
                st.dataframe(top.rename(columns={"score": "المجموع"}), use_container_width=True)


# =============================================================================
# Reports - Advanced Reports & Statistics Page
# =============================================================================
def _export_to_csv_bytes(df):
    """Convert DataFrame to CSV bytes for download."""
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding='utf-8-sig')
    buf.seek(0)
    return buf.getvalue()


def _export_to_excel_with_charts(report_title, df, charts_list=None):
    """
    إنشاء ملف Excel (.xlsx) يحتوي على التقرير + الرسوم البيانية كصور داخل الملف.
    charts_list: list of (plotly_figure, sheet_name) tuples
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write data sheet
        df.to_excel(writer, sheet_name='التقرير', index=False)
        workbook = writer.book
        ws = writer.sheets['التقرير']
        
        # Style header
        header_font = Font(name='Cairo', bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column widths
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(
                df[col].astype(str).map(len).max() if not df.empty else 0,
                len(str(col))
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
        
        # Add charts as images if provided
        if charts_list:
            import warnings
            for fig, sheet_name in charts_list:
                try:
                    img_bytes = pio.to_image(fig, format='png', width=1200, height=500, scale=2)
                    img_stream = io.BytesIO(img_bytes)
                    from openpyxl.drawing.image import Image as XLImage
                    img = XLImage(img_stream)
                    img.width = 800
                    img.height = 350
                    
                    # Create new sheet for chart
                    ws_chart = workbook.create_sheet(title=sheet_name)
                    ws_chart.add_image(img, 'A1')
                except Exception:
                    warnings.warn(f"Chart export skipped for {sheet_name}: Kaleido not available")
                    note_ws = workbook.create_sheet(title=sheet_name)
                    note_ws['A1'] = "الرسوم البيانية غير متاحة في بيئة التشغيل الحالية. استخدم خيار CSV أو شاهد الرسوم في التطبيق."
    
    output.seek(0)
    return output.getvalue()


def show_reports_page(db):
    """صفحة التقارير والإحصائيات المتقدمة"""
    st.markdown(hero_header("التقارير والإحصائيات", "📊 عرض التقارير والرسوم البيانية"), unsafe_allow_html=True)
    
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")
    user_section_id = user.get("section_id", "")
    
    # Load data
    attendance = db.get_attendance()
    students = db.get_students()
    sections = db.get_sections()
    stages = db.get_stages()
    events = db.get_events()
    
    # RBAC filtering
    if role == "Service Manager" and db and user_id:
        section_ids = get_sections_for_supervisor(db, user_id)
        if section_ids:
            if not students.empty and "section_id" in students.columns:
                students = students[students.section_id.isin(section_ids)]
            if not attendance.empty and "section_id" in attendance.columns:
                attendance = attendance[attendance.section_id.isin(section_ids)]
        else:
            st.info("لا توجد فصول معينة لك.")
            return
    elif role == "Teacher":
        if user_section_id:
            if not students.empty and "section_id" in students.columns:
                students = students[students.section_id == user_section_id]
            if not attendance.empty and "section_id" in attendance.columns:
                attendance = attendance[attendance.section_id == user_section_id]
    
    if attendance.empty:
        st.info("لا توجد بيانات حضور كافية لإنشاء التقارير.")
        return
    
    # Parse dates
    if "date" in attendance.columns:
        attendance["date"] = pd.to_datetime(attendance["date"], errors="coerce")
    if "student_id" in attendance.columns and not students.empty and "student_id" in students.columns:
        attendance = attendance.merge(students[["student_id", "full_name", "section_id"]], on="student_id", how="left")
    
    # =========================================================================
    # FILTERS
    # =========================================================================
    st.markdown("### 🔍 الفلاتر")
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    
    with col_f1:
        # Date range filter
        min_date = attendance["date"].min().date() if not attendance.empty and "date" in attendance.columns else get_cairo_now().date() - timedelta(days=30)
        max_date = attendance["date"].max().date() if not attendance.empty and "date" in attendance.columns else get_cairo_now().date()
        date_from = st.date_input("من تاريخ", min_date)
        date_to = st.date_input("إلى تاريخ", max_date)
    
    with col_f2:
        # Section filter
        section_options = ["الكل"]
        if not sections.empty and "section_id" in sections.columns:
            section_options += sections["section_id"].tolist()
        section_filter = st.selectbox(
            "الفصل", section_options,
            format_func=lambda x: "الكل" if x == "الكل" else (
                sections[sections.section_id == x]["section_name"].values[0] if not sections.empty and x in sections["section_id"].values else x
            )
        )
    
    with col_f3:
        # Event type filter (for event reports)
        event_type_filter = st.selectbox("نوع الحدث", ["الكل"] + EVENT_TYPES)
    
    with col_f4:
        # Report type selector
        report_type = st.selectbox(
            "نوع التقرير",
            [
                "تقرير الحضور الأسبوعي (آخر 7 أيام)",
                "تقرير الحضور الشهري (شهر محدد)",
                "تقرير الأعضاء الجدد (آخر 30 يوم)",
                "تقرير الأعضاء الغائبين (أكثر من 3 أيام)"
            ]
        )
    
    # Apply date filter
    filtered_attendance = attendance.copy()
    if "date" in filtered_attendance.columns:
        filtered_attendance = filtered_attendance[
            (filtered_attendance["date"].dt.date >= date_from) &
            (filtered_attendance["date"].dt.date <= date_to)
        ]
    
    # Apply section filter
    if section_filter != "الكل" and "section_id" in filtered_attendance.columns:
        filtered_attendance = filtered_attendance[filtered_attendance["section_id"] == section_filter]
    
    # Apply event type filter (for events)
    filtered_events = events.copy() if not events.empty else pd.DataFrame()
    if event_type_filter != "الكل" and not filtered_events.empty and "event_type" in filtered_events.columns:
        filtered_events = filtered_events[filtered_events["event_type"] == event_type_filter]
    
    # =========================================================================
    # REPORT GENERATION
    # =========================================================================
    st.markdown("---")
    st.markdown("### 📋 التقرير")
    
    report_df = pd.DataFrame()
    report_title = ""
    charts_to_export = []
    
    if report_type == "تقرير الحضور الأسبوعي (آخر 7 أيام)":
        report_title = "تقرير الحضور الأسبوعي"
        week_ago = get_cairo_now().date() - timedelta(days=7)
        weekly = filtered_attendance[filtered_attendance["date"].dt.date >= week_ago].copy()
        
        if not weekly.empty:
            # Summary by day
            daily_summary = weekly.groupby([weekly["date"].dt.date, "status"]).size().reset_index(name="العدد")
            daily_pivot = daily_summary.pivot(index="date", columns="status", values="العدد").fillna(0).reset_index()
            daily_pivot.columns.name = None
            daily_pivot["date"] = pd.to_datetime(daily_pivot["date"]).dt.strftime("%Y-%m-%d")
            daily_pivot = daily_pivot.rename(columns={"date": "التاريخ"})
            
            # Add total
            status_cols = [c for c in daily_pivot.columns if c != "التاريخ"]
            if status_cols:
                daily_pivot["الإجمالي"] = daily_pivot[status_cols].sum(axis=1)
            
            report_df = daily_pivot
            
            # Line chart for weekly attendance
            fig_line = px.line(
                weekly.groupby([weekly["date"].dt.date, "status"]).size().reset_index(name="العدد"),
                x="date", y="العدد", color="status",
                title="الحضور اليومي - آخر 7 أيام",
                color_discrete_map={"حاضر": "#28a745", "غائب": "#dc3545", "متأخر": "#ffc107"}
            )
            fig_line.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis_title="التاريخ", yaxis_title="العدد",
                font=dict(family="Cairo")
            )
            st.plotly_chart(fig_line, use_container_width=True)
            charts_to_export.append((fig_line, "الحضور اليومي"))
            
            # Pie chart for status distribution
            status_counts = weekly["status"].value_counts().reset_index()
            status_counts.columns = ["الحالة", "العدد"]
            fig_pie = px.pie(
                status_counts, names="الحالة", values="العدد",
                title="توزيع الحالات - آخر 7 أيام",
                color_discrete_map={"حاضر": "#28a745", "غائب": "#dc3545", "متأخر": "#ffc107"}
            )
            fig_pie.update_layout(font=dict(family="Cairo"))
            st.plotly_chart(fig_pie, use_container_width=True)
            charts_to_export.append((fig_pie, "توزيع الحالات"))
        else:
            st.info("لا توجد بيانات للأيام السبعة الماضية.")
    
    elif report_type == "تقرير الحضور الشهري (شهر محدد)":
        report_title = "تقرير الحضور الشهري"
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            month = st.selectbox("الشهر", range(1, 13), index=get_cairo_now().month - 1)
        with col_m2:
            year = st.number_input("السنة", value=get_cairo_now().year, min_value=2020, max_value=2100)
        
        monthly = filtered_attendance[
            (filtered_attendance["date"].dt.month == month) &
            (filtered_attendance["date"].dt.year == year)
        ].copy()
        
        if not monthly.empty:
            # Student-level summary
            student_summary = monthly.groupby(["student_id", "full_name", "status"]).size().reset_index(name="عدد الأيام")
            student_pivot = student_summary.pivot(index=["student_id", "full_name"], columns="status", values="عدد الأيام").fillna(0).reset_index()
            student_pivot.columns.name = None
            
            # Rename columns
            student_pivot = student_pivot.rename(columns={"full_name": "اسم الطالبة"})
            if "student_id" in student_pivot.columns:
                student_pivot = student_pivot.drop(columns=["student_id"])
            
            # Add total
            status_cols = [c for c in student_pivot.columns if c != "اسم الطالبة"]
            if status_cols:
                student_pivot["إجمالي الأيام"] = student_pivot[status_cols].sum(axis=1)
            
            report_df = student_pivot
            
            # Bar chart comparing sections
            if "section_id" in monthly.columns:
                section_attendance = monthly.groupby(["section_id", "status"]).size().reset_index(name="العدد")
                section_pivot = section_attendance.pivot(index="section_id", columns="status", values="العدد").fillna(0).reset_index()
                section_pivot.columns.name = None
                
                if not sections.empty and "section_id" in sections.columns:
                    section_pivot = section_pivot.merge(
                        sections[["section_id", "section_name"]], on="section_id", how="left"
                    )
                    section_pivot["section_id"] = section_pivot["section_name"].fillna(section_pivot["section_id"])
                
                fig_bar = px.bar(
                    section_pivot, x="section_id",
                    y=[c for c in ["حاضر", "غائب", "متأخر"] if c in section_pivot.columns],
                    title=f"مقارنة بين الفصول - شهر {month}/{year}",
                    barmode="group",
                    color_discrete_map={"حاضر": "#28a745", "غائب": "#dc3545", "متأخر": "#ffc107"}
                )
                fig_bar.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    xaxis_title="الفصل", yaxis_title="العدد",
                    font=dict(family="Cairo")
                )
                st.plotly_chart(fig_bar, use_container_width=True)
                charts_to_export.append((fig_bar, "مقارنة الفصول"))
            
            # Pie chart
            status_counts = monthly["status"].value_counts().reset_index()
            status_counts.columns = ["الحالة", "العدد"]
            fig_pie = px.pie(
                status_counts, names="الحالة", values="العدد",
                title=f"توزيع الحالات - شهر {month}/{year}",
                color_discrete_map={"حاضر": "#28a745", "غائب": "#dc3545", "متأخر": "#ffc107"}
            )
            fig_pie.update_layout(font=dict(family="Cairo"))
            st.plotly_chart(fig_pie, use_container_width=True)
            charts_to_export.append((fig_pie, "توزيع الحالات"))
        else:
            st.info(f"لا توجد بيانات للشهر {month}/{year}.")
    
    elif report_type == "تقرير الأعضاء الجدد (آخر 30 يوم)":
        report_title = "تقرير الأعضاء الجدد"
        thirty_days_ago = get_cairo_now().date() - timedelta(days=30)
        
        # New students
        new_students = pd.DataFrame()
        if not students.empty:
            if "created_by" in students.columns:
                # Try to find creation date from attendance first record
                student_ids = students["student_id"].tolist()
                first_attendance = attendance[attendance["student_id"].isin(student_ids)].copy()
                if not first_attendance.empty:
                    first_dates = first_attendance.groupby("student_id")["date"].min().reset_index()
                    first_dates.columns = ["student_id", "first_date"]
                    new_students = students.merge(first_dates, on="student_id", how="inner")
                    new_students = new_students[new_students["first_date"].dt.date >= thirty_days_ago]
                else:
                    # No attendance records, show all students as new
                    new_students = students.copy()
                    new_students["first_date"] = get_cairo_now()
            else:
                # No created_by column, use first attendance date
                student_ids = students["student_id"].tolist()
                first_attendance = attendance[attendance["student_id"].isin(student_ids)].copy()
                if not first_attendance.empty:
                    first_dates = first_attendance.groupby("student_id")["date"].min().reset_index()
                    first_dates.columns = ["student_id", "first_date"]
                    new_students = students.merge(first_dates, on="student_id", how="inner")
                    new_students = new_students[new_students["first_date"].dt.date >= thirty_days_ago]
        
        if not new_students.empty:
            new_students["first_date_str"] = new_students["first_date"].dt.strftime("%Y-%m-%d")
            report_df = new_students[["full_name", "phone", "section_id", "first_date_str"]].rename(
                columns={"full_name": "الاسم", "phone": "الهاتف", "section_id": "الفصل", "first_date_str": "تاريخ أول حضور"}
            )
            
            # Add section names
            if not sections.empty and "section_id" in sections.columns:
                report_df = report_df.merge(
                    sections[["section_id", "section_name"]],
                    left_on="الفصل", right_on="section_id", how="left"
                )
                report_df["الفصل"] = report_df["section_name"].fillna(report_df["الفصل"])
                report_df = report_df.drop(columns=["section_id", "section_name"], errors="ignore")
            
            st.success(f"✅ تم إضافة {len(new_students)} أعضاء جدد في آخر 30 يوم")
            
            # Bar chart for new members by section
            if not report_df.empty and "الفصل" in report_df.columns:
                section_counts = report_df["الفصل"].value_counts().reset_index()
                section_counts.columns = ["الفصل", "العدد"]
                fig_bar = px.bar(
                    section_counts, x="الفصل", y="العدد",
                    title="الأعضاء الجدد حسب الفصل",
                    color="العدد", color_continuous_scale="Viridis"
                )
                fig_bar.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(family="Cairo")
                )
                st.plotly_chart(fig_bar, use_container_width=True)
                charts_to_export.append((fig_bar, "أعضاء جدد حسب الفصل"))
        else:
            st.info("لا يوجد أعضاء جدد في آخر 30 يوم.")
    
    elif report_type == "تقرير الأعضاء الغائبين (أكثر من 3 أيام)":
        report_title = "تقرير الأعضاء الغائبين"
        month_start = get_cairo_now().replace(day=1).date()
        month_attendance = filtered_attendance[filtered_attendance["date"].dt.date >= month_start].copy()
        
        if not month_attendance.empty and "status" in month_attendance.columns:
            # Count absences per student
            absent_counts = month_attendance[month_attendance["status"] == "غائب"].groupby(
                ["student_id", "full_name"]
            ).size().reset_index(name="أيام الغياب")
            
            absent_counts = absent_counts[absent_counts["أيام الغياب"] > 3].sort_values("أيام الغياب", ascending=False)
            
            if not absent_counts.empty:
                report_df = absent_counts.rename(columns={"full_name": "اسم الطالبة", "أيام الغياب": "أيام الغياب"})
                
                # Add section info
                if not students.empty and "section_id" in students.columns:
                    report_df = report_df.merge(
                        students[["student_id", "section_id"]], on="student_id", how="left"
                    )
                    if not sections.empty and "section_id" in sections.columns:
                        report_df = report_df.merge(
                            sections[["section_id", "section_name"]], on="section_id", how="left"
                        )
                        report_df["الفصل"] = report_df["section_name"].fillna("")
                        report_df = report_df.drop(columns=["section_id", "section_name"], errors="ignore")
                    else:
                        report_df["الفصل"] = report_df.get("section_id", "")
                        report_df = report_df.drop(columns=["section_id"], errors="ignore")
                
                report_df = report_df.drop(columns=["student_id"], errors="ignore")
                
                st.warning(f"⚠️ يوجد {len(absent_counts)} طالبة غائبة أكثر من 3 أيام هذا الشهر")
                
                # Bar chart
                fig_bar = px.bar(
                    absent_counts.head(15), x="full_name", y="أيام الغياب",
                    title="أكثر الطالبات غياباً (أكثر من 3 أيام)",
                    color="أيام الغياب", color_continuous_scale="Reds"
                )
                fig_bar.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    xaxis_title="الطالبة", yaxis_title="أيام الغياب",
                    font=dict(family="Cairo")
                )
                st.plotly_chart(fig_bar, use_container_width=True)
                charts_to_export.append((fig_bar, "الطالبات الغائبات"))
            else:
                st.success("✅ لا توجد طالبات غائبات أكثر من 3 أيام هذا الشهر.")
        else:
            st.info("لا توجد بيانات كافية.")
    
    # =========================================================================
    # DISPLAY REPORT TABLE
    # =========================================================================
    if not report_df.empty:
        st.markdown("#### 📊 بيانات التقرير")
        st.dataframe(report_df, use_container_width=True)
    
    # =========================================================================
    # INTERACTIVE CHARTS SECTION
    # =========================================================================
    st.markdown("---")
    st.markdown("### 📈 الرسوم البيانية التفاعلية")
    
    tab_chart1, tab_chart2, tab_chart3 = st.tabs([
        "📈 الحضور عبر الزمن", "📊 مقارنة بين الفصول", "🥧 توزيع الحالات"
    ])
    
    with tab_chart1:
        # Line chart: attendance over time (last 30 days)
        thirty_days_ago = get_cairo_now().date() - timedelta(days=30)
        last_30 = filtered_attendance[filtered_attendance["date"].dt.date >= thirty_days_ago].copy()
        
        if not last_30.empty:
            daily = last_30.groupby([last_30["date"].dt.date, "status"]).size().reset_index(name="العدد")
            fig_line = px.line(
                daily, x="date", y="العدد", color="status",
                title="الحضور عبر الزمن - آخر 30 يوم",
                color_discrete_map={"حاضر": "#28a745", "غائب": "#dc3545", "متأخر": "#ffc107"},
                markers=True
            )
            fig_line.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis_title="التاريخ", yaxis_title="العدد",
                font=dict(family="Cairo"),
                hovermode="x unified"
            )
            st.plotly_chart(fig_line, use_container_width=True)
        else:
            st.info("لا توجد بيانات كافية لآخر 30 يوم.")
    
    with tab_chart2:
        # Bar chart: compare sections
        if not filtered_attendance.empty and "section_id" in filtered_attendance.columns:
            section_comp = filtered_attendance.groupby(["section_id", "status"]).size().reset_index(name="العدد")
            section_pivot = section_comp.pivot(index="section_id", columns="status", values="العدد").fillna(0).reset_index()
            section_pivot.columns.name = None
            
            if not sections.empty and "section_id" in sections.columns:
                section_pivot = section_pivot.merge(
                    sections[["section_id", "section_name"]], on="section_id", how="left"
                )
                section_pivot["الفصل"] = section_pivot["section_name"].fillna(section_pivot["section_id"])
            else:
                section_pivot["الفصل"] = section_pivot["section_id"]
            
            fig_bar = px.bar(
                section_pivot, x="الفصل",
                y=[c for c in ["حاضر", "غائب", "متأخر"] if c in section_pivot.columns],
                title="مقارنة الحضور بين الفصول",
                barmode="group",
                color_discrete_map={"حاضر": "#28a745", "غائب": "#dc3545", "متأخر": "#ffc107"}
            )
            fig_bar.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis_title="الفصل", yaxis_title="العدد",
                font=dict(family="Cairo"),
                legend_title="الحالة"
            )
            st.plotly_chart(fig_bar, use_container_width=True)
        else:
            st.info("لا توجد بيانات كافية للمقارنة بين الفصول.")
    
    with tab_chart3:
        # Pie chart: status distribution
        if not filtered_attendance.empty and "status" in filtered_attendance.columns:
            status_counts = filtered_attendance["status"].value_counts().reset_index()
            status_counts.columns = ["الحالة", "العدد"]
            
            fig_pie = px.pie(
                status_counts, names="الحالة", values="العدد",
                title="توزيع الحالات (حاضر / غائب / متأخر)",
                color_discrete_map={"حاضر": "#28a745", "غائب": "#dc3545", "متأخر": "#ffc107"},
                hole=0.3
            )
            fig_pie.update_traces(textposition='inside', textinfo='percent+label')
            fig_pie.update_layout(font=dict(family="Cairo"))
            st.plotly_chart(fig_pie, use_container_width=True)
        else:
            st.info("لا توجد بيانات كافية لتوزيع الحالات.")
    
    # =========================================================================
    # EXPORT BUTTONS
    # =========================================================================
    if not report_df.empty:
        st.markdown("---")
        st.markdown("### 📥 تصدير التقرير")
        
        col_exp1, col_exp2 = st.columns(2)
        
        with col_exp1:
            # CSV Export
            csv_bytes = _export_to_csv_bytes(report_df)
            st.download_button(
                label="📄 تصدير CSV",
                data=csv_bytes,
                file_name=f"{report_title}_{get_cairo_now().strftime('%Y-%m-%d')}.csv",
                mime="text/csv",
                use_container_width=True,
                key="export_csv_btn"
            )
        
        with col_exp2:
            # Excel Export
            try:
                excel_bytes = _export_to_excel_with_charts(report_title, report_df, charts_to_export)
                st.download_button(
                    label="📗 تصدير Excel مع الرسوم البيانية",
                    data=excel_bytes,
                    file_name=f"{report_title}_{get_cairo_now().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="export_excel_btn"
                )
            except Exception as e:
                st.error("تعذر تصدير Excel: " + str(e))
    
    # =========================================================================
    # EVENT REPORT (if events data available)
    # =========================================================================
    if not filtered_events.empty and event_type_filter != "الكل":
        st.markdown("---")
        st.markdown("### 📅 تقرير الفعاليات")
        st.dataframe(filtered_events[["event_name", "event_type", "event_date", "location", "status"]].rename(
            columns={
                "event_name": "اسم الفعالية", "event_type": "النوع",
                "event_date": "التاريخ", "location": "المكان", "status": "الحالة"
            }
        ), use_container_width=True)


# =============================================================================
# Events Management
# ==============================================================================
def inject_events_css():
    """Design system already applies globally — no-op for backwards compat."""
    pass


def get_event_type_label(event_type):
    labels = {"اجتماع": "اجتماع", "خدمة": "خدمة", "رحلة": "رحلة", "احتفال": "احتفال"}
    return labels.get(event_type, event_type)


def get_event_type_css(event_type):
    css_map = {"اجتماع": "meeting", "خدمة": "service", "رحلة": "trip", "احتفال": "celebration"}
    return css_map.get(event_type, "meeting")


def show_upcoming_events(db, user, role):
    """عرض الفعاليات القادمة مع خيار التسجيل (RSVP)"""
    events = db.get_events()
    if events.empty or "event_date" not in events.columns:
        st.info("لا توجد فعاليات مسجلة.")
        return pd.DataFrame()

    today_str = get_cairo_now().strftime("%Y-%m-%d")
    events["event_date_clean"] = pd.to_datetime(events["event_date"], errors="coerce").dt.date
    upcoming = events[events["event_date_clean"] >= pd.to_datetime(today_str).date()].copy()

    if upcoming.empty:
        st.info("لا توجد فعاليات قادمة.")
        return pd.DataFrame()

    event_type_filter = st.selectbox("تصفية حسب النوع", ["الكل"] + EVENT_TYPES)
    if event_type_filter != "الكل" and "event_type" in upcoming.columns:
        upcoming = upcoming[upcoming.event_type == event_type_filter]

    if upcoming.empty:
        st.info("لا توجد فعاليات مطابقة.")
        return pd.DataFrame()

    for _, ev in upcoming.iterrows():
        ev_id = ev.get("event_id", "")
        ev_name = ev.get("event_name", "بدون اسم")
        ev_type = ev.get("event_type", "")
        ev_date = ev.get("event_date", "")
        ev_time = ev.get("event_time", "")
        ev_location = ev.get("location", "")
        ev_capacity = ev.get("max_capacity", "")
        ev_desc = ev.get("description", "")

        type_label = get_event_type_label(ev_type)
        type_css = get_event_type_css(ev_type)

        st.markdown(f"""
        <div class='event-card'>
            <span class='event-badge {type_css}'>{type_label}</span>
            <h3 style='margin-top:0.5rem;'>{ev_name}</h3>
            <p>📅 {ev_date} | ⏰ {ev_time}</p>
            <p>📍 {ev_location}</p>
            <p>👥 السعة القصوى: {ev_capacity}</p>
            <p>📝 {ev_desc}</p>
        </div>
        """, unsafe_allow_html=True)

        # RSVP for non-admin users
        if role not in ["System Admin", "Father Account"]:
            rsvp_df = db.get_event_rsvps(ev_id)
            already_rsvped = False
            if not rsvp_df.empty and "student_id" in rsvp_df.columns:
                student_id = user.get("user_id", "")
                already_rsvped = not rsvp_df[rsvp_df.student_id == student_id].empty

            if already_rsvped:
                rsvp_row = rsvp_df[rsvp_df.student_id == user.get("user_id", "")]
                rsvp_status = rsvp_row.iloc[0].get("rsvp_status", "") if not rsvp_row.empty else ""
                st.success(f"✅ تم تسجيل حضورك المتوقع: {rsvp_status}")
            else:
                if st.button("📝 تسجيل حضور متوقع", key=f"rsvp_{ev_id}", use_container_width=True):
                    st.session_state[f"rsvp_event_{ev_id}"] = True
                    st.rerun()

            if st.session_state.get(f"rsvp_event_{ev_id}", False):
                with st.form(f"rsvp_form_{ev_id}"):
                    rsvp_status = st.selectbox("حالة الحضور", RSVP_STATUSES)
                    submitted = st.form_submit_button("حفظ", use_container_width=True)
                    if submitted:
                        db.add_event_rsvp({
                            "rsvp_id": str(uuid.uuid4()),
                            "event_id": ev_id,
                            "student_id": user.get("user_id", ""),
                            "student_name": user.get("full_name", ""),
                            "rsvp_status": rsvp_status,
                            "rsvp_date": get_cairo_now().strftime("%Y-%m-%d %H:%M:%S")
                        })
                        st.success("✅ تم تسجيل الحضور المتوقع")
                        st.session_state[f"rsvp_event_{ev_id}"] = False
                        time.sleep(1)
                        st.rerun()

        st.markdown("---")

    return upcoming


def add_event_form(db, user):
    """نموذج إضافة فعالية جديدة"""
    st.subheader("➕ إضافة فعالية جديدة")
    with st.form("add_event_form"):
        col1, col2 = st.columns(2)
        with col1:
            event_name = st.text_input("اسم الفعالية*")
            event_type = st.selectbox("نوع الفعالية", EVENT_TYPES)
            event_date = st.date_input("التاريخ", get_cairo_now().date())
            event_time = st.text_input("الوقت", placeholder="مثال: 05:00 م")
        with col2:
            location = st.text_input("المكان*")
            max_capacity = st.number_input("السعة القصوى", min_value=1, value=50)
        description = st.text_area("الوصف")
        submitted = st.form_submit_button("💾 حفظ الفعالية", use_container_width=True)
        if submitted:
            if not event_name or not location:
                st.error("يرجى ملء اسم الفعالية والمكان")
            else:
                event_id = str(uuid.uuid4())
                db.add_event({
                    "event_id": event_id,
                    "event_name": event_name.strip(),
                    "event_type": event_type,
                    "event_date": event_date.strftime("%Y-%m-%d"),
                    "event_time": event_time,
                    "location": location.strip(),
                    "max_capacity": str(max_capacity),
                    "description": description.strip(),
                    "created_by": user.get("user_id", ""),
                    "status": "upcoming"
                 })
                st.success("✅ تمت إضافة الفعالية بنجاح")
                time.sleep(1)
                st.rerun()


def show_event_actual_attendance(db, user):
    """تسجيل الحضور الفعلي للفعاليات المنتهية مع ملخص إحصائي"""
    st.subheader("📊 تسجيل الحضور الفعلي")
    events = db.get_events()
    if events.empty or "event_date" not in events.columns:
        st.info("لا توجد فعاليات مسجلة.")
        return

    today_str = get_cairo_now().strftime("%Y-%m-%d")
    events["event_date_clean"] = pd.to_datetime(events["event_date"], errors="coerce").dt.date
    past = events[events["event_date_clean"] < pd.to_datetime(today_str).date()].copy()

    if past.empty:
        st.info("لا توجد فعاليات منتهية لتسجيل الحضور الفعلي.")
        return

    event_options = past["event_id"].tolist()
    selected_event = st.selectbox(
        "اختر الفعالية", event_options,
        format_func=lambda x: past[past.event_id == x]["event_name"].values[0]
    )

    if not selected_event:
        return

    # Get RSVPs and students
    rsvp_df = db.get_event_rsvps(selected_event)
    existing_attendance = db.get_event_attendance(selected_event)
    students = db.get_students()

    already_recorded = not existing_attendance.empty
    if already_recorded:
        st.warning("⚠️ تم تسجيل الحضور الفعلي مسبقاً")
        st.info("سيتم إعادة حفظ الحضور الجديد محل القديم.")

    st.markdown("#### اختر الحاضرات فعلياً")

    # Build options from RSVP list or all students
    options = []
    labels = []
    if not rsvp_df.empty and not students.empty:
        merged = rsvp_df.merge(students[["student_id", "full_name"]], on="student_id", how="left")
        options = merged["student_id"].tolist()
        labels = merged.apply(
            lambda row: f"{row['full_name']} ({row['rsvp_status']})" if row['full_name'] else row['student_id'],
            axis=1
        ).tolist()

    if not options:
        st.info("لا توجد تسجيلات حضور متوقع لهذه الفعالية.")
        return

    selected = st.multiselect(
        "الطالبات الحاضرات",
        options=options,
        format_func=lambda x: labels[options.index(x)] if x in options else x
    )

    if st.button("💾 حفظ الحضور الفعلي", use_container_width=True):
        # Remove old records if any
        if already_recorded:
            for _, rec in existing_attendance.iterrows():
                db.delete_event_attendance(rec["record_id"])

        for sid in selected:
            db.add_event_attendance({
                "record_id": str(uuid.uuid4()),
                "event_id": selected_event,
                "student_id": sid,
                "status": "حاضر",
                "notes": ""
            })
        # Add absent records for RSVPed but not present
        for _, row in rsvp_df.iterrows():
            sid = row["student_id"]
            if sid not in selected:
                db.add_event_attendance({
                    "record_id": str(uuid.uuid4()),
                    "event_id": selected_event,
                    "student_id": sid,
                    "status": "غائب",
                    "notes": ""
                })
        st.success(f"✅ تم تسجيل حضور {len(selected)} طالبة")
        time.sleep(1)
        st.rerun()

    st.markdown("---")
    st.subheader("📈 ملخص إحصائي")
    event_row = past[past.event_id == selected_event].iloc[0].to_dict()
    max_cap = int(event_row.get("max_capacity", 0) or 0)
    total_rsvp = len(rsvp_df)

    # Reload attendance records for summary
    attendance_df = db.get_event_attendance(selected_event)
    total_present = len(attendance_df[attendance_df["status"] == "حاضر"]) if not attendance_df.empty else 0
    total_absent = len(attendance_df[attendance_df["status"] == "غائب"]) if not attendance_df.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("المسجلين (RSVP)", total_rsvp)
    c2.metric("الحاضرين فعلياً", total_present)
    c3.metric("الغائبين", total_absent)
    c4.metric("السعة القصوى", max_cap)

    # Show detailed table
    if not attendance_df.empty and not students.empty:
        detail = attendance_df.merge(students[["student_id", "full_name"]], on="student_id", how="left")
        st.dataframe(detail[["full_name", "status", "notes"]].rename(
            columns={"full_name": "الاسم", "status": "الحالة", "notes": "ملاحظات"}
        ), use_container_width=True)


def show_events_page(db):
    """الصفحة الرئيسية لإدارة الفعاليات"""
    inject_events_css()
    st.markdown(hero_header("إدارة الفعاليات", "📅 إنشاء وإدارة الفعاليات والأنشطة"), unsafe_allow_html=True)
    user = st.session_state.user
    role = user.get("role", "")
    user_id = user.get("user_id", "")

    if role not in ["System Admin", "Father Account", "Service Manager", "Teacher", "Student"]:
        st.error("🚫 غير مصرح")
        return

    tab1, tab2, tab3 = st.tabs(["📋 الفعاليات القادمة", "➕ إضافة فعالية", "📊 سجل الحضور الفعلي"])

    with tab1:
        show_upcoming_events(db, user, role)

    with tab2:
        if role in ["System Admin", "Service Manager", "Teacher"]:
            add_event_form(db, user)
        else:
            st.info("👁️ يمكنك فقط مشاهدة الفعاليات")

    with tab3:
        show_event_actual_attendance(db, user)


# =============================================================================
# PHASE 2 — QR CODE ATTENDANCE SYSTEM (Future Feature — NOT yet active)
# =============================================================================
# This module is RESERVED for Phase 2 implementation.
# The scaffolding (record_qr_attendance, generate_qr_image, generate_student_id_card,
# generate_a4_qr_printable_page, process_qr_scan, show_qr_scanner_page) is in place,
# but the full QR attendance system features (encrypted JWT tokens, opening/closing
# times, sound feedback, dashboard widgets, advanced reports) are NOT yet implemented.
# See: FUTURE_FEATURES_ROADMAP.md → "PHASE 2 — MODULE A: QR CODE ATTENDANCE SYSTEM"
# =============================================================================
def generate_qr_image(data: str, size: int = 250) -> Image.Image:
    """Generate PIL Image QR code from string data."""
    qr = qrcode.QRCode(version=3, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#2563eb", back_color="white").convert("RGB")
    img = img.resize((size, size))
    return img


def generate_student_id_card(student: dict, section_name: str) -> bytes:
    """
    Generate a vertical student ID card (600x900) as PNG bytes.
    Contains: header, name, code, section, QR code, footer.
    """
    width, height = 600, 900
    img = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Try to load Arabic font, fallback to default
    try:
        font_title = ImageFont.truetype("Cairo-Bold.ttf", 36)
        font_name = ImageFont.truetype("Cairo-Bold.ttf", 48)
        font_info = ImageFont.truetype("Cairo-Regular.ttf", 28)
        font_footer = ImageFont.truetype("Cairo-Bold.ttf", 24)
    except Exception:
        try:
            font_title = ImageFont.truetype("arial.ttf", 36)
            font_name = ImageFont.truetype("arial.ttf", 48)
            font_info = ImageFont.truetype("arial.ttf", 28)
            font_footer = ImageFont.truetype("arial.ttf", 24)
        except Exception:
            font_title = ImageFont.load_default()
            font_name = ImageFont.load_default()
            font_info = ImageFont.load_default()
            font_footer = ImageFont.load_default()
    
    # Header - Church name
    header_height = 120
    draw.rectangle([0, 0, width, header_height], fill="#2563eb")
    header_text = "كنيسة الشهيدة دميانة"
    bbox = draw.textbbox((0, 0), header_text, font=font_title)
    text_width = bbox[2] - bbox[0]
    draw.text(((width - text_width) // 2, 40), header_text, fill="white", font=font_title)
    
    # Student Name
    full_name = student.get("full_name", "غير معروف")
    y_pos = 160
    draw.text((width // 2, y_pos), full_name, fill="#0f172a", font=font_name, anchor="mm")
    
    # Student Code
    y_pos = 260
    student_code = student.get("student_code", "")
    code_text = f"الكود: {student_code}"
    draw.text((width // 2, y_pos), code_text, fill="#64748b", font=font_info, anchor="mm")
    
    # Section Name
    y_pos = 320
    section_text = f"الفصل: {section_name if section_name else 'غير محدد'}"
    draw.text((width // 2, y_pos), section_text, fill="#64748b", font=font_info, anchor="mm")
    
    # QR Code
    qr_data = f"SCODE:{student.get('student_code', '')}|PWD:{student.get('student_password', '')}"
    qr_img = generate_qr_image(qr_data, size=250)
    qr_y = 400
    img.paste(qr_img, ((width - 250) // 2, qr_y))
    
    # Footer
    footer_y = 820
    draw.rectangle([0, footer_y, width, height], fill="#7c3aed")
    footer_text = "نظام الحضور الذكي"
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_width = bbox[2] - bbox[0]
    draw.text(((width - text_width) // 2, footer_y + 45), footer_text, fill="white", font=font_footer)
    
    # Convert to bytes
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.getvalue()


def generate_a4_qr_printable_page(students_list: list, section_name: str) -> bytes:
    """
    Generate A4 landscape page (3508x2480px @ 300dpi) with 6 student QR cards arranged in 2x3 grid.
    Each cell: student name, section name, QR code (300x300).
    Return PNG bytes.
    """
    # A4 landscape at 300dpi: 3508x2480
    width, height = 3508, 2480
    img = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Try to load fonts
    try:
        font_name = ImageFont.truetype("Cairo-Bold.ttf", 60)
        font_section = ImageFont.truetype("Cairo-Regular.ttf", 48)
    except Exception:
        try:
            font_name = ImageFont.truetype("arial.ttf", 60)
            font_section = ImageFont.truetype("arial.ttf", 48)
        except Exception:
            font_name = ImageFont.load_default()
            font_section = ImageFont.load_default()
    
    # Calculate grid: 2 columns x 3 rows
    cols, rows = 2, 3
    cell_width = width // cols
    cell_height = height // rows
    
    for idx, student in enumerate(students_list[:6]):
        row = idx // cols
        col = idx % cols
        
        x_offset = col * cell_width
        y_offset = row * cell_height
        
        # Draw cell border
        draw.rectangle([x_offset + 20, y_offset + 20, x_offset + cell_width - 20, y_offset + cell_height - 20], 
                      outline="#e2e8f0", width=5)
        
        # Student Name
        full_name = student.get("full_name", "غير معروف")
        y_pos = y_offset + 60
        bbox = draw.textbbox((0, 0), full_name, font=font_name)
        text_width = bbox[2] - bbox[0]
        draw.text((x_offset + (cell_width - text_width) // 2, y_pos), full_name, fill="#0f172a", font=font_name)
        
        # Section Name
        y_pos = y_pos + 80
        section_text = section_name if section_name else "غير محدد"
        bbox = draw.textbbox((0, 0), section_text, font=font_section)
        text_width = bbox[2] - bbox[0]
        draw.text((x_offset + (cell_width - text_width) // 2, y_pos), section_text, fill="#64748b", font=font_section)
        
        # QR Code
        qr_data = f"SCODE:{student.get('student_code', '')}|PWD:{student.get('student_password', '')}"
        qr_img = generate_qr_image(qr_data, size=300)
        qr_x = x_offset + (cell_width - 300) // 2
        qr_y = y_pos + 100
        img.paste(qr_img, (qr_x, qr_y))
    
    # Convert to bytes
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.getvalue()


def process_qr_scan(db, scanned_raw: str, recorded_by_user_id: str):
    """
    Process QR scan data and record attendance.
    Returns dict with success status and message.
    """
    import re
    print("=" * 60)
    print("🔹 QR Received 🔹")
    print(f"[DEBUG] process_qr_scan called")
    print(f"[DEBUG] scanned_raw: {scanned_raw}")
    print(f"[DEBUG] recorded_by_user_id: {recorded_by_user_id}")
    
    pattern = r"SCODE:(.*?)\|PWD:(.*)"
    match = re.search(pattern, scanned_raw)
    
    if not match:
        print("[DEBUG] ❌ QR pattern not matched")
        return {"success": False, "message": "❌ كود QR غير صالح"}
    
    parsed_code = match.group(1).strip()
    parsed_pwd = match.group(2).strip()
    print(f"[DEBUG] parsed_code: {parsed_code}")
    print(f"[DEBUG] parsed_pwd: {parsed_pwd}")
    
    # Lookup student
    students = db.get_students()
    if students.empty:
        print("[DEBUG] ❌ No students data")
        return {"success": False, "message": "❌ لا توجد بيانات طالبات"}
    
    print(f"[DEBUG] Total students: {len(students)}")
    
    # Handle NaN in student_code and student_password
    codes_clean = students["student_code"].fillna("").astype(str).str.strip()
    pwds_clean = students["student_password"].fillna("").astype(str).str.strip()
    
    student_match = students[
        (codes_clean == parsed_code) &
        (pwds_clean == parsed_pwd)
    ]
    
    print("🔹 Student Found 🔹")
    print(f"[DEBUG] Student match found: {not student_match.empty}")
    
    if student_match.empty:
        print("[DEBUG] ❌ Student not found with code and password")
        return {"success": False, "message": "❌ كود غير صالح أو كلمة مرور خاطئة"}
    
    student = student_match.iloc[0].to_dict()
    student_id = student.get("student_id", "")
    student_name = student.get("full_name", "")
    section_id = student.get("section_id", "")
    student_status = str(student.get("status", "active")).strip().lower()
    
    print(f"[DEBUG] student_id: {student_id}")
    print(f"[DEBUG] student_name: {student_name}")
    print(f"[DEBUG] section_id: {section_id}")
    print(f"[DEBUG] student_status: {student_status}")
    
    if student_status != "active":
        print(f"[DEBUG] ❌ Student not active: {student_status}")
        return {"success": False, "message": f"⛔ الطالبة {student_name} غير نشطة", "student": student}
    
    # Get stage_id from section
    stage_id = ""
    sections = db.get_sections()
    if not sections.empty and section_id:
        sec_match = sections[sections["section_id"] == section_id]
        if not sec_match.empty:
            stage_id = sec_match.iloc[0].get("stage_id", "")
    
    print(f"[DEBUG] stage_id: {stage_id}")
    
    # Check for duplicate attendance today
    today = get_cairo_now().strftime("%Y-%m-%d")
    existing = db.get_attendance_by_date_user(today, student_id)
    if not existing.empty:
        qr_attendance = existing[existing["attendance_method"] == "QR_SCAN"]
        if not qr_attendance.empty:
            print("[DEBUG] ⚠️ Already recorded today")
            return {"success": False, "message": "⚠️ تم تسجيل الحضور مسبقاً اليوم", "student": student}
    
    print("🔹 Password Matched 🔹")
    print("🔹 Attendance Writing 🔹")
    
    # Record attendance
    print(f"[DEBUG] Calling record_qr_attendance(student_id={student_id}, student_name={student_name}, section_id={section_id}, stage_id={stage_id}, recorded_by={recorded_by_user_id})")
    record_result = db.record_qr_attendance(student_id, student_name, section_id, stage_id, recorded_by_user_id)
    success = record_result.get("success", False)
    msg = record_result.get("message", "")
    print(f"[DEBUG] record_qr_attendance result: success={success}, msg={msg}")
    
    if success:
        print("🔹 Attendance Saved 🔹")
    
    if success:
        # Get section name for message
        section_name = ""
        if not sections.empty and section_id:
            sec_match = sections[sections["section_id"] == section_id]
            if not sec_match.empty:
                section_name = sec_match.iloc[0].get("section_name", "")
        
        # Get stage name
        stage_name = ""
        stages = db.get_stages()
        if not stages.empty and stage_id:
            stg_match = stages[stages["stage_id"] == stage_id]
            if not stg_match.empty:
                stage_name = stg_match.iloc[0].get("stage_name", "")
        
        # Add audit log
        db.add_log(recorded_by_user_id, "تسجيل حضور QR", 
                   f"الطالبة: {student_name} | الفصل: {section_name}")
        
        print(f"[DEBUG] ✅ Attendance recorded successfully for {student_name}")
        print("=" * 60)
        
        return {
            "success": True, 
            "message": f"✅ تم تسجيل حضور {student_name} | الفصل: {section_name}",
            "student": student,
            "student_name": student_name,
            "section_name": section_name,
            "stage_name": stage_name,
            "time": get_cairo_now().strftime("%Y-%m-%d %I:%M:%S %p")
        }
    else:
        print(f"[DEBUG] ❌ Failed to record attendance: {msg}")
        print("=" * 60)
        return {"success": False, "message": f"❌ فشل في تسجيل الحضور: {msg}", "student": student}


# =============================================================================
# PHASE 2 — QR CODE ATTENDANCE SYSTEM (Future Feature — NOT yet active)
# =============================================================================
# This module is RESERVED for Phase 2 implementation.
# See: FUTURE_FEATURES_ROADMAP.md → "PHASE 2 — MODULE A: QR CODE ATTENDANCE SYSTEM"
# =============================================================================
def show_qr_scanner_page(db):
    """QR Code Scanner page - Under Development."""
    under_development_page(
        title="ماسح QR Code",
        subtitle="📷 نظام الحضور الذكي",
        message="نعمل حاليًا على تطوير نظام الحضور عبر QR Code المتكامل الذي سيضم:<br><br>"
                "✅ مسح QR Code لكشف هوية الطالبة<br>"
                "✅ تسجيل الحضور تلقائياً<br>"
                "✅ دعم الكاميرا الأمامية والخلفية<br>"
                "✅ سجل عمليات المسح<br>"
                "✅ تقارير الحضور التفصيلية<br><br>"
                "سيتم إطلاق الميزة قريبًا بإذن الله. شكرًا لصبركم.",
        button_label="العودة إلى لوحة التحكم",
        button_key="qr_under_dev_back"
    )


# =============================================================================
# User Card Helpers
# ==============================================================================

def render_user_card(user, sections_df=None, stages_df=None, is_selected=False, db=None):
    user_id = user.get("user_id", "")
    full_name = user.get("full_name", "غير معروف")
    role = user.get("role", "")
    section_id = user.get("section_id", "")
    phone = user.get("phone", "")
    email = user.get("email", "")
    status = get_user_status(user)
    initials = get_initials(full_name)
    role_css = get_role_css_class(role)
    status_css = get_status_css_class(status)
    section_name = ""
    if sections_df is not None and not sections_df.empty and section_id:
        sec = sections_df[sections_df.section_id == section_id]
        if not sec.empty:
            section_name = sec.iloc[0].get("section_name", "")
    stage_name = ""
    if stages_df is not None and not stages_df.empty and section_id:
        sec = sections_df[sections_df.section_id == section_id] if sections_df is not None else pd.DataFrame()
        if not sec.empty:
            stage_id = sec.iloc[0].get("stage_id", "")
            if stage_id:
                stage = stages_df[stages_df.stage_id == stage_id]
                if not stage.empty:
                    stage_name = stage.iloc[0].get("stage_name", "")
    role_label = {"System Admin": "مدير النظام", "Father Account": "أب كاهن", "Service Manager": "أمين الخدمة", "Teacher": "مدرسة", "Student": "طالبة"}.get(role, role)
    status_label = {"active": "نشط", "inactive": "غير نشط", "suspended": "موقوف", "archived": "مؤرشف"}.get(status, "نشط")
    border = "2px solid #667eea" if is_selected else "1px solid rgba(0,0,0,0.05)"
    reg_date = user.get("registration_date", "")
    if reg_date:
        try:
            reg_date = pd.to_datetime(reg_date).strftime("%Y-%m-%d")
        except:
            reg_date = ""
    last_login = user.get("last_login", "")
    if last_login:
        try:
            last_login = pd.to_datetime(last_login).strftime("%Y-%m-%d %I:%M %p")
        except:
            last_login = ""
    return f"""
    <div class="user-card" style="border: {border};" data-user-id="{user_id}">
        <div class="card-badge {status_css}">{status_label}</div>
        <div style="display:flex; align-items:center; gap:1rem; margin-top:1.5rem;">
            <div class="user-avatar">{initials}</div>
            <div style="flex:1;">
                <div style="font-weight:700; font-size:1.1rem;">{full_name}</div>
                <span class="role-badge {role_css}">{role_label}</span>
                <div style="font-size:0.8rem; color:#6c757d; margin-top:0.3rem;">📞 {phone if phone else '—'}</div>
            </div>
        </div>
        <div style="margin-top:1rem; padding:0.8rem; background:rgba(102,126,234,0.05); border-radius:10px;">
            {('<div style="font-size:0.85rem; margin-bottom:0.3rem;">📚 <strong>المرحلة:</strong> ' + stage_name + '</div>') if stage_name else ''}
            {('<div style="font-size:0.85rem; margin-bottom:0.3rem;">🏫 <strong>الفصل:</strong> ' + section_name + '</div>') if section_name else ''}
            <div style="font-size:0.8rem; color:#6c757d;">📅 التسجيل: {reg_date if reg_date else 'غير متاح'}</div>
            <div style="font-size:0.8rem; color:#6c757d;">⏰ آخر دخول: {last_login if last_login else 'غير متاح'}</div>
        </div>
        <div style="display:flex; gap:0.5rem; margin-top:0.8rem; flex-wrap:wrap; align-items:center;">
            <span style="font-size:0.75rem; background:#f8f9fa; padding:0.2rem 0.6rem; border-radius:8px;">🆔 {user_id[:12]}...</span>
            {('<span style="font-size:0.75rem; background:#f8f9fa; padding:0.2rem 0.6rem; border-radius:8px;">📧 ' + email[:25] + '</span>') if email else ''}
        </div>
    </div>
    """

def filter_users_df(df, search_term="", role_filter="الكل", status_filter="الكل", section_filter="الكل"):
    filtered = df.copy()
    if search_term:
        search_mask = pd.Series(False, index=filtered.index)
        for col in ["full_name", "username", "phone", "email"]:
            if col in filtered.columns:
                search_mask |= filtered[col].astype(str).str.contains(search_term, na=False, case=False)
        filtered = filtered[search_mask]
    if role_filter != "الكل" and "role" in filtered.columns:
        filtered = filtered[filtered["role"] == role_filter]
    if status_filter != "الكل":
        if "status" in filtered.columns:
            st_map = {"نشط": "active", "غير نشط": "inactive", "موقوف": "suspended", "مؤرشف": "archived"}
            eng_status = st_map.get(status_filter, status_filter)
            filtered = filtered[filtered["status"] == eng_status]
        else:
            eng_status = "active" if status_filter == "نشط" else "inactive"
            filtered = filtered[filtered.get("status", "active") == eng_status]
    if section_filter != "الكل" and "section_id" in filtered.columns:
        filtered = filtered[filtered["section_id"] == section_filter]
    return filtered


# =============================================================================
# Student Profile Page
# ==============================================================================
def show_student_profile(db, student_id):
    students_df = db.get_students()
    student_row = students_df[students_df.student_id == student_id]
    if student_row.empty:
        st.error("لم يتم العثور على الطالبة")
        if st.button("🔙 العودة"):
            st.session_state.profile_user_id = None
            st.rerun()
        return
    
    student = student_row.iloc[0].to_dict()
    sections = db.get_sections()
    user = st.session_state.user
    role = user.get("role", "")
    
    # Get section name
    section_name = ""
    sec_id = student.get("section_id", "")
    if not sections.empty and sec_id:
        sec_match = sections[sections["section_id"] == sec_id]
        if not sec_match.empty:
            section_name = sec_match.iloc[0].get("section_name", "")
    
    full_name = student.get("full_name", "غير معروف")
    initials = get_initials(full_name)
    status = student.get("status", "active")
    status_label = {"active": "نشطة", "inactive": "غير نشطة"}.get(status, "نشطة")
    
    st.markdown(f"""
    <div class="profile-header">
        <div style="display:flex; align-items:center; gap:2rem;">
            <div style="width:100px;height:100px;border-radius:50%;background:rgba(255,255,255,0.2); display:flex;align-items:center;justify-content:center;font-size:2.5rem;font-weight:700;">{initials}</div>
            <div>
                <h1 style="margin:0;font-size:1.8rem;">{full_name}</h1>
                <p style="margin:0.3rem 0;opacity:0.9;">طالبة</p>
                <p style="margin:0;opacity:0.8;font-size:0.85rem;">🆔 {student_id[:12]}...</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        st.markdown('<div class="profile-stat-card">', unsafe_allow_html=True)
        st.markdown(f"<h3>{status_label}</h3>", unsafe_allow_html=True)
        st.markdown("<p>📌 الحالة</p>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="profile-stat-card">', unsafe_allow_html=True)
        birthdate = student.get("birthdate", "")
        age = get_student_age(birthdate)
        st.markdown(f"<h3>{age if age else '—'}</h3>", unsafe_allow_html=True)
        st.markdown("<p>🎂 العمر</p>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    with col3:
        st.markdown('<div class="profile-stat-card">', unsafe_allow_html=True)
        st.markdown(f"<h3>{section_name or '—'}</h3>", unsafe_allow_html=True)
        st.markdown("<p>🏫 الفصل</p>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with st.expander("📋 المعلومات الشخصية", expanded=True):
        info_cols = st.columns(2)
        with info_cols[0]:
            st.markdown(f"**👤 الاسم الكامل:** {full_name}")
            st.markdown(f"**📱 الهاتف:** {student.get('phone', '—') or '—'}")
            st.markdown(f"**📱 رقم ولي الأمر:** {student.get('parent_phone', '—') or '—'}")
            st.markdown(f"**🎂 تاريخ الميلاد:** {birthdate or '—'}")
        with info_cols[1]:
            st.markdown(f"**🏫 الفصل:** {section_name or '—'}")
            st.markdown(f"**🏫 المدرسة:** {student.get('school', '—') or '—'}")
            st.markdown(f"**📍 العنوان:** {student.get('address', '—') or '—'}")
            st.markdown(f"**📌 الحالة:** {status_label}")
    
    if student.get("notes"):
        with st.expander("📝 ملاحظات"):
            st.markdown(student.get("notes", ""))
    
    st.markdown("---")
    
    # Action buttons based on role
    if role in ["System Admin", "Service Manager"]:
        act_col1, act_col2, act_col3 = st.columns(3)
        with act_col1:
            if st.button("✏️ تعديل", use_container_width=True):
                st.session_state.edit_student_id = student_id
                st.session_state.profile_user_id = None
                st.rerun()
        with act_col2:
            if status == "active":
                if st.button("⏸️ تعطيل", use_container_width=True):
                    db.update_student(student_id, {"status": "inactive"})
                    db.add_log(user.get("user_id", ""), f"تعطيل طالبة {student_id}", f"تم تعطيل {full_name}")
                    st.success("✅ تم التعطيل")
                    time.sleep(1)
                    st.rerun()
            else:
                if st.button("▶️ تفعيل", use_container_width=True):
                    db.update_student(student_id, {"status": "active"})
                    db.add_log(user.get("user_id", ""), f"تفعيل طالبة {student_id}", f"تم تفعيل {full_name}")
                    st.success("✅ تم التفعيل")
                    time.sleep(1)
                    st.rerun()
        with act_col3:
            if st.button("🗑️ حذف", use_container_width=True):
                db.delete_student(student_id)
                db.add_log(user.get("user_id", ""), f"حذف طالبة {student_id}", f"تم حذف {full_name}")
                st.success("✅ تم الحذف")
                st.session_state.profile_user_id = None
                time.sleep(1)
                st.rerun()
    elif role == "Teacher":
        st.info("👁️ وضع العرض فقط - لا يمكنك التعديل على بيانات الطالبات")
    
    if st.button("🔙 العودة", use_container_width=True):
        st.session_state.profile_user_id = None
        st.rerun()
    
    # Edit form for System Admin and Service Manager
    if role in ["System Admin", "Service Manager"] and st.session_state.get("edit_student_id") == student_id:
        st.markdown("---")
        with st.expander("✏️ تعديل بيانات الطالبة", expanded=True):
            with st.form("edit_student_profile_form"):
                edit_name = st.text_input("الاسم الكامل*", value=full_name)
                edit_phone = st.text_input("الهاتف", value=student.get("phone", ""))
                edit_parent_phone = st.text_input("رقم ولي الأمر", value=student.get("parent_phone", ""))
                bd_value = pd.to_datetime(birthdate).date() if birthdate else None
                edit_birthdate = st.date_input("تاريخ الميلاد", value=bd_value)
                edit_address = st.text_input("العنوان", value=student.get("address", ""))
                edit_school = st.text_input("المدرسة", value=student.get("school", ""))
                edit_notes = st.text_area("ملاحظات", value=student.get("notes", ""))
                
                sec_options = sections["section_id"].tolist() if not sections.empty else []
                current_sec = sec_id if sec_id in sec_options else (sec_options[0] if sec_options else "")
                edit_section = st.selectbox("الفصل", sec_options, 
                                           index=sec_options.index(current_sec) if current_sec in sec_options else 0,
                                           format_func=lambda x: sections[sections.section_id == x]["section_name"].values[0]) if sec_options else ""
                
                edit_status = st.selectbox("الحالة", ["نشطة", "غير نشطة"], index=0 if status == "active" else 1)
                
                if st.form_submit_button("💾 حفظ التعديلات"):
                    db.update_student(student_id, {
                        "full_name": edit_name,
                        "phone": edit_phone,
                        "section_id": edit_section,
                        "parent_phone": edit_parent_phone,
                        "birthdate": edit_birthdate.strftime("%Y-%m-%d") if edit_birthdate else "",
                        "address": edit_address,
                        "school": edit_school,
                        "notes": edit_notes,
                        "status": "active" if edit_status == "نشطة" else "inactive"
                    })
                    db.add_log(user.get("user_id", ""), "تعديل طالبة", f"تم تعديل {edit_name}")
                    st.session_state.edit_student_id = None
                    st.success("✅ تم التحديث")
                    time.sleep(1)
                    st.rerun()


# =============================================================================
# User Profile Page
# =============================================================================
def show_user_profile(db, user_id):
    users_df = db.get_users()
    user_row = users_df[users_df.user_id == user_id]
    if user_row.empty:
        st.error("لم يتم العثور على المستخدم")
        if st.button("🔙 العودة"):
            st.session_state.profile_user_id = None
            st.rerun()
        return
    user = user_row.iloc[0].to_dict()
    sections = db.get_sections()
    stages = db.get_stages()
    logs = db.get_logs()
    user_logs = logs[logs.user_id == user_id] if not logs.empty and "user_id" in logs.columns else pd.DataFrame()
    section_name = ""
    if not sections.empty:
        sec = sections[sections.section_id == user.get("section_id", "")]
        section_name = sec.iloc[0]["section_name"] if not sec.empty else ""
    initials = get_initials(user.get("full_name", ""))
    role = user.get("role", "")
    role_label = {"System Admin": "مدير النظام", "Father Account": "أب كاهن", "Service Manager": "أمين الخدمة", "Teacher": "مدرسة"}.get(role, role)
    status = get_user_status(user)
    status_label = {"active": "نشط", "inactive": "غير نشط", "suspended": "موقوف", "archived": "مؤرشف"}.get(status, "نشط")
    st.markdown(f"""
    <div class="profile-header">
        <div style="display:flex; align-items:center; gap:2rem;">
            <div style="width:100px;height:100px;border-radius:50%;background:rgba(255,255,255,0.2); display:flex;align-items:center;justify-content:center;font-size:2.5rem;font-weight:700;">{initials}</div>
            <div>
                <h1 style="margin:0;font-size:1.8rem;">{user.get('full_name', '')}</h1>
                <p style="margin:0.3rem 0;opacity:0.9;">{role_label}</p>
                <p style="margin:0;opacity:0.8;font-size:0.85rem;">🆔 {user.get('user_id', '')[:12]}... | 📅 تاريخ التسجيل: {user.get('registration_date', get_cairo_now().strftime('%Y-%m-%d'))[:10]}</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        st.markdown('<div class="profile-stat-card">', unsafe_allow_html=True)
        st.markdown(f"<h3>{len(user_logs)}</h3>", unsafe_allow_html=True)
        st.markdown("<p>📋 نشاطات</p>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="profile-stat-card">', unsafe_allow_html=True)
        now = get_cairo_now()
        if user.get("birthdate"):
            try:
                age = now.year - pd.to_datetime(user["birthdate"]).year
                st.markdown(f"<h3>{age}</h3>", unsafe_allow_html=True)
                st.markdown("<p>🎂 العمر</p>", unsafe_allow_html=True)
            except:
                st.markdown("<h3>—</h3>", unsafe_allow_html=True)
                st.markdown("<p>🎂 العمر</p>", unsafe_allow_html=True)
        else:
            st.markdown("<h3>—</h3>", unsafe_allow_html=True)
            st.markdown("<p>🎂 العمر</p>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    with col3:
        st.markdown('<div class="profile-stat-card">', unsafe_allow_html=True)
        st.markdown(f'<h3><span class="status-badge {status}">{status_label}</span></h3>', unsafe_allow_html=True)
        st.markdown("<p>📌 الحالة</p>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    with st.expander("📋 المعلومات الشخصية", expanded=True):
        info_cols = st.columns(2)
        with info_cols[0]:
            st.markdown(f"**👤 الاسم الكامل:** {user.get('full_name', '')}")
            st.markdown(f"**👤 اسم المستخدم:** {user.get('username', '')}")
            st.markdown(f"**📱 الهاتف:** {user.get('phone', '—')}")
            st.markdown(f"**📧 البريد:** {user.get('email', '—')}")
        with info_cols[1]:
            st.markdown(f"**🎭 الدور:** {role_label}")
            st.markdown(f"**🏫 الفصل:** {section_name or '—'}")
            st.markdown(f"**📌 الحالة:** {status_label}")
            st.markdown(f"**🆔 المعرف:** {user.get('user_id', '')}")
    with st.expander("📜 سجل النشاطات"):
        if not user_logs.empty:
            display_logs = user_logs.sort_values("timestamp", ascending=False).head(20)
            for _, log_row in display_logs.iterrows():
                ts = log_row.get("timestamp", "")
                action = log_row.get("action", "")
                details = log_row.get("details", "")
                st.markdown(f"- **{str(ts)[:19]}** — {action} {('(' + details + ')') if details else ''}")
        else:
            st.info("لا توجد نشاطات مسجلة لهذا المستخدم.")
    st.markdown("---")
    act_col1, act_col2, act_col3, act_col4 = st.columns(4)
    with act_col1:
        if st.button("✏️ تعديل", use_container_width=True):
            st.session_state.edit_user_id = user_id
            st.session_state.profile_user_id = None
            st.rerun()
    with act_col2:
        if status == "active":
            if st.button("⏸️ تعطيل", use_container_width=True):
                db.update_user(user_id, {"status": "inactive"})
                db.add_log(st.session_state.user.get("user_id", ""), f"تعطيل مستخدم {user_id}", f"تم تعطيل {user.get('full_name', '')}")
                st.rerun()
        else:
            if st.button("▶️ تفعيل", use_container_width=True):
                db.update_user(user_id, {"status": "active"})
                db.add_log(st.session_state.user.get("user_id", ""), f"تفعيل مستخدم {user_id}", f"تم تفعيل {user.get('full_name', '')}")
                st.rerun()
    with act_col3:
        if st.button("🗑️ حذف", use_container_width=True):
            if user_id == st.session_state.user.get("user_id"):
                st.error("لا يمكنك حذف حسابك الحالي!")
            else:
                db.delete_user(user_id)
                db.add_log(st.session_state.user.get("user_id", ""), f"حذف مستخدم {user_id}", f"تم حذف {user.get('full_name', '')}")
                st.success("✅ تم حذف المستخدم")
                st.session_state.profile_user_id = None
                time.sleep(1)
                st.rerun()
    with act_col4:
        if st.button("🔙 العودة للقائمة", use_container_width=True):
            st.session_state.profile_user_id = None
            st.rerun()


# =============================================================================
# Audit Log Page - سجل التدقيق
# =============================================================================
def show_logs(db):
    st.markdown(hero_header("سجل العمليات", "📜 عرض سجل العمليات والتدقيق"), unsafe_allow_html=True)
    logs = db.get_audit_log()
    if not logs.empty:
        if "timestamp" in logs.columns:
            logs["timestamp"] = pd.to_datetime(logs["timestamp"])
        
        # إضافة عمليات البحث والتصفية
        st.markdown("#### 🔍 تصفية السجلات")
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            # فلتر حسب المستخدم
            user_ids = ["الكل"] + sorted(logs["user_id"].dropna().unique().tolist()) if "user_id" in logs.columns else ["الكل"]
            if user_ids:
                selected_user = st.selectbox("المستخدم", user_ids, format_func=lambda x: "الكل" if x == "الكل" else str(x)[:30])
        with col_f2:
            # فلتر حسب نوع العملية
            actions = ["الكل"] + sorted(logs["action"].dropna().unique().tolist()) if "action" in logs.columns else ["الكل"]
            selected_action = st.selectbox("نوع العملية", actions)
        with col_f3:
            # فلتر حسب التاريخ
            date_options = ["الكل", "آخر 24 ساعة", "آخر 7 أيام", "آخر 30 يوم"]
            selected_date_range = st.selectbox("الفترة الزمنية", date_options)
        
        filtered_logs = logs.copy()
        
        # تطبيق فلتر المستخدم
        if selected_user != "الكل" and "user_id" in filtered_logs.columns:
            filtered_logs = filtered_logs[filtered_logs["user_id"] == selected_user]
        
        # تطبيق فلتر العملية
        if selected_action != "الكل" and "action" in filtered_logs.columns:
            filtered_logs = filtered_logs[filtered_logs["action"] == selected_action]
        
        # تطبيق فلتر التاريخ
        if selected_date_range != "الكل":
            now = get_cairo_now()
            if selected_date_range == "آخر 24 ساعة":
                cutoff = now - timedelta(hours=24)
                filtered_logs = filtered_logs[filtered_logs["timestamp"] >= cutoff]
            elif selected_date_range == "آخر 7 أيام":
                cutoff = now - timedelta(days=7)
                filtered_logs = filtered_logs[filtered_logs["timestamp"] >= cutoff]
            elif selected_date_range == "آخر 30 يوم":
                cutoff = now - timedelta(days=30)
                filtered_logs = filtered_logs[filtered_logs["timestamp"] >= cutoff]
        
        # أعمدة العرض
        display_columns = ["timestamp", "username", "user_id", "action", "details", 
                          "ip_address", "country", "city", "browser", "os", "device_type"]
        available = [c for c in display_columns if c in filtered_logs.columns]
        
        st.markdown(f"**عدد السجلات:** {len(filtered_logs)}")
        st.dataframe(
            filtered_logs[available].sort_values("timestamp", ascending=False),
            use_container_width=True,
            column_config={
                "timestamp": "الوقت",
                "username": "اسم المستخدم",
                "user_id": "المعرف",
                "action": "العملية",
                "details": "التفاصيل",
                "ip_address": "IP",
                "country": "الدولة",
                "city": "المدينة",
                "browser": "المتصفح",
                "os": "نظام التشغيل",
                "device_type": "الجهاز"
            }
        )
        
        # حذف السجلات
        if st.session_state.user.get("role") == "System Admin" and "log_id" in filtered_logs.columns:
            st.markdown("---")
            st.subheader("🗑️ حذف سجل")
            del_id = st.selectbox("اختر سجلاً لحذفه", filtered_logs["log_id"], key="del_log_sel")
            if st.button("حذف السجل"):
                db.delete_audit_log(del_id)
                st.success("تم الحذف")
                time.sleep(1)
                st.rerun()
    else:
        st.info("لا توجد سجلات تدقيق بعد.")


# =============================================================================
# Change Password
# =============================================================================
def change_password(db):
    st.markdown(hero_header("تغيير كلمة المرور", "🔒 تحديث كلمة المرور الخاصة بك"), unsafe_allow_html=True)
    with st.form("change_password_form"):
        old = st.text_input("كلمة المرور الحالية", type="password", placeholder="").strip()
        new = st.text_input("كلمة المرور الجديدة", type="password", placeholder="").strip()
        confirm = st.text_input("تأكيد كلمة المرور الجديدة", type="password", placeholder="").strip()
        if st.form_submit_button("تغيير كلمة المرور"):
            if not old or not new or not confirm:
                st.error("الرجاء ملء جميع الحقول")
            elif not verify_password(old, st.session_state.user.get("password", "")):
                st.error("كلمة المرور الحالية غير صحيحة")
            elif len(new) < 4:
                st.error("كلمة المرور الجديدة يجب أن تكون 4 أحرف على الأقل")
            elif new != confirm:
                st.error("كلمتا المرور غير متطابقتين")
            else:
                hashed = hash_password(new)
                db.update_user(st.session_state.user["user_id"], {"password": hashed})
                st.session_state.user["password"] = hashed
                db.add_log(st.session_state.user["user_id"], "تغيير كلمة المرور", "تم تغيير كلمة المرور بنجاح")
                st.success("✅ تم تغيير كلمة المرور بنجاح!")


# =============================================================================
# PHASE 2 — EXAMS PORTAL (Future Feature — NOT yet active)
# =============================================================================
# This module is RESERVED for Phase 2 implementation.
# See: FUTURE_FEATURES_ROADMAP.md → "PHASE 2 — MODULE B: EXAMS PORTAL"
# =============================================================================
def show_student_exam_portal(db):
    """Legacy redirect — exam portal merged into unified student assessments page."""
    st.info("تم دمج بوابة الامتحانات مع صفحة المسابقات والاختبارات الموحدة.")
    if st.session_state.get("student_logged_in"):
        st.session_state.student_dashboard_page = STUDENT_ASSESSMENTS_PAGE
        st.rerun()


def show_exam_interface(db):
    """Legacy redirect — use unified assessment taking interface."""
    show_unified_assessment_taking_interface(db)


# =============================================================================
# PHASE 2 — EXAMS PORTAL (Future Feature — NOT yet active)
# =============================================================================
# This module is RESERVED for Phase 2 implementation.
# See: FUTURE_FEATURES_ROADMAP.md → "PHASE 2 — MODULE B: EXAMS PORTAL"
# =============================================================================
def _submit_exam(db, exam_id, student_id, student_name, answers, attempt_id=None, start_time=None, end_time=None):
    """
    تسليم الامتحان: تصحيح، حساب الدرجة، نسبة النجاح، الصحيح والخطأ، الوقت، ترتيب الفصل، وحفظ النتيجة.
    returns: dict contains score, total_marks, correct_count, wrong_count, percentage,
             pass_rate, grade, time_taken, class_rank, total_class, saved
    """
    # ===== تصحيح الامتحان =====
    score, total_marks, correct_count, wrong_count = db.grade_exam_attempt(exam_id, answers)

    # ===== حساب الدرجة والنسبة المئوية =====
    percentage = (score / total_marks * 100) if total_marks > 0 else 0

    # ===== نسبة النجاح (النجاح = 50% فأكثر) =====
    passed = percentage >= 50
    pass_rate = 100 if passed else 0

    # ===== التقدير =====
    if percentage >= 90:
        grade = "ممتاز"
    elif percentage >= 80:
        grade = "جيد جداً"
    elif percentage >= 70:
        grade = "جيد"
    elif percentage >= 60:
        grade = "مقبول"
    else:
        grade = "راسب"

    # ===== حساب الوقت المستغرق =====
    time_taken = ""
    if start_time and end_time:
        try:
            delta = end_time - start_time
            total_sec = int(delta.total_seconds())
            if total_sec < 0:
                total_sec = 0
            mins, secs = divmod(total_sec, 60)
            hours, mins = divmod(mins, 60)
            if hours > 0:
                time_taken = f"{hours} ساعة و {mins} دقيقة و {secs} ثانية"
            elif mins > 0:
                time_taken = f"{mins} دقيقة و {secs} ثانية"
            else:
                time_taken = f"{secs} ثانية"
        except Exception:
            time_taken = ""

    # ===== حفظ النتيجة =====
    saved = False
    if attempt_id:
        answers_json = json.dumps(answers, ensure_ascii=False)
        db.submit_exam_attempt(attempt_id, score, answers_json)
        saved = True

    # ===== ترتيب الفصل =====
    class_rank = None
    total_class = None
    try:
        results_df = db.get_exam_results(exam_id)
        students_df = db.get_students()
        if not results_df.empty and "student_id" in results_df.columns and not students_df.empty:
            submitted = results_df[results_df["status"] == "submitted"]
            if not submitted.empty:
                submitted = submitted.copy()
                submitted["score"] = pd.to_numeric(submitted["score"], errors="coerce").fillna(0)
                # تحديد فصل الطالبة
                student_section = ""
                match = students_df[students_df["student_id"] == student_id]
                if not match.empty:
                    student_section = match.iloc[0].get("section_id", "")
                # قائمة طالبات نفس الفصل
                if student_section:
                    section_students = students_df[students_df["section_id"] == student_section]["student_id"].tolist()
                    class_results = submitted[submitted["student_id"].isin(section_students)].copy()
                else:
                    class_results = submitted.copy()
                # تضمين نتيجة الطالبة الحالية إذا لم تكن موجودة
                if class_results.empty or student_id not in class_results["student_id"].tolist():
                    class_results = pd.concat([
                        class_results,
                        pd.DataFrame([{"student_id": student_id, "score": score}])
                    ], ignore_index=True)
                class_results = class_results.sort_values("score", ascending=False).reset_index(drop=True)
                total_class = len(class_results)
                ranks = {sid: i + 1 for i, sid in enumerate(class_results["student_id"])}
                class_rank = ranks.get(student_id, total_class)
    except Exception:
        class_rank = None
        total_class = None

    return {
        "score": score,
        "total_marks": total_marks,
        "correct_count": correct_count,
        "wrong_count": wrong_count,
        "percentage": round(percentage, 1),
        "pass_rate": pass_rate,
        "grade": grade,
        "passed": passed,
        "time_taken": time_taken,
        "class_rank": class_rank,
        "total_class": total_class,
        "saved": saved,
    }


def show_exam_results(db, result):
    """
    عرض نتائج الامتحان بعد التسليم:
    - الدرجة والنسبة المئوية
    - الصحيح والخطأ
    - نسبة النجاح
    - الوقت المستغرق
    - ترتيب الفصل
    """
    score = result.get("score", 0)
    total_marks = result.get("total_marks", 0)
    correct = result.get("correct_count", 0)
    wrong = result.get("wrong_count", 0)
    percentage = result.get("percentage", 0)
    grade = result.get("grade", "")
    passed = result.get("passed", False)
    pass_rate = result.get("pass_rate", 0)
    time_taken = result.get("time_taken", "")
    class_rank = result.get("class_rank", None)
    total_class = result.get("total_class", None)

    st.markdown(hero_header("نتيجة الامتحان", "📊 نتيجة الامتحان بعد التصحيح"), unsafe_allow_html=True)
    st.success("✅ تم تسليم الامتحان بنجاح!")

    # ===== الدرجة والنسبة =====
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📝 الدرجة", f"{score} / {total_marks}")
    col2.metric("📊 النسبة المئوية", f"{percentage}%")
    col3.metric("✅ صحيحة", correct)
    col4.metric("❌ خاطئة", wrong)

    # ===== نسبة النجاح والتقدير والوقت =====
    col5, col6, col7 = st.columns(3)
    col5.metric("🎯 التقدير", grade)
    col6.metric("📈 نسبة النجاح", f"{pass_rate}%")
    col7.metric("⏱️ الوقت المستغرق", time_taken or "—")

    # ===== ترتيب الفصل =====
    if class_rank is not None and total_class is not None:
        st.info(f"🏆 ترتيبك في الفصل: **{class_rank}** من **{total_class}** طالبة")
    else:
        st.info("🏆 ترتيب الفصل غير متاح حالياً.")

    # ===== حالة النجاح =====
    if passed:
        st.success("🎉 مبروك! لقد نجحت في الامتحان.")
    else:
        st.error("❌ لم تنجحي في هذا الامتحان. حاولي مرة أخرى في المرة القادمة.")


# =============================================================================
# PHASE 2 — EXAMS PORTAL (Future Feature — NOT yet active)
# =============================================================================
# This module is RESERVED for Phase 2 implementation.
# See: FUTURE_FEATURES_ROADMAP.md → "PHASE 2 — MODULE B: EXAMS PORTAL"
# =============================================================================
def show_exam_dashboard(db):
    """
    لوحة تحكم شاملة للامتحانات:
    - إجمالي الامتحانات
    - عدد الطالبات
    - متوسط الدرجات
    - نسبة النجاح
    - Histogram
    - Pie Chart
    - Bar Chart
    - جدول النتائج
    """
    st.markdown(hero_header("لوحة تحكم الامتحانات", "📊 إحصائيات ورسوم بيانية شاملة"), unsafe_allow_html=True)

    # ===== جلب البيانات =====
    exams_df = db.get_exams()
    results_df = db.get_exam_results()
    students_df = db.get_students()

    # ===== الفلاتر =====
    st.markdown("### 🔍 الفلاتر")
    col_f1, col_f2 = st.columns(2)

    with col_f1:
        exam_options = ["الكل"]
        if not exams_df.empty and "exam_id" in exams_df.columns:
            exam_options += exams_df["exam_id"].tolist()
        selected_exam = st.selectbox(
            "الامتحان", exam_options,
            format_func=lambda x: "الكل" if x == "الكل" else (
                exams_df[exams_df.exam_id == x]["title"].values[0]
                if not exams_df.empty and x in exams_df["exam_id"].values
                else str(x)
            ),
            key="exam_dashboard_exam_filter"
        )

    with col_f2:
        section_options = ["الكل"]
        if not students_df.empty and "section_id" in students_df.columns:
            section_options += sorted([s for s in students_df["section_id"].dropna().unique().tolist() if str(s).strip()])
        selected_section = st.selectbox(
            "الفصل", section_options,
            format_func=lambda x: "الكل" if x == "الكل" else str(x),
            key="exam_dashboard_section_filter"
        )

    # ===== تصفية البيانات =====
    # تصفية الامتحانات
    filtered_exams = exams_df.copy() if not exams_df.empty else pd.DataFrame()
    if selected_exam != "الكل" and not filtered_exams.empty and "exam_id" in filtered_exams.columns:
        filtered_exams = filtered_exams[filtered_exams["exam_id"] == selected_exam]

    # تصفية الطالبات
    filtered_students = students_df.copy() if not students_df.empty else pd.DataFrame()
    if selected_section != "الكل" and not filtered_students.empty and "section_id" in filtered_students.columns:
        filtered_students = filtered_students[filtered_students["section_id"] == selected_section]

    # تصفية النتائج
    filtered_results = results_df.copy() if not results_df.empty else pd.DataFrame()
    if selected_exam != "الكل" and not filtered_results.empty and "exam_id" in filtered_results.columns:
        filtered_results = filtered_results[filtered_results["exam_id"] == selected_exam]
    if selected_section != "الكل" and not filtered_results.empty and "student_id" in filtered_results.columns:
        section_student_ids = filtered_students["student_id"].tolist() if not filtered_students.empty and "student_id" in filtered_students.columns else []
        if section_student_ids:
            filtered_results = filtered_results[filtered_results["student_id"].isin(section_student_ids)]

    # النتائج المسلّمة فقط
    if not filtered_results.empty and "status" in filtered_results.columns:
        submitted = filtered_results[filtered_results["status"].astype(str).str.strip().str.lower() == "submitted"].copy()
    else:
        submitted = pd.DataFrame()

    # حساب النسب المئوية
    if not submitted.empty:
        if "score" in submitted.columns:
            submitted["score"] = pd.to_numeric(submitted["score"], errors="coerce").fillna(0)
        if "total_marks" in submitted.columns:
            submitted["total_marks"] = pd.to_numeric(submitted["total_marks"], errors="coerce").fillna(0)
        if "score" in submitted.columns and "total_marks" in submitted.columns:
            submitted["percentage"] = submitted.apply(
                lambda r: (r["score"] / r["total_marks"] * 100) if r["total_marks"] > 0 else 0,
                axis=1
            )
        elif "score" in submitted.columns:
            submitted["percentage"] = submitted["score"]

    # ===== الإحصائيات العامة =====
    total_exams = len(filtered_exams) if not filtered_exams.empty else 0
    total_students = filtered_students["student_id"].nunique() if not filtered_students.empty and "student_id" in filtered_students.columns else 0
    avg_score = submitted["percentage"].mean() if not submitted.empty and "percentage" in submitted.columns else 0
    pass_count = len(submitted[submitted["percentage"] >= 50]) if not submitted.empty and "percentage" in submitted.columns else 0
    pass_rate = (pass_count / len(submitted) * 100) if len(submitted) > 0 else 0

    st.markdown("### 📊 الإحصائيات العامة")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📝 إجمالي الامتحانات", total_exams)
    col2.metric("👩‍🎓 عدد الطالبات", total_students)
    col3.metric("📈 متوسط الدرجات", f"{avg_score:.1f}%")
    col4.metric("🎯 نسبة النجاح", f"{pass_rate:.1f}%")

    st.markdown("---")

    # ===== الرسوم البيانية =====
    if submitted.empty:
        st.info("لا توجد نتائج امتحانات مسلّمة للعرض.")
    else:
        # 1) Histogram - توزيع الدرجات
        st.markdown("### 📊 توزيع الدرجات (Histogram)")
        fig_hist = px.histogram(
            submitted, x="percentage", nbins=20,
            title="توزيع نسب الطالبات في الامتحانات",
            color_discrete_sequence=["#2563eb"],
            labels={"percentage": "النسبة المئوية", "count": "عدد الطالبات"}
        )
        fig_hist.update_layout(
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            xaxis_title="النسبة المئوية", yaxis_title="عدد الطالبات",
            font=dict(family="Cairo")
        )
        st.plotly_chart(fig_hist, use_container_width=True)

        # 2) Pie Chart - نسبة النجاح والرسوب
        st.markdown("### 🥧 نسبة النجاح والرسوب (Pie Chart)")
        pie_counts = pd.DataFrame({
            "الحالة": ["ناجحة", "راسبة"],
            "العدد": [pass_count, max(len(submitted) - pass_count, 0)]
        })
        pie_counts = pie_counts[pie_counts["العدد"] > 0]
        if not pie_counts.empty:
            fig_pie = px.pie(
                pie_counts, names="الحالة", values="العدد",
                title="نسبة النجاح والرسوب",
                color_discrete_map={"ناجحة": "#28a745", "راسبة": "#dc3545"},
                hole=0.3
            )
            fig_pie.update_layout(font=dict(family="Cairo"))
            st.plotly_chart(fig_pie, use_container_width=True)

        # 3) Bar Chart - متوسط الدرجات لكل امتحان
        if not exams_df.empty and "exam_id" in exams_df.columns and "exam_id" in submitted.columns:
            st.markdown("### 📊 متوسط الدرجات لكل امتحان (Bar Chart)")
            exam_avg = submitted.groupby("exam_id")["percentage"].mean().reset_index()
            exam_avg = exam_avg.merge(exams_df[["exam_id", "title"]], on="exam_id", how="left")
            exam_avg["title"] = exam_avg["title"].fillna("بدون عنوان")
            exam_avg = exam_avg.sort_values("percentage", ascending=False)

            fig_bar = px.bar(
                exam_avg, x="title", y="percentage",
                title="متوسط الدرجات لكل امتحان",
                color="percentage",
                color_continuous_scale="Blues",
                text_auto=".1f",
                labels={"title": "الامتحان", "percentage": "متوسط النسبة المئوية (%)"}
            )
            fig_bar.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis_title="الامتحان", yaxis_title="متوسط النسبة المئوية (%)",
                font=dict(family="Cairo")
            )
            st.plotly_chart(fig_bar, use_container_width=True)

        st.markdown("---")

        # ===== جدول النتائج =====
        st.markdown("### 📋 جدول النتائج")
        display = submitted.copy()

        # دمج أسماء الامتحانات
        if not exams_df.empty and "exam_id" in exams_df.columns and "exam_id" in display.columns:
            display = display.merge(exams_df[["exam_id", "title"]], on="exam_id", how="left")
            display.rename(columns={"title": "الامتحان"}, inplace=True)

        # دمج أسماء الطالبات والفصول
        if not students_df.empty and "student_id" in students_df.columns and "student_id" in display.columns:
            display = display.merge(
                students_df[["student_id", "full_name", "section_id"]],
                on="student_id", how="left"
            )
            display["full_name"] = display["full_name"].fillna(display.get("student_name", ""))
            display.rename(columns={"full_name": "اسم الطالبة", "section_id": "الفصل"}, inplace=True)
 
        # تحديد الأعمدة المعروضة
        display_cols = []
        if "اسم الطالبة" in display.columns:
            display_cols.append("اسم الطالبة")
        if "الامتحان" in display.columns:
            display_cols.append("الامتحان")
        if "الفصل" in display.columns:
            display_cols.append("الفصل")
        if "score" in display.columns:
            display_cols.append("score")
        if "total_marks" in display.columns:
            display_cols.append("total_marks")
        if "percentage" in display.columns:
            display_cols.append("percentage")
        if "status" in display.columns:
            display_cols.append("status")
        if "submission_time" in display.columns:
            display_cols.append("submission_time")

        if display_cols:
            sort_col = "percentage" if "percentage" in display_cols else display_cols[0]
            sorted_display = display[display_cols].sort_values(sort_col, ascending=False)
            st.dataframe(
                sorted_display,
                use_container_width=True,
                column_config={
                    "اسم الطالبة": "اسم الطالبة",
                    "الامتحان": "الامتحان",
                    "الفصل": "الفصل",
                    "score": "الدرجة",
                    "total_marks": "الدرجة الكلية",
                    "percentage": "النسبة المئوية",
                    "status": "الحالة",
                    "submission_time": "وقت التسليم"
                }
            )
        else:
            st.dataframe(display, use_container_width=True)

        # ===== ملخص إضافي =====
        with st.expander("📈 إحصائيات تفصيلية"):
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("🎯 أعلى درجة", f"{submitted['percentage'].max():.1f}%" if not submitted.empty and "percentage" in submitted.columns else "—")
            c2.metric("📉 أدنى درجة", f"{submitted['percentage'].min():.1f}%" if not submitted.empty and "percentage" in submitted.columns else "—")
            c3.metric("✅ عدد الناجحات", pass_count)
            c4.metric("❌ عدد الراسبات", max(len(submitted) - pass_count, 0))


# =============================================================================
# مراجعة الواجبات
# =============================================================================
def show_homework_review(db):
    """
    مراجعة واجبات الطالبات:
    - رفع صور الواجب (تسليم الواجب بالصور)
    - مراجعة وتقييم التسليمات
    - قبول أو رفض
    - إضافة ملاحظات
    - منح درجات
    - إرسال إشعار للطالبة بنتيجة المراجعة
    """
    st.markdown(hero_header("مراجعة الواجبات", "📸 عرض ومراجعة واجبات الطالبات وتقييمها"), unsafe_allow_html=True)

    user = st.session_state.get("user", {})
    role = user.get("role", "")

    homeworks = db.get_homeworks()
    if homeworks.empty:
        st.info("لا توجد واجبات مسجلة بعد.")
        return

    # المعلم يرى واجبات فصله فقط
    if role not in ["System Admin", "Father Account", "Service Manager"]:
        teacher_section = user.get("section_id", "")
        if teacher_section and "section_id" in homeworks.columns:
            homeworks = homeworks[homeworks["section_id"] == teacher_section]
        if homeworks.empty:
            st.info("لا توجد واجبات لفصلك.")
            return

    submissions = db.get_homework_submissions()
    students = db.get_students()

    # اختيار الواجب
    hw_options = {
        str(row["homework_id"]): f"{row.get('title', '')} | {row.get('subject', '')} | {row.get('section_id', '')}"
        for _, row in homeworks.iterrows()
    }
    selected_id = st.selectbox("📌 اختر الواجب للمراجعة", list(hw_options.keys()), format_func=lambda x: hw_options.get(x, x))
    if not selected_id:
        return

    homework = homeworks[homeworks["homework_id"] == selected_id].iloc[0].to_dict()
    total_marks = homework.get("total_marks", 0)
    try:
        total_marks = float(total_marks or 0)
    except (TypeError, ValueError):
        total_marks = 0.0

    # تفاصيل الواجب
    st.markdown("### 📋 تفاصيل الواجب")
    d1, d2, d3, d4 = st.columns(4)
    d1.metric("العنوان", homework.get("title", "-"))
    d2.metric("المادة", homework.get("subject", "-"))
    d3.metric("الفصل", homework.get("section_id", "-"))
    d4.metric("الدرجة الكلية", f"{total_marks:g}")
    if homework.get("description"):
        st.caption(f"📝 {homework['description']}")

    st.markdown("---")

    # ===== رفع صور واجب لطالبة =====
    with st.expander("📤 رفع صور واجب طالبة", expanded=False):
        st.caption("أضف تسليم واجب لإحدى الطالبات (مع رفع صور) ليتم مراجعته وتقييمه.")
        if students.empty:
            st.info("لا توجد طالبات مسجلات.")
        else:
            sec_id = homework.get("section_id", "")
            class_students = students
            if sec_id and "section_id" in students.columns:
                class_students = students[students["section_id"] == sec_id]
            if class_students.empty:
                st.info("لا توجد طالبات في هذا الفصل.")
            else:
                student_options = {
                    str(row["student_id"]): f"{row.get('full_name', '')} ({row.get('section_id', '')})"
                    for _, row in class_students.iterrows()
                }
                add_c1, add_c2 = st.columns(2)
                with add_c1:
                    sel_student = st.selectbox(
                        "👩‍🎓 الطالبة", list(student_options.keys()),
                        format_func=lambda x: student_options.get(x, x),
                        key=f"add_stu_{selected_id}"
                    )
                with add_c2:
                    add_note = st.text_area("📝 ملاحظة الطالبة", height=90, key=f"add_note_{selected_id}")
                uploaded = st.file_uploader(
                    "📸 رفع صور الواجب (jpg, jpeg, png, webp, gif)",
                    type=["jpg", "jpeg", "png", "webp", "gif"],
                    accept_multiple_files=True,
                    key=f"up_{selected_id}"
                )
                if st.button("💾 تسجيل التسليم", key=f"save_{selected_id}"):
                    try:
                        stu_row = class_students[class_students["student_id"] == sel_student].iloc[0]
                        images_base64, image_names = [], []
                        if uploaded:
                            for uf in uploaded:
                                b64 = base64.b64encode(uf.getvalue()).decode("utf-8")
                                if len(b64) > 45000:
                                    st.warning(f"⚠️ الصورة {uf.name} كبيرة جداً (الحد الأقصى 45 ألف حرف) ولن تُحفظ.")
                                    continue
                                images_base64.append(b64)
                                image_names.append(uf.name)
                        sub_data = {
                            "submission_id": str(uuid.uuid4()),
                            "homework_id": selected_id,
                            "student_id": str(stu_row.get("student_id", "")),
                            "student_name": str(stu_row.get("full_name", "")),
                            "section_id": str(stu_row.get("section_id", "")),
                            "image_data": json.dumps(images_base64) if images_base64 else "",
                            "image_name": json.dumps(image_names) if image_names else "",
                            "submission_note": add_note or "",
                            "status": "pending",
                            "grade": "",
                            "feedback": "",
                            "submitted_at": get_cairo_now().isoformat(),
                            "reviewed_by": "",
                            "reviewed_at": ""
                        }
                        db.add_homework_submission(sub_data)
                        st.success("✅ تم حفظ تسليم الواجب بنجاح.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ فشل حفظ التسليم: {e}")

    st.markdown("---")

    # ===== مراجعة التسليمات =====
    hw_subs = pd.DataFrame()
    if not submissions.empty and "homework_id" in submissions.columns:
        hw_subs = submissions[submissions["homework_id"] == selected_id]
    if hw_subs.empty:
        st.markdown(empty_state("لا توجد تسليمات لهذا الواجب بعد.", "📭"), unsafe_allow_html=True)
        return

    st.markdown(f"### 📝 التسليمات ({len(hw_subs)})")
    status_map = {"pending": "⏳ بانتظار المراجعة", "accepted": "✅ مقبول", "rejected": "❌ مرفوض"}

    for _, sub_row in hw_subs.iterrows():
        sub = sub_row.to_dict()
        sub_id = str(sub.get("submission_id", ""))
        cur_status = sub.get("status", "pending")

        # وقت التسليم
        sub_time_display = "غير متاح"
        try:
            sub_dt = pd.to_datetime(sub.get("submitted_at", ""))
            if pd.isna(sub_dt):
                sub_time_display = "غير متاح"
            else:
                if sub_dt.tzinfo is None:
                    sub_dt = sub_dt.replace(tzinfo=CAIRO_TZ)
                sub_time_display = format_cairo_time(sub_dt)
        except Exception:
            sub_time_display = "غير متاح"

        with st.expander(
            f"👩‍🎓 {sub.get('student_name', '')} — {status_map.get(cur_status, cur_status)} — 🕒 {sub_time_display}",
            expanded=False
        ):
            info_cols = st.columns(4)
            info_cols[0].markdown(f"**الفصل:** {sub.get('section_id', '-')}")
            cur_grade = sub.get("grade", "")
            info_cols[1].markdown(f"**الدرجة الحالية:** {cur_grade if cur_grade not in ('', None) else '—'} من {total_marks:g}")
            info_cols[2].markdown(f"**الحالة:** {status_map.get(cur_status, cur_status)}")
            info_cols[3].markdown(f"**وقت التسليم:** {sub_time_display}")

            if sub.get("submission_note"):
                st.markdown(f"**📝 ملاحظة الطالبة:** {sub['submission_note']}")

            # عرض الصور
            img_json = sub.get("image_data", "")
            if img_json:
                try:
                    imgs = json.loads(img_json)
                    names = json.loads(sub.get("image_name", "[]") or "[]")
                    if not isinstance(imgs, list):
                        imgs = [imgs]
                    if not isinstance(names, list):
                        names = [names]
                    st.markdown("**📸 صور الواجب:**")
                    img_cols = st.columns(min(len(imgs), 3) if imgs else 1)
                    for idx_i, b64 in enumerate(imgs):
                        try:
                            img_bytes = base64.b64decode(b64)
                            caption = f"صورة {idx_i + 1}"
                            if idx_i < len(names) and names[idx_i]:
                                caption = names[idx_i]
                            with img_cols[idx_i % len(img_cols)]:
                                st.image(img_bytes, caption=caption, use_container_width=True)
                        except Exception:
                            st.warning(f"تعذر عرض الصورة رقم {idx_i + 1}.")
                except Exception:
                    st.error("تعذر قراءة بيانات الصور.")

            # نموذج المراجعة
            st.markdown("#### 🧑‍🏫 المراجعة")
            grade_init = 0.0
            try:
                grade_init = float(sub.get("grade", 0) or 0)
            except (TypeError, ValueError):
                grade_init = 0.0
            decision_init = 0
            if cur_status == "rejected":
                decision_init = 1

            with st.form(key=f"review_{sub_id}"):
                r1, r2 = st.columns(2)
                with r1:
                    new_grade = st.number_input(
                        "🎯 الدرجة",
                        min_value=0.0,
                        max_value=total_marks if total_marks > 0 else 100.0,
                        value=grade_init,
                        step=0.5,
                        key=f"g_{sub_id}"
                    )
                with r2:
                    decision = st.radio(
                        "القرار",
                        ["accepted", "rejected"],
                        format_func=lambda x: "✅ قبول" if x == "accepted" else "❌ رفض",
                        key=f"d_{sub_id}",
                        index=decision_init
                    )
                new_feedback = st.text_area("📌 ملاحظات المراجعة", value=sub.get("feedback", ""), key=f"f_{sub_id}")
                save_btn = st.form_submit_button("💾 حفظ المراجعة وإرسال الإشعار")

            if save_btn:
                try:
                    updates = {
                        "status": decision,
                        "grade": str(new_grade),
                        "feedback": new_feedback,
                        "reviewed_by": user.get("full_name", user.get("username", "")),
                        "reviewed_at": get_cairo_now().isoformat()
                    }
                    db.update_homework_submission(sub_id, updates)

                    # إشعار للطالبة
                    if decision == "accepted":
                        n_title = "✅ تم قبول واجبك"
                        n_msg = f"تم قبول واجب: {homework.get('title', '')}"
                        if total_marks > 0:
                            n_msg += f" | الدرجة: {new_grade:g} من {total_marks:g}"
                    else:
                        n_title = "❌ تم رفض واجبك"
                        n_msg = f"تم رفض واجب: {homework.get('title', '')}"
                    if new_feedback:
                        n_msg += f" | ملاحظات: {new_feedback}"
                    db.add_notification({
                        "notification_id": str(uuid.uuid4()),
                        "user_id": str(sub.get("student_id", "")),
                        "title": n_title,
                        "message": n_msg,
                        "notification_type": "homework_review",
                        "is_read": "False",
                        "created_at": get_cairo_now().isoformat()
                    })
                    st.success(f"✅ تم حفظ المراجعة وإرسال إشعار للطالبة {sub.get('student_name', '')}.")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ فشل حفظ المراجعة: {e}")

    # ===== ملخص المراجعة =====
    st.markdown("---")
    st.markdown("### 📊 ملخص المراجعة")
    acc = hw_subs[hw_subs["status"] == "accepted"]
    rej = hw_subs[hw_subs["status"] == "rejected"]
    pend = hw_subs[hw_subs["status"] == "pending"]
    grade_series = pd.to_numeric(hw_subs.get("grade", pd.Series(dtype=float)), errors="coerce").dropna()
    avg_grade = grade_series.mean() if not grade_series.empty else 0.0
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("✅ مقبول", len(acc))
    m2.metric("❌ مرفوض", len(rej))
    m3.metric("⏳ بانتظار المراجعة", len(pend))
    m4.metric("📈 متوسط الدرجات", f"{avg_grade:.1f}" if not grade_series.empty else "—")


# =============================================================================
# Notifications Panel - لوحة الإشعارات
# =============================================================================
def get_notification_icon(notif_type):
    """إرجاع أيقونة مناسبة لنوع الإشعار."""
    icons = {
        "exam_open": "📝",
        "exam_result": "📊",
        "homework": "📚",
        "homework_review": "📝",
        "general": "🔔",
        "event": "📅",
        "attendance": "📋",
        "system": "⚙️"
    }
    return icons.get(notif_type, "🔔")


def get_notification_color(notif_type):
    """إرجاع لون مناسب لنوع الإشعار."""
    colors = {
        "exam_open": "#2563eb",
        "exam_result": "#059669",
        "homework": "#d97706",
        "homework_review": "#7c3aed",
        "general": "#64748b",
        "event": "#dc2626",
        "attendance": "#0891b2",
        "system": "#475569"
    }
    return colors.get(notif_type, "#64748b")


def get_unread_notification_count(db, user_id):
    """حساب عدد الإشعارات غير المقروءة للمستخدم."""
    try:
        notifications = db.get_notifications(user_id)
        if notifications.empty or "is_read" not in notifications.columns:
            return 0
        unread = notifications[notifications["is_read"].astype(str).str.strip().str.lower() != "true"]
        return len(unread)
    except Exception:
        return 0


def show_notifications_panel(db):
    """
    لوحة الإشعارات الكاملة:
    - عرض جميع الإشعارات مع شارة العدد غير المقروء
    - قراءة الإشعارات (تحديد كـ مقروء)
    - إنشاء إشعار جديد (للمدرسين والمسؤولين)
    - إشعار فتح الامتحان
    - إشعار النتيجة
    - إشعار الواجب
    """
    st.markdown(hero_header("الإشعارات", "🔔 عرض وإدارة جميع الإشعارات"), unsafe_allow_html=True)

    user = st.session_state.get("user", {})
    user_id = user.get("user_id", "")
    role = user.get("role", "")

    # ===== جلب الإشعارات =====
    notifications = db.get_notifications(user_id)
    if notifications.empty:
        notifications = pd.DataFrame(columns=db.NOTIFICATION_COLUMNS)

    # ===== الإحصائيات =====
    total_count = len(notifications)
    unread_count = get_unread_notification_count(db, user_id)
    read_count = total_count - unread_count

    col1, col2, col3 = st.columns(3)
    col1.metric("🔔 إجمالي الإشعارات", total_count)
    col2.metric("📬 غير مقروء", unread_count)
    col3.metric("✅ مقروء", read_count)

    st.markdown("---")

    # ===== تبويبات: الإشعارات / إنشاء إشعار =====
    tab1, tab2 = st.tabs(["📥 الإشعارات", "➕ إنشاء إشعار"])

    # ===== تبويب عرض الإشعارات =====
    with tab1:
        # ===== فلاتر =====
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            type_filter = st.selectbox(
                "نوع الإشعار",
                ["الكل", "فتح الامتحان", "النتيجة", "الواجب", "عام", "فعالية", "حضور", "نظام"]
            )
        with col_f2:
            read_filter = st.selectbox("الحالة", ["الكل", "غير مقروء", "مقروء"])

        # ===== تصفية الإشعارات =====
        filtered = notifications.copy()
        if not filtered.empty:
            # تصفية حسب النوع
            type_map = {
                "فتح الامتحان": "exam_open",
                "النتيجة": "exam_result",
                "الواجب": "homework",
                "عام": "general",
                "فعالية": "event",
                "حضور": "attendance",
                "نظام": "system"
            }
            if type_filter != "الكل" and "notification_type" in filtered.columns:
                filtered = filtered[filtered["notification_type"] == type_map.get(type_filter, type_filter)]

            # تصفية حسب الحالة
            if read_filter == "غير مقروء" and "is_read" in filtered.columns:
                filtered = filtered[filtered["is_read"].astype(str).str.strip().str.lower() != "true"]
            elif read_filter == "مقروء" and "is_read" in filtered.columns:
                filtered = filtered[filtered["is_read"].astype(str).str.strip().str.lower() == "true"]

        if filtered.empty:
            st.markdown(empty_state("لا توجد إشعارات مطابقة.", "🔕"), unsafe_allow_html=True)
        else:
            # ===== زر قراءة الكل =====
            if unread_count > 0:
                if st.button("✅ تحديد الكل كمقروء", use_container_width=True, key="mark_all_read_btn"):
                    for _, notif in filtered.iterrows():
                        nid = notif.get("notification_id", "")
                        if nid and str(notif.get("is_read", "False")).strip().lower() != "true":
                            db.mark_notification_read(nid)
                    st.success("✅ تم تحديد جميع الإشعارات كمقروءة")
                    time.sleep(1)
                    st.rerun()
                st.markdown("---")

            # ===== عرض الإشعارات =====
            # ترتيب من الأحدث إلى الأقدم
            if "created_at" in filtered.columns:
                filtered = filtered.sort_values("created_at", ascending=False)

            for _, notif_row in filtered.iterrows():
                notif = notif_row.to_dict()
                nid = notif.get("notification_id", "")
                ntype = notif.get("notification_type", "general")
                ntitle = notif.get("title", "إشعار")
                nmsg = notif.get("message", "")
                nread = str(notif.get("is_read", "False")).strip().lower() == "true"
                ncreated = notif.get("created_at", "")

                # تنسيق الوقت
                time_display = "غير متاح"
                try:
                    dt = pd.to_datetime(ncreated)
                    if pd.isna(dt):
                        time_display = "غير متاح"
                    else:
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=CAIRO_TZ)
                        time_display = format_cairo_time(dt)
                except Exception:
                    time_display = "غير متاح"

                icon = get_notification_icon(ntype)
                color = get_notification_color(ntype)

                # بطاقة الإشعار
                bg_color = "#f8fafc" if nread else "#dbeafe"
                border_color = "#e2e8f0" if nread else color

                st.markdown(f"""
                <div style="background:{bg_color}; border:1px solid {border_color}; border-right:4px solid {color};
                            border-radius:12px; padding:1rem 1.2rem; margin-bottom:0.8rem;">
                    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:0.3rem;">
                        <div style="font-weight:700; color:#0f172a; font-size:0.95rem;">
                            {icon} {ntitle}
                            {'' if nread else '<span style="background:#dc2626; color:white; font-size:0.65rem; padding:0.1rem 0.5rem; border-radius:9999px; margin-right:0.5rem;">جديد</span>'}
                        </div>
                        <div style="font-size:0.7rem; color:#64748b;">🕒 {time_display}</div>
                    </div>
                    <div style="color:#475569; font-size:0.85rem; margin-top:0.3rem;">{nmsg}</div>
                </div>
                """, unsafe_allow_html=True)

                # أزرار الإجراءات
                act_cols = st.columns([1, 1, 3])
                if not nread:
                    with act_cols[0]:
                        if st.button("✅ قراءة", key=f"read_{nid}", use_container_width=True):
                            db.mark_notification_read(nid)
                            st.rerun()
                else:
                    with act_cols[0]:
                        st.markdown("<div style='text-align:center; color:#059669; font-size:0.8rem; padding:0.3rem;'>✅ مقروء</div>", unsafe_allow_html=True)

                # زر حذف الإشعار
                with act_cols[1]:
                    if st.button("🗑️ حذف", key=f"del_notif_{nid}", use_container_width=True):
                        try:
                            all_notifs = db.get_notifications()
                            if not all_notifs.empty and "notification_id" in all_notifs.columns:
                                all_notifs = all_notifs[all_notifs["notification_id"] != nid]
                                db._df_to_sheet("Notifications", all_notifs, db.NOTIFICATION_COLUMNS)
                                st.success("✅ تم حذف الإشعار")
                                time.sleep(1)
                                st.rerun()
                        except Exception as e:
                            st.error(f"❌ فشل حذف الإشعار: {e}")

                st.markdown("---")

    # ===== تبويب إنشاء إشعار =====
    with tab2:
        # المدرسين والمسؤولين فقط يمكنهم إنشاء إشعارات
        if role in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
            st.markdown("### ➕ إنشاء إشعار جديد")

            # جلب المستخدمين والطالبات
            users_df = db.get_users()
            students_df = db.get_students()

            # بناء قائمة المستلمين
            recipients = {}
            if not users_df.empty and "user_id" in users_df.columns:
                for _, u in users_df.iterrows():
                    u_role = u.get("role", "")
                    if u_role in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
                        recipients[str(u.get("user_id", ""))] = f"👤 {u.get('full_name', '')} ({u_role})"
            if not students_df.empty and "student_id" in students_df.columns:
                for _, s in students_df.iterrows():
                    recipients[str(s.get("student_id", ""))] = f"👩‍🎓 {s.get('full_name', '')}"

            if not recipients:
                st.info("لا يوجد مستخدمون أو طالبات لإرسال الإشعارات إليهم.")
            else:
                with st.form("create_notification_form"):
                    col1, col2 = st.columns(2)
                    with col1:
                        notif_type = st.selectbox(
                            "نوع الإشعار",
                            ["عام", "فتح الامتحان", "النتيجة", "الواجب", "فعالية", "حضور", "نظام"]
                        )
                    with col2:
                        recipient_type = st.selectbox("المستلم", ["مستخدم محدد", "جميع الطالبات", "جميع المستخدمين"])

                    selected_recipient = ""
                    if recipient_type == "مستخدم محدد":
                        selected_recipient = st.selectbox(
                            "اختر المستلم",
                            list(recipients.keys()),
                            format_func=lambda x: recipients.get(x, x)
                        )

                    notif_title = st.text_input("عنوان الإشعار*", placeholder="أدخل عنوان الإشعار")
                    notif_message = st.text_area("نص الإشعار*", placeholder="أدخل نص الإشعار", height=120)

                    submitted = st.form_submit_button("📨 إرسال الإشعار", use_container_width=True)

                    if submitted:
                        if not notif_title or not notif_message:
                            st.error("⚠️ عنوان الإشعار ونصه مطلوبان")
                        else:
                            type_map = {
                                "عام": "general",
                                "فتح الامتحان": "exam_open",
                                "النتيجة": "exam_result",
                                "الواجب": "homework",
                                "فعالية": "event",
                                "حضور": "attendance",
                                "نظام": "system"
                            }
                            notif_type_eng = type_map.get(notif_type, "general")

                            # تحديد المستلمين
                            target_ids = []
                            if recipient_type == "مستخدم محدد":
                                target_ids = [selected_recipient]
                            elif recipient_type == "جميع الطالبات":
                                if not students_df.empty and "student_id" in students_df.columns:
                                    target_ids = students_df["student_id"].astype(str).tolist()
                            else:  # جميع المستخدمين
                                if not users_df.empty and "user_id" in users_df.columns:
                                    target_ids = users_df["user_id"].astype(str).tolist()

                            # إرسال الإشعارات
                            sent_count = 0
                            for target_id in target_ids:
                                if not target_id:
                                    continue
                                db.add_notification({
                                    "notification_id": str(uuid.uuid4()),
                                    "user_id": str(target_id),
                                    "title": notif_title.strip(),
                                    "message": notif_message.strip(),
                                    "notification_type": notif_type_eng,
                                    "is_read": "False",
                                    "created_at": get_cairo_now().isoformat()
                                })
                                sent_count += 1

                            db.add_log(user_id, "إنشاء إشعار", f"تم إرسال {sent_count} إشعار: {notif_title}")
                            st.success(f"✅ تم إرسال {sent_count} إشعار بنجاح")
                            time.sleep(1)
                            st.rerun()
        else:
            st.info("👁️ يمكنك فقط عرض الإشعارات. المدرسون والمسؤولون يمكنهم إنشاء إشعارات جديدة.")


# =============================================================================
# Main App
# =============================================================================
def main():
    inject_css()
    init_session()
    init_data_cache()
    if 'db_instance' not in st.session_state:
        try:
            creds = get_credentials()
            st.session_state.db_instance = Database(creds, get_spreadsheet_id())
        except Exception as e:
            st.error(f"❌ خطأ في الاتصال: {e}")
            st.stop()
    db = st.session_state.db_instance
    jwt_secret = get_jwt_secret()
    if st.session_state.get("authenticated"):
        try:
            migrated = db.migrate_single_supervisors()
            if migrated and migrated > 0:
                st.caption(f"تم نقل {migrated} تعيين من مسؤول المرحلة القديم إلى نظام المشرفين المتعددين.")
        except Exception:
            pass
    st.markdown('<div class="help-float-container"></div>', unsafe_allow_html=True)
    if st.session_state.get("student_logged_in", False):
        show_student_dashboard(db)
    else:
        if not st.session_state.authenticated:
            show_login_page(db, jwt_secret)
        else:
            token_data = verify_token(st.session_state.token, jwt_secret)
            if not token_data:
                st.error("⏰ انتهت صلاحية الجلسة.")
                st.session_state.clear()
                time.sleep(2)
                st.rerun()
                return
            if not st.session_state.get("data_validated"):
                errors = validate_data_integrity(db)
                st.session_state.data_errors = errors
                st.session_state.data_validated = True
            render_admin_top_bar(show_menu_button=not st.session_state.show_sidebar)
            if not st.session_state.show_sidebar:
                st.markdown("""<style>section[data-testid="stSidebar"] { transform: translateX(100%) !important; }</style>""", unsafe_allow_html=True)
            else:
                st.markdown("""<style>section[data-testid="stSidebar"] { transform: translateX(0) !important; }</style>""", unsafe_allow_html=True)
                choice = show_sidebar_navigation(db)
            if not st.session_state.show_sidebar:
                choice = normalize_admin_menu_choice(st.session_state.get("menu_choice", "🏠 لوحة التحكم"))
                if choice != st.session_state.get("menu_choice"):
                    st.session_state.menu_choice = choice
                role = st.session_state.user.get("role", "")
                menu_items = get_role_menu(role)
                if choice not in menu_items:
                    choice = menu_items[0] if menu_items else "🏠 لوحة التحكم"
                    st.session_state.menu_choice = choice
            st.markdown("<div class='content-area'>", unsafe_allow_html=True)
            if st.session_state.get("profile_user_id"):
                profile_id = st.session_state.profile_user_id
                # Check if this ID exists in Students sheet first
                students_df = db.get_students()
                if not students_df.empty and "student_id" in students_df.columns:
                    student_match = students_df[students_df["student_id"] == profile_id]
                    if not student_match.empty:
                        show_student_profile(db, profile_id)
                    else:
                        show_user_profile(db, profile_id)
                else:
                    show_user_profile(db, profile_id)
            elif choice == "🏠 لوحة التحكم":
                show_dashboard(db)
            elif choice == "🔔 الإشعارات":
                show_notifications_panel(db)
            elif choice == "🏫 إدارة المراحل الدراسية":
                if st.session_state.user.get("role") in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
                    show_stages_page(db)
                else:
                    st.error("🚫 غير مصرح")
            elif choice == "📚 إدارة الفصول":
                if st.session_state.user.get("role") in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
                    show_sections_page(db)
                else:
                    st.error("🚫 غير مصرح")
            elif choice == "👥 إدارة الأعضاء":
                if st.session_state.user.get("role") in ["System Admin", "Father Account", "Service Manager", "Teacher"]:
                    show_members_cards_page(db)
                else:
                    st.error("🚫 غير مصرح")
            elif choice == "📋 الحضور":
                show_attendance(db)
            elif choice == "💬 الافتقاد":
                show_followup(db)
            elif choice == ADMIN_ASSESSMENTS_PAGE or choice == LEGACY_ADMIN_ASSESSMENTS_PAGE:
                show_unified_assessments_admin(db)
            elif choice == "📊 التقارير والإحصائيات":
                show_reports_page(db)
            elif choice == "📅 إدارة الفعاليات":
                show_events_page(db)
            elif choice == "📜 سجل العمليات":
                if st.session_state.user.get("role") == "System Admin":
                    show_logs(db)
                else:
                    st.error("🚫 غير مصرح")
            elif choice == "📷 ماسح QR":
                show_qr_scanner_page(db)
            elif choice == "🔒 تغيير كلمة المرور":
                change_password(db)
            st.markdown("</div>", unsafe_allow_html=True)
    if st.session_state.get("open_help_dialog"):
        show_help_dialog()
        st.session_state.open_help_dialog = False


if __name__ == "__main__":
    main()
